#!/usr/bin/env python3
"""
Classify DO↔Lightsail pull-list rows into confident vs not-confident for
Handpan-style carousel / variant gallery sync.

Reads:
  data/compare/do-ls-pull-list.xlsx (Pull List + Excluded)
  data/compare/do_products.csv
  data/compare/do_variants.csv
  backend/prisma/wc-products.csv (ACF carousel meta)

Writes:
  data/compare/do-ls-media-confidence.xlsx
    - Confident
    - Not Confident
    - Summary

Usage:
  python3 backend/scripts/generate-do-ls-media-confidence.py
"""
from __future__ import annotations

import csv
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
PULL_LIST = ROOT / "data/compare/do-ls-pull-list.xlsx"
DO_PRODUCTS = ROOT / "data/compare/do_products.csv"
DO_VARIANTS = ROOT / "data/compare/do_variants.csv"
DO_CAROUSEL = ROOT / "data/compare/do_carousel_by_product.csv"
WC_CSV = ROOT / "backend/prisma/wc-products.csv"
OUT = ROOT / "data/compare/do-ls-media-confidence.xlsx"

HEADERS = [
    "Lightsail Product name",
    "DO DB product name",
    "Lightsail variant name",
    "DO variant name",
    "Lightsail SKU",
    "DO SKU",
    "Note on mismatch",
    "Match source",
    "Confidence reason",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN = PatternFill("solid", fgColor="E2EFDA")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
RED = PatternFill("solid", fgColor="FCE4D6")


@dataclass
class PullRow:
    ls_product: str
    do_product: str
    ls_variant: str
    do_variant: str
    ls_sku: str
    do_sku: str
    note: str
    source: str


@dataclass
class ProductMedia:
    woo_id: int
    slug: str
    product_type: str
    has_carousel: bool
    carousel_slots: int
    has_gallery: bool
    has_thumb: bool
    image_count: int


def norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def norm_sku(s: str) -> str:
    return (s or "").strip().upper()


def norm_variant(s: str) -> str:
    s = norm_text(s)
    return re.sub(r"[\s|/·,–—\-]+", " ", s).strip()


def parse_do_variant_name(attrs: str, title: str, product_name: str) -> str:
    if attrs:
        parts = []
        for seg in attrs.split(";"):
            if "=" in seg:
                parts.append(seg.split("=", 1)[1].strip())
        if parts:
            return " / ".join(parts)
    if title and " - " in title:
        tail = title.split(" - ", 1)[1].strip()
        if tail and norm_text(tail) != norm_text(product_name):
            return tail
    return "Standard"


def load_do_product_by_name() -> dict[str, dict]:
    out: dict[str, dict] = {}
    with DO_PRODUCTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            name = r.get("name") or ""
            out[norm_text(name)] = r
    return out


def load_do_variants_by_product() -> dict[str, list[dict]]:
    products = {}
    with DO_PRODUCTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() == "publish":
                products[r["id"]] = r.get("name") or ""

    by_name: dict[str, list[dict]] = defaultdict(list)
    with DO_VARIANTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            pname = products.get(r.get("parent_id") or "", "")
            by_name[norm_text(pname)].append(r)
    return by_name


def load_do_carousel() -> dict[int, int]:
    """Woo ID → carousel slot count (from live DO dump CSV)."""
    out: dict[int, int] = {}
    if not DO_CAROUSEL.exists():
        return out
    with DO_CAROUSEL.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            try:
                out[int(r["id"])] = int(r.get("carousel_slots") or 0)
            except (ValueError, KeyError):
                continue
    return out


def load_wc_media() -> dict[int, ProductMedia]:
    """Map Woo product ID → carousel / gallery info from wc-products.csv."""
    out: dict[int, ProductMedia] = {}
    if not WC_CSV.exists():
        return out

    with WC_CSV.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ptype = (row.get("Type") or "").strip().lower()
            if ptype not in ("simple", "variable"):
                continue
            parent = (row.get("Parent") or "").strip()
            if parent:
                continue
            try:
                woo_id = int(row.get("ID") or "")
            except ValueError:
                continue

            carousel_slots = 0
            has_carousel = False
            for key, val in row.items():
                if not key or not val:
                    continue
                if "product_gallery_carousel_image_linked_with_" not in key:
                    continue
                if key.endswith("_image") and not key.startswith("Meta: _"):
                    v = str(val).strip()
                    if v.isdigit():
                        carousel_slots += 1
                        has_carousel = True
                if "_iframe" in key and "youtube" in str(val).lower():
                    carousel_slots += 1
                    has_carousel = True

            images_raw = (row.get("Images") or "").strip()
            image_urls = [u.strip() for u in images_raw.split(",") if u.strip().startswith("http")]
            thumb = ""
            for key, val in row.items():
                if key.endswith("_thumbnail_id") or key == "Meta: _thumbnail_id":
                    if str(val).strip().isdigit():
                        thumb = str(val).strip()

            out[woo_id] = ProductMedia(
                woo_id=woo_id,
                slug=(row.get("Name") or "").strip(),
                product_type=ptype,
                has_carousel=has_carousel,
                carousel_slots=carousel_slots,
                has_gallery=bool(image_urls),
                has_thumb=bool(thumb),
                image_count=len(image_urls),
            )
    return out


def read_pull_rows() -> tuple[list[PullRow], list[list]]:
    wb = load_workbook(PULL_LIST, read_only=True, data_only=True)
    pull: list[PullRow] = []
    ws = wb["Pull List"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        pull.append(
            PullRow(
                ls_product=str(row[0] or ""),
                do_product=str(row[1] or ""),
                ls_variant=str(row[2] or "—"),
                do_variant=str(row[3] or "—"),
                ls_sku=str(row[4] or ""),
                do_sku=str(row[5] or ""),
                note=str(row[6] or ""),
                source=str(row[7] or ""),
            )
        )

    excluded: list[list] = []
    if "Excluded" in wb.sheetnames:
        ex = wb["Excluded"]
        for row in ex.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                excluded.append([str(c or "") for c in row])
    wb.close()
    return pull, excluded


def sku_ok(ls_sku: str, do_sku: str) -> tuple[bool, str]:
    if not ls_sku and not do_sku:
        return True, ""
    if not ls_sku or not do_sku:
        return True, "SKU missing on one side"
    if norm_sku(ls_sku) == norm_sku(do_sku):
        return True, ""
    return False, f"SKU differs (LS {ls_sku} vs DO {do_sku})"


def variant_match_ok(note: str) -> tuple[bool, str]:
    n = note.lower()
    if "only" in n or "not found" in n or "extra" in n:
        return False, note
    if "fuzzy" in n and "100" not in n and "exact" not in n:
        m = re.search(r"fuzzy match \((\d+)", n)
        if m and int(m.group(1)) < 85:
            return False, note
    return True, ""


def classify_product(
    rows: list[PullRow],
    do_product: dict | None,
    do_variants: list[dict],
    wc: ProductMedia | None,
    carousel_slots: int,
) -> tuple[bool, str]:
    """Return (confident, reason)."""

    # 1) Full variant parity
    unmatched = [r for r in rows if r.ls_variant == "—" or r.do_variant == "—"]
    if unmatched:
        return False, f"Partial variant match ({len(unmatched)} row(s) missing LS or DO variant)"

    ls_count = len([r for r in rows if r.ls_variant != "—"])
    do_count = len(do_variants) if do_variants else (1 if do_product and (do_product.get("product_type") or "").lower() == "simple" else 0)
    if do_variants and ls_count != do_count:
        return False, f"Variant count mismatch (LS {ls_count} vs DO {do_count})"

    # 2) Per-row SKU + name checks
    for r in rows:
        ok, msg = variant_match_ok(r.note)
        if not ok:
            return False, msg
        sku_good, sku_msg = sku_ok(r.ls_sku, r.do_sku)
        if not sku_good:
            return False, sku_msg

    # 3) DO media source
    if not do_product:
        return False, "DO product not found in do_products.csv"

    woo_id = int(do_product["id"])
    ptype = (do_product.get("product_type") or "").lower()

    if carousel_slots > 0:
        return True, (
            f"Handpan-like: ACF carousel on DO ({carousel_slots} slot(s)); "
            "full LS↔DO variant parity — carousel sync path"
        )

    if ptype == "simple":
        has_media = (
            (wc and (wc.has_gallery or wc.has_thumb))
            or bool((do_product.get("thumb_id") or "").strip())
            or bool((do_product.get("gallery") or "").strip())
        )
        if has_media:
            return True, "Simple product; shared gallery / thumb on DO"
        return False, "Simple product but no DO gallery or thumb"

    # Variable without carousel
    thumbs = sum(1 for v in do_variants if (v.get("thumb_id") or "").strip())
    if thumbs == len(do_variants) and thumbs > 0:
        return True, f"All {thumbs} DO variants have featured image (no ACF carousel)"

    if thumbs > 0:
        return False, f"Only {thumbs}/{len(do_variants)} DO variants have featured image; no carousel"

    gallery = (do_product.get("gallery") or "").strip()
    if gallery or (wc and wc.has_gallery):
        return False, "Shared Woo gallery only — no per-variant carousel/thumbs (risk wrong thumbs like Handpan had)"

    return False, "No DO carousel, gallery, or variant thumbs"


def write_sheet(ws, headers: list[str], rows: list[list], fill_col: int | None = None, fill: PatternFill | None = None):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
    for ri, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=col, value=val)
            if fill_col and col == fill_col and fill:
                cell.fill = fill
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(52, max(14, len(h) + 2))


def main() -> None:
    for p in (PULL_LIST, DO_PRODUCTS, DO_VARIANTS):
        if not p.exists():
            raise SystemExit(f"Missing {p}")

    pull_rows, excluded = read_pull_rows()
    do_by_name = load_do_product_by_name()
    do_vars_by_name = load_do_variants_by_product()
    wc_media = load_wc_media()
    do_carousel = load_do_carousel()

    by_product: dict[tuple[str, str], list[PullRow]] = defaultdict(list)
    for r in pull_rows:
        by_product[(r.ls_product, r.do_product)].append(r)

    confident_rows: list[list] = []
    not_confident_rows: list[list] = []
    confident_products: set[tuple[str, str]] = set()
    not_confident_products: set[tuple[str, str]] = set()
    handpan_like_products: set[tuple[str, str]] = set()
    thumb_only_products: set[tuple[str, str]] = set()

    for (ls_name, do_name), rows in sorted(by_product.items(), key=lambda x: x[0][0].lower()):
        do_prod = do_by_name.get(norm_text(do_name))
        do_vars = do_vars_by_name.get(norm_text(do_name), [])
        wc = wc_media.get(int(do_prod["id"])) if do_prod and do_prod.get("id") else None
        carousel_slots = do_carousel.get(int(do_prod["id"]), 0) if do_prod and do_prod.get("id") else 0

        ok, reason = classify_product(rows, do_prod, do_vars, wc, carousel_slots)
        bucket = confident_rows if ok else not_confident_rows
        pset = confident_products if ok else not_confident_products
        pset.add((ls_name, do_name))
        if ok:
            if carousel_slots > 0:
                handpan_like_products.add((ls_name, do_name))
            else:
                thumb_only_products.add((ls_name, do_name))

        for r in sorted(rows, key=lambda x: x.ls_variant.lower()):
            bucket.append(
                [
                    r.ls_product,
                    r.do_product,
                    r.ls_variant,
                    r.do_variant,
                    r.ls_sku,
                    r.do_sku,
                    r.note,
                    r.source,
                    reason,
                ]
            )

    # Excluded → not confident
    for ex in excluded:
        # ex may have 9 cols with exclude reason last
        base = ex[:8] if len(ex) >= 8 else ex + [""] * (8 - len(ex))
        reason = ex[8] if len(ex) > 8 else "Excluded from pull list"
        not_confident_rows.append(
            [
                base[0] if len(base) > 0 else "",
                base[1] if len(base) > 1 else "",
                base[2] if len(base) > 2 else "—",
                base[3] if len(base) > 3 else "—",
                base[4] if len(base) > 4 else "",
                base[5] if len(base) > 5 else "",
                base[6] if len(base) > 6 else "",
                base[7] if len(base) > 7 else "",
                reason,
            ]
        )
        ls_p = base[0] if base else ""
        do_p = base[1] if len(base) > 1 else ""
        if ls_p or do_p:
            not_confident_products.add((ls_p, do_p))

    wb = Workbook()
    conf = wb.active
    conf.title = "Confident"
    write_sheet(conf, HEADERS, confident_rows, fill_col=len(HEADERS), fill=GREEN)

    nc = wb.create_sheet("Not Confident")
    write_sheet(nc, HEADERS, not_confident_rows, fill_col=len(HEADERS), fill=YELLOW)

    sm = wb.create_sheet("Summary")
    conf_var_rows = len(confident_rows)
    nc_var_rows = len(not_confident_rows)
    handpan_rows = sum(
        1 for r in confident_rows if r[8] and str(r[8]).startswith("Handpan-like:")
    )
    sm_data = [
        ["Generated at (UTC)", datetime.now(timezone.utc).isoformat()],
        ["Source pull list", str(PULL_LIST.relative_to(ROOT))],
        ["DO carousel map", str(DO_CAROUSEL.relative_to(ROOT))],
        ["", ""],
        ["CONFIDENT — total products", len(confident_products)],
        ["CONFIDENT — total variant rows", conf_var_rows],
        ["  ↳ Handpan-like (ACF carousel on DO)", len(handpan_like_products)],
        ["  ↳ Handpan-like variant rows", handpan_rows],
        ["  ↳ Variant-thumb / simple gallery only", len(thumb_only_products)],
        ["  ↳ Variant-thumb variant rows", conf_var_rows - handpan_rows],
        ["", ""],
        ["NOT CONFIDENT — total products", len(not_confident_products)],
        ["NOT CONFIDENT — total variant rows", nc_var_rows],
        ["", ""],
        ["Confident criteria", "Every LS variant matched to DO + no SKU conflict + DO media exists"],
        ["Handpan-like", "DO has ACF product_gallery_carousel_* slots (live DO dump)"],
        ["Variant-thumb path", "No carousel but every DO variant has _thumbnail_id, or simple gallery"],
        ["Not confident", "Partial match, LS/DO-only, excluded, missing DO media, SKU mismatch"],
    ]
    write_sheet(sm, ["Metric", "Value"], sm_data)

    wb.save(OUT)

    print(f"Confident products: {len(confident_products)} ({conf_var_rows} variant rows)")
    print(f"  Handpan-like (carousel): {len(handpan_like_products)} ({handpan_rows} variant rows)")
    print(f"  Variant-thumb/simple: {len(thumb_only_products)} ({conf_var_rows - handpan_rows} variant rows)")
    print(f"Not confident products: {len(not_confident_products)} ({nc_var_rows} variant rows)")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
