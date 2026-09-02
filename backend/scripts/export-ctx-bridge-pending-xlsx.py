#!/usr/bin/env python3
"""Export CTX bridge pending rows to Excel for manual review."""
from __future__ import annotations

import json
import re
import csv
from pathlib import Path
from urllib.parse import urlparse, parse_qs
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise SystemExit("pip install openpyxl") from e

REPO = Path(__file__).resolve().parents[2]
AUDIT = REPO / "docs/audit/google-merchant-native-compatibility"
CTX_XML = AUDIT / "ctx_india_authoritative.xml"
DO_MAP = AUDIT / "do_lightsail_sku_map.json"
APPLY_SUMMARY = AUDIT / "ctx_bridge_apply_summary.json"
MAPPING_TSV = REPO / "docs/audit/merchant_woo_sarveda_mapping.tsv"
OUT_XLSX = AUDIT / "ctx_bridge_pending_review.xlsx"


def parse_ctx_xml(path: Path) -> dict[str, dict]:
    xml = path.read_text(encoding="utf-8", errors="replace")
    items = re.findall(r"<item>([\s\S]*?)</item>", xml)
    out: dict[str, dict] = {}

    def tag(block: str, name: str, ns: bool = False) -> str:
        n = f"g:{name}" if ns else name
        m = re.search(rf"<{re.escape(n)}>([\s\S]*?)</{re.escape(n)}>", block, re.I)
        return (m.group(1) if m else "").replace("<![CDATA[", "").replace("]]>", "").strip()

    for block in items:
        gid = tag(block, "id", True)
        if not gid or not gid.isdigit():
            continue
        link = tag(block, "link")
        attrs: list[str] = []
        try:
            q = parse_qs(urlparse(link).query)
            for k, vals in sorted(q.items()):
                if k.startswith("attribute_") and vals:
                    attrs.append(f"{k}={vals[0]}")
        except Exception:
            pass
        out[gid] = {
            "ctx_g_id": gid,
            "ctx_item_group_id": tag(block, "item_group_id", True),
            "ctx_title": tag(block, "title"),
            "ctx_product_type": tag(block, "product_type", True),
            "ctx_link": link,
            "ctx_legacy_attributes": "; ".join(attrs),
            "ctx_price": tag(block, "price", True),
            "ctx_sale_price": tag(block, "sale_price", True),
            "ctx_availability": tag(block, "availability", True),
            "ctx_brand": tag(block, "brand", True),
            "ctx_image_link": tag(block, "image_link", True),
        }
    return out


def read_mapping() -> dict[str, dict]:
    out: dict[str, dict] = {}
    if not MAPPING_TSV.exists():
        return out
    with MAPPING_TSV.open(encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            oid = (row.get("woo_offer_id") or "").strip()
            if oid:
                out[oid] = row
    return out


def load_apply_skip_reasons() -> dict[str, str]:
    if not APPLY_SUMMARY.exists():
        return {}
    data = json.loads(APPLY_SUMMARY.read_text())
    return {
        str(r["wooOfferId"]): r["reason"]
        for r in data.get("rows", [])
        if r.get("action") == "skip"
    }


def autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(60, max(10, max(len(str(c.value or "")) for c in col) + 2))
        ws.column_dimensions[letter].width = width


def write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    autosize(ws)


def main() -> None:
    ctx = parse_ctx_xml(CTX_XML)
    do_map = json.loads(DO_MAP.read_text()) if DO_MAP.exists() else {}
    mapping = read_mapping()
    apply_skip = load_apply_skip_reasons()

    ls_path = Path("/tmp/ls_offers_current.json")
    if not ls_path.exists():
        raise SystemExit("Run Lightsail export to /tmp/ls_offers_current.json first")
    ls = json.loads(ls_path.read_text())
    offers = ls["offers"]
    by_sku = {v["sku"].strip(): v for v in ls["variants"]}
    by_woo_var = {
        str(v["wooCommerceVariationId"]): v
        for v in ls["variants"]
        if v.get("wooCommerceVariationId")
    }
    variant_to_publish_offer: dict[str, str] = {}
    for o in offers:
        if o.get("classification") == "PUBLISH" and o.get("sarvedaVariantId"):
            variant_to_publish_offer[o["sarvedaVariantId"]] = str(o["wooOfferId"])

    headers = [
        "ctx_g_id",
        "ctx_item_group_id",
        "ctx_title",
        "ctx_product_type",
        "ctx_link",
        "ctx_legacy_attributes",
        "ctx_price",
        "ctx_sale_price",
        "ctx_availability",
        "classification",
        "do_lightsail_sku",
        "ls_variant_sku",
        "ls_product_slug",
        "ls_product_name",
        "ls_variant_status",
        "pending_reason",
        "audit_match_method",
        "audit_match_confidence",
        "audit_notes",
        "published_ctx_g_id_for_same_variant",
    ]

    no_bridge: list[list] = []
    variant_conflict: list[list] = []
    remaining_manual: list[list] = []
    remaining_exclude: list[list] = []

    for o in offers:
        gid = str(o["wooOfferId"])
        c = ctx.get(gid, {})
        m = mapping.get(gid, {})
        do_ls_sku = do_map.get(gid, "")
        variant = None
        if do_ls_sku:
            variant = by_sku.get(do_ls_sku.strip())
        elif o.get("sarvedaVariantId"):
            variant = next((v for v in ls["variants"] if v["id"] == o["sarvedaVariantId"]), None)

        cls = o.get("classification", "")
        if cls == "PUBLISH":
            continue

        if not do_ls_sku:
            pending = "NO_DO_LIGHTSAIL_SKU"
        elif apply_skip.get(gid) == "VARIANT_ALREADY_PUBLISHED":
            pending = "VARIANT_ALREADY_PUBLISHED"
        elif apply_skip.get(gid):
            pending = apply_skip[gid]
        elif cls == "MANUAL_REVIEW":
            pending = o.get("excludeReason") or "MANUAL_REVIEW"
        else:
            pending = o.get("excludeReason") or cls

        pub_other = ""
        if variant and variant["id"] in variant_to_publish_offer:
            pub_other = variant_to_publish_offer[variant["id"]]
            if pub_other == gid:
                pub_other = ""

        row = [
            gid,
            c.get("ctx_item_group_id", o.get("ctxItemGroupId", "")),
            c.get("ctx_title", o.get("ctxTitle", "")),
            c.get("ctx_product_type", o.get("ctxProductType", "")),
            c.get("ctx_link", o.get("ctxLegacyLink", "")),
            c.get("ctx_legacy_attributes", ""),
            c.get("ctx_price", ""),
            c.get("ctx_sale_price", ""),
            c.get("ctx_availability", ""),
            cls,
            do_ls_sku,
            variant["sku"] if variant else (m.get("sarveda_sku") or ""),
            (variant or {}).get("productRel", {}).get("slug") or m.get("sarveda_slug", ""),
            (variant or {}).get("productRel", {}).get("name", ""),
            variant.get("status", "") if variant else "",
            pending,
            m.get("match_method", ""),
            m.get("match_confidence", ""),
            m.get("notes", o.get("notes", "")),
            pub_other,
        ]

        if cls == "MANUAL_REVIEW":
            remaining_manual.append(row)
            if pending == "NO_DO_LIGHTSAIL_SKU":
                no_bridge.append(row)
            elif pending == "VARIANT_ALREADY_PUBLISHED":
                variant_conflict.append(row)
        elif cls == "INTENTIONALLY_EXCLUDE":
            remaining_exclude.append(row)

    wb = Workbook()
    wb.remove(wb.active)

    s1 = wb.create_sheet("no_bridge_85")
    write_sheet(s1, headers, no_bridge)

    s2 = wb.create_sheet("variant_conflict_31")
    write_sheet(s2, headers, variant_conflict)

    s3 = wb.create_sheet("remaining_manual_116")
    write_sheet(s3, headers, remaining_manual)

    s4 = wb.create_sheet("intentionally_exclude_11")
    write_sheet(s4, headers, remaining_exclude)

    s5 = wb.create_sheet("all_non_publish_127")
    all_non = no_bridge + variant_conflict + [
        r for r in remaining_manual if r not in no_bridge and r not in variant_conflict
    ] + remaining_exclude
    write_sheet(s5, headers, all_non)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(
        json.dumps(
            {
                "output": str(OUT_XLSX),
                "no_bridge_85": len(no_bridge),
                "variant_conflict_31": len(variant_conflict),
                "remaining_manual_116": len(remaining_manual),
                "intentionally_exclude_11": len(remaining_exclude),
                "all_non_publish": len(all_non),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
