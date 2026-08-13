#!/usr/bin/env python3
"""
Build enhanced DO→Lightsail media pull list (v2).

Rules (Lightsail = source of truth):
- Product name pairs from existing fuzzy + MATCHED ONES (unchanged)
- Variant match by name only (ignore DO SKU); LS SKU always kept
- LS-only variant → skip (no DO pull), unless DNA Tuning Fork fallback
- DO-only variant → ignore (removed on new site)
- Secondary fuzzy on variant names for partial-match products
- HOLD: Etched Chau Gongs (entire product skipped)

Writes:
  data/compare/do-ls-media-pull-v2.json
  data/compare/do-ls-media-pull-v2.xlsx

Usage:
  python3 backend/scripts/generate-do-ls-media-pull-v2.py
"""
from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

ROOT = Path(__file__).resolve().parents[2]
FUZZY_XLSX = ROOT / "data/compare/do-vs-lightsail-fuzzy.xlsx"
DO_PRODUCTS = ROOT / "data/compare/do_products.csv"
DO_VARIANTS = ROOT / "data/compare/do_variants.csv"
LS_EXPORT = ROOT / "data/compare/lightsail-catalog-export.json"
CONFIDENCE_XLSX = ROOT / "data/compare/do-ls-media-confidence.xlsx"
OUT_JSON = ROOT / "data/compare/do-ls-media-pull-v2.json"
OUT_XLSX = ROOT / "data/compare/do-ls-media-pull-v2.xlsx"

VARIANT_MIN = 80
HOLD_LS_PRODUCTS = {"etched chau gongs"}
DNA_LS_PRODUCT = "dna tuning fork"

HEADERS = [
    "Lightsail Product name",
    "DO DB product name",
    "Lightsail variant name",
    "DO variant name",
    "Lightsail SKU",
    "DO SKU",
    "Note on mismatch",
    "Action",
    "DO variation ID",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN = PatternFill("solid", fgColor="E2EFDA")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GRAY = PatternFill("solid", fgColor="EDEDED")


@dataclass
class DoVar:
    id: str
    name: str
    sku: str
    thumb_id: str
    attrs: dict[str, str]


@dataclass
class LsVar:
    sku: str
    name: str
    variant_id: str
    product_id: str


@dataclass
class ProductPair:
    ls_name: str
    do_name: str
    ls_slug: str
    woo_id: int | None
    source: str


@dataclass
class MatchRow:
    ls_product: str
    do_product: str
    ls_variant: str
    do_variant: str
    ls_sku: str
    do_sku: str
    note: str
    action: str
    do_variation_id: str = ""


def norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def norm_variant(s: str) -> str:
    s = norm_text(s)
    s = s.replace("inch", "in").replace("inches", "in")
    s = re.sub(r"\s*in\b", " in", s)
    s = re.sub(r"[\s|/·,–—\-]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def parse_attrs(raw: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for seg in (raw or "").split(";"):
        if "=" in seg:
            k, v = seg.split("=", 1)
            out[k.replace("attribute_", "").replace("pa_", "")] = v.strip()
    return out


def parse_do_variant_name(attrs: dict[str, str], title: str, product_name: str) -> str:
    if attrs:
        parts = list(attrs.values())
        if parts:
            return " / ".join(parts)
    if title and " - " in title:
        tail = title.split(" - ", 1)[1].strip()
        if tail and norm_text(tail) != norm_text(product_name):
            return tail
    return "Standard"


def color_tokens(name: str) -> set[str]:
    colors = {
        "lavender",
        "misty blue",
        "rouge pink",
        "dark grey",
        "grey",
        "gray",
        "sage",
        "rose",
        "light grey",
    }
    n = norm_text(name)
    found = {c for c in colors if c in n}
    for part in re.split(r"[/|]", name):
        p = norm_text(part)
        if p in colors:
            found.add(p)
    return found


def size_token(name: str) -> str | None:
    m = re.search(r"(\d+(?:\.\d+)?)\s*in", norm_variant(name))
    return m.group(1) if m else None


def type_tokens(name: str) -> set[str]:
    known = {
        "buddha",
        "mantra",
        "om",
        "yin yang",
        "sacred geometry",
        "chakra",
        "buddhist om",
        "deep",
    }
    n = norm_text(name)
    return {k for k in known if k in n}


def variant_score(ls: str, do: str, ls_product: str) -> float:
    ls_n, do_n = norm_variant(ls), norm_variant(do)
    if ls_n == do_n:
        return 100.0
    if set(ls_n.split()) == set(do_n.split()):
        return 99.0

    base = float(fuzz.token_sort_ratio(ls_n, do_n))

    ls_size, do_size = size_token(ls), size_token(do)
    if ls_size and do_size:
        if ls_size == do_size:
            base = max(base, 92.0)
        else:
            return 0.0  # size mismatch — never cross-match (e.g. 13 in vs 7 in)

    ls_colors = color_tokens(ls)
    do_colors = color_tokens(do)
    if ls_colors and do_colors and ls_colors & do_colors:
        base = max(base, 90.0)

    ls_types = type_tokens(ls)
    do_types = type_tokens(do)
    if ls_types and do_types:
        if ls_types & do_types:
            base = max(base, 88.0)
        else:
            base -= 20

    # Crescent Zafu Wide Cotton: LS "Lavender / Lavender" ↔ DO "Lavender / Large"
    if "crescent zafu" in norm_text(ls_product) and "wide" in norm_text(ls_product):
        if ls_colors & do_colors:
            base = max(base, 93.0)

    return base


def load_ls_catalog() -> dict[str, list[LsVar]]:
    data = json.loads(LS_EXPORT.read_text(encoding="utf-8"))
    grouped: dict[str, list[LsVar]] = defaultdict(list)
    for r in data["rows"]:
        name = r.get("name") or ""
        key = norm_text(name)
        grouped[key].append(
            LsVar(
                sku=(r.get("sku") or "").strip(),
                name=(r.get("variantName") or "").strip() or "Standard",
                variant_id=r.get("variantId") or "",
                product_id=r.get("productId") or "",
            )
        )
    return grouped


def load_do_catalog() -> tuple[dict[str, dict], dict[str, list[DoVar]]]:
    products: dict[str, dict] = {}
    with DO_PRODUCTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            products[r["id"]] = r

    by_name: dict[str, list[DoVar]] = defaultdict(list)
    with DO_VARIANTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            parent = products.get(r.get("parent_id") or "")
            if not parent:
                continue
            pname = parent.get("name") or ""
            attrs = parse_attrs(r.get("attrs") or "")
            by_name[norm_text(pname)].append(
                DoVar(
                    id=r["id"],
                    name=parse_do_variant_name(attrs, r.get("title") or "", pname),
                    sku=(r.get("sku") or "").strip(),
                    thumb_id=(r.get("thumb_id") or "").strip(),
                    attrs=attrs,
                )
            )
    name_map = {norm_text(p["name"]): p for p in products.values()}
    return name_map, by_name


def load_product_pairs() -> list[ProductPair]:
    pairs: dict[tuple[str, str], ProductPair] = {}

    # From confidence xlsx (both sheets) — product names already matched
    if CONFIDENCE_XLSX.exists():
        wb = load_workbook(CONFIDENCE_XLSX, read_only=True, data_only=True)
        for sheet in ("Confident", "Not Confident"):
            if sheet not in wb.sheetnames:
                continue
            for row in wb[sheet].iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                ls, do = str(row[0]), str(row[1] or row[0])
                key = (norm_text(ls), norm_text(do))
                if key not in pairs:
                    pairs[key] = ProductPair(
                        ls_name=ls,
                        do_name=do,
                        ls_slug="",
                        woo_id=None,
                        source="confidence-xlsx",
                    )
        wb.close()

    # Enrich woo_id + slug from LS export
    ls_data = json.loads(LS_EXPORT.read_text(encoding="utf-8"))
    ls_meta = {}
    for r in ls_data["rows"]:
        k = norm_text(r.get("name") or "")
        if k not in ls_meta:
            ls_meta[k] = {
                "slug": r.get("slug") or "",
                "woo": r.get("wooCommerceId"),
            }
    for p in pairs.values():
        m = ls_meta.get(norm_text(p.ls_name), {})
        p.ls_slug = m.get("slug") or ""
        woo = m.get("woo")
        p.woo_id = int(woo) if woo else None

    return sorted(pairs.values(), key=lambda p: p.ls_name.lower())


def match_product(pair: ProductPair, ls_vars: list[LsVar], do_vars: list[DoVar]) -> list[MatchRow]:
    rows: list[MatchRow] = []
    if norm_text(pair.ls_name) in HOLD_LS_PRODUCTS:
        for lv in ls_vars:
            rows.append(
                MatchRow(
                    pair.ls_name,
                    pair.do_name,
                    lv.name,
                    "—",
                    lv.sku,
                    "",
                    "Product on HOLD — Etched Chau Gongs",
                    "hold",
                )
            )
        return rows

    used_do: set[str] = set()

    # DNA Tuning Fork: single new LS variant → fallback to any DO variant with media
    if norm_text(pair.ls_name) == DNA_LS_PRODUCT and len(ls_vars) == 1 and do_vars:
        lv = ls_vars[0]
        donor = next((d for d in do_vars if d.thumb_id), do_vars[0])
        rows.append(
            MatchRow(
                pair.ls_name,
                pair.do_name,
                lv.name,
                donor.name,
                lv.sku,
                donor.sku,
                "DNA new single variant — fallback image from DO variant",
                "pull_fallback",
                donor.id,
            )
        )
        return rows

    for lv in ls_vars:
        # Zafu Wide Cotton: only pull Lavender, Misty Blue, Rouge Pink
        if norm_text(pair.ls_name) == norm_text("Crescent Zafu Cushion - Wide -Cotton"):
            colors = color_tokens(lv.name)
            allowed = {"lavender", "misty blue", "rouge pink"}
            if not (colors & allowed):
                rows.append(
                    MatchRow(
                        pair.ls_name,
                        pair.do_name,
                        lv.name,
                        "—",
                        lv.sku,
                        "",
                        "New Lightsail-only variant (no DO colour match)",
                        "skip_ls_only",
                    )
                )
                continue

        best: DoVar | None = None
        best_score = 0.0
        for dv in do_vars:
            if dv.id in used_do:
                continue
            s = variant_score(lv.name, dv.name, pair.ls_name)
            if s > best_score:
                best_score, best = s, dv

        if best and best_score >= VARIANT_MIN:
            used_do.add(best.id)
            note = (
                "exact variant name"
                if best_score >= 99
                else f"secondary fuzzy variant ({best_score:.0f}%)"
            )
            rows.append(
                MatchRow(
                    pair.ls_name,
                    pair.do_name,
                    lv.name,
                    best.name,
                    lv.sku,
                    best.sku,
                    note,
                    "pull",
                    best.id,
                )
            )
        else:
            rows.append(
                MatchRow(
                    pair.ls_name,
                    pair.do_name,
                    lv.name,
                    "—",
                    lv.sku,
                    "",
                    "Lightsail-only variant — no DO match",
                    "skip_ls_only",
                )
            )

    for dv in do_vars:
        if dv.id in used_do:
            continue
        rows.append(
            MatchRow(
                pair.ls_name,
                pair.do_name,
                "—",
                dv.name,
                "",
                dv.sku,
                "DO-only variant — removed on Lightsail",
                "skip_do_only",
                dv.id,
            )
        )

    return rows


def resolve_conflicts(rows: list[MatchRow]) -> list[MatchRow]:
    """If one DO variation maps to multiple LS SKUs, skip those pulls (ambiguous)."""
    by_do: dict[str, list[MatchRow]] = defaultdict(list)
    for r in rows:
        if r.action in ("pull", "pull_fallback") and r.do_variation_id:
            by_do[r.do_variation_id].append(r)

    conflict_ids = {did for did, rs in by_do.items() if len(rs) > 1}
    if not conflict_ids:
        return rows

    out: list[MatchRow] = []
    for r in rows:
        if r.do_variation_id in conflict_ids and r.action in ("pull", "pull_fallback"):
            out.append(
                MatchRow(
                    r.ls_product,
                    r.do_product,
                    r.ls_variant,
                    r.do_variant,
                    r.ls_sku,
                    r.do_sku,
                    "DO variation matched to multiple LS SKUs — skipped (conflict)",
                    "skip_conflict",
                    r.do_variation_id,
                )
            )
        else:
            out.append(r)
    return out


def write_xlsx(rows: list[MatchRow]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pull List v2"
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for ri, r in enumerate(rows, 2):
        vals = [
            r.ls_product,
            r.do_product,
            r.ls_variant,
            r.do_variant,
            r.ls_sku,
            r.do_sku,
            r.note,
            r.action,
            r.do_variation_id,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=col, value=v)
            if r.action == "pull":
                cell.fill = GREEN
            elif r.action in ("skip_ls_only", "skip_do_only"):
                cell.fill = YELLOW
            elif r.action == "hold":
                cell.fill = GRAY
    for i, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(48, max(12, len(h) + 2))

    sm = wb.create_sheet("Summary")
    pull = [r for r in rows if r.action in ("pull", "pull_fallback")]
    sm_data = [
        ["Generated (UTC)", datetime.now(timezone.utc).isoformat()],
        ["Total rows", len(rows)],
        ["Pull rows", len(pull)],
        ["Skip LS-only", sum(1 for r in rows if r.action == "skip_ls_only")],
        ["Skip DO-only", sum(1 for r in rows if r.action == "skip_do_only")],
        ["Hold products", sum(1 for r in rows if r.action == "hold")],
        [
            "Unique products with ≥1 pull",
            len({r.ls_product for r in pull}),
        ],
    ]
    for a, b in sm_data:
        sm.append([a, b])

    wb.save(OUT_XLSX)


def main() -> None:
    for p in (DO_PRODUCTS, DO_VARIANTS, LS_EXPORT):
        if not p.exists():
            raise SystemExit(f"Missing {p}")

    ls_catalog = load_ls_catalog()
    _, do_by_name = load_do_catalog()
    pairs = load_product_pairs()

    all_rows: list[MatchRow] = []
    for pair in pairs:
        ls_vars = ls_catalog.get(norm_text(pair.ls_name), [])
        do_vars = do_by_name.get(norm_text(pair.do_name), [])
        if not ls_vars:
            continue
        if not do_vars and (do_by_name.get(norm_text(pair.ls_name))):
            do_vars = do_by_name[norm_text(pair.ls_name)]
        all_rows.extend(match_product(pair, ls_vars, do_vars))

    all_rows = resolve_conflicts(all_rows)

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "rows": [asdict(r) for r in all_rows],
    }
    OUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    write_xlsx(all_rows)

    pull = [r for r in all_rows if r.action in ("pull", "pull_fallback")]
    print(f"Products in scope: {len(pairs)}")
    print(f"Pull rows: {len(pull)}")
    print(f"Unique products with pulls: {len({r.ls_product for r in pull})}")
    print(f"Wrote {OUT_JSON}")
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
