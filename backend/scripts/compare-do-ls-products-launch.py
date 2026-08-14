#!/usr/bin/env python3
"""
DO (sarveda.com) vs Lightsail (sarveda-demo.xyz) — product-level match report.

Products only (no variant parity). Match tiers:
  1. exact_slug — same slug
  2. exact_woo_id — same Woo product ID (DO id = Lightsail wooCommerceId)
  3. fuzzy_name — token_sort_ratio >= threshold (default 82)
  4. lightsail_only / do_only — no match on the other side

Inputs (refresh same session):
  data/compare/do_products.csv
  data/compare/lightsail-catalog-export.json

Output:
  data/compare/launch-do-vs-ls-products.xlsx
  data/compare/launch-do-vs-ls-products-summary.json

Usage:
  python3 backend/scripts/compare-do-ls-products-launch.py
"""
from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

ROOT = Path(__file__).resolve().parents[2]
DO_PRODUCTS = ROOT / "data/compare/do_products.csv"
LS_EXPORT = ROOT / "data/compare/lightsail-catalog-export.json"
OUT_XLSX = ROOT / "data/compare/launch-do-vs-ls-products.xlsx"
OUT_JSON = ROOT / "data/compare/launch-do-vs-ls-products-summary.json"

FUZZY_MIN = 82

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN = PatternFill("solid", fgColor="E2EFDA")
BLUE = PatternFill("solid", fgColor="DDEBF7")
RED = PatternFill("solid", fgColor="FCE4D6")
YELLOW = PatternFill("solid", fgColor="FFF2CC")


@dataclass
class Product:
    source: str
    key: str
    slug: str
    name: str
    woo_id: int | None
    status: str
    variant_count: int = 0


def norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def is_shop_slug(slug: str) -> bool:
    s = (slug or "").lower()
    return not (s.startswith("course-checkout-") or s.startswith("event-checkout-"))


def load_do() -> list[Product]:
    out: list[Product] = []
    with DO_PRODUCTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            slug = (r.get("slug") or "").strip()
            if not is_shop_slug(slug):
                continue
            pid = int(r["id"])
            out.append(
                Product(
                    source="do",
                    key=f"do:{pid}",
                    slug=slug,
                    name=(r.get("name") or "").strip(),
                    woo_id=pid,
                    status="publish",
                )
            )
    return out


def load_ls() -> list[Product]:
    data = json.loads(LS_EXPORT.read_text(encoding="utf-8"))
    grouped: dict[str, Product] = {}
    for r in data["rows"]:
        slug = (r.get("slug") or "").strip()
        if not is_shop_slug(slug):
            continue
        if slug not in grouped:
            grouped[slug] = Product(
                source="ls",
                key=f"ls:{slug}",
                slug=slug,
                name=(r.get("name") or "").strip(),
                woo_id=int(r["wooCommerceId"]) if r.get("wooCommerceId") else None,
                status="ACTIVE",
            )
        grouped[slug].variant_count += 1
    return sorted(grouped.values(), key=lambda p: p.slug)


def fuzzy_score(a: Product, b: Product) -> float:
    return float(fuzz.token_sort_ratio(norm_text(a.name), norm_text(b.name)))


def write_sheet(ws, headers: list[str], rows: list[list], widths: list[int] | None = None):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
    for ri, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=ri, column=col, value=val)
    for i, h in enumerate(headers, 1):
        w = widths[i - 1] if widths else min(48, max(12, len(h) + 2))
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> None:
    for p in (DO_PRODUCTS, LS_EXPORT):
        if not p.exists():
            raise SystemExit(f"Missing {p} — refresh live dumps first")

    do_products = load_do()
    ls_products = load_ls()

    ls_by_slug = {norm_text(p.slug): p for p in ls_products if p.slug}
    do_by_slug = {norm_text(p.slug): p for p in do_products if p.slug}
    ls_by_woo = {p.woo_id: p for p in ls_products if p.woo_id}
    do_by_woo = {p.woo_id: p for p in do_products if p.woo_id}

    used_ls: set[str] = set()
    used_do: set[str] = set()

    exact_slug_pairs: list[tuple[Product, Product]] = []
    exact_woo_pairs: list[tuple[Product, Product]] = []

    for lp in ls_products:
        ns = norm_text(lp.slug)
        if not ns or ns not in do_by_slug:
            continue
        dp = do_by_slug[ns]
        if dp.key in used_do or lp.key in used_ls:
            continue
        exact_slug_pairs.append((lp, dp))
        used_ls.add(lp.key)
        used_do.add(dp.key)

    for lp in ls_products:
        if lp.key in used_ls:
            continue
        if lp.woo_id and lp.woo_id in do_by_woo:
            dp = do_by_woo[lp.woo_id]
            if dp.key in used_do:
                continue
            exact_woo_pairs.append((lp, dp))
            used_ls.add(lp.key)
            used_do.add(dp.key)

    ls_remaining = [p for p in ls_products if p.key not in used_ls]
    do_remaining = [p for p in do_products if p.key not in used_do]

    fuzzy_pairs: list[tuple[Product, Product, float]] = []
    candidates: list[tuple[float, int, int]] = []
    for li, lp in enumerate(ls_remaining):
        for di, dp in enumerate(do_remaining):
            s = fuzzy_score(lp, dp)
            if s >= FUZZY_MIN:
                candidates.append((s, li, di))
    candidates.sort(reverse=True)

    used_lr: set[int] = set()
    used_dr: set[int] = set()
    for s, li, di in candidates:
        if li in used_lr or di in used_dr:
            continue
        used_lr.add(li)
        used_dr.add(di)
        fuzzy_pairs.append((ls_remaining[li], do_remaining[di], s))
        used_ls.add(ls_remaining[li].key)
        used_do.add(do_remaining[di].key)

    ls_only = [p for p in ls_products if p.key not in used_ls]
    do_only = [p for p in do_products if p.key not in used_do]

    total_matched = len(exact_slug_pairs) + len(exact_woo_pairs) + len(fuzzy_pairs)

    summary = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "do_products_publish": len(do_products),
        "lightsail_products_active": len(ls_products),
        "matched_total": total_matched,
        "exact_slug_matches": len(exact_slug_pairs),
        "exact_woo_id_matches": len(exact_woo_pairs),
        "fuzzy_name_matches": len(fuzzy_pairs),
        "lightsail_only": len(ls_only),
        "do_only": len(do_only),
        "fuzzy_threshold": FUZZY_MIN,
        "sources": {
            "do": "DigitalOcean MySQL (sarveda.com Woo publish products)",
            "lightsail": "Lightsail Postgres (sarveda-demo.xyz ACTIVE storefront)",
        },
        "inputs": {"do_products": str(DO_PRODUCTS), "lightsail_export": str(LS_EXPORT)},
        "outputs": {"xlsx": str(OUT_XLSX), "json": str(OUT_JSON)},
    }
    OUT_JSON.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    wb = Workbook()
    sm = wb.active
    sm.title = "Summary"
    write_sheet(
        sm,
        ["Metric", "Count"],
        [
            ["DO sarveda.com — publish shop products", len(do_products)],
            ["Lightsail demo — ACTIVE shop products", len(ls_products)],
            ["", ""],
            ["TOTAL matched (any tier)", total_matched],
            ["  Exact slug match", len(exact_slug_pairs)],
            ["  Exact Woo ID match (slug differed)", len(exact_woo_pairs)],
            ["  Fuzzy name match (≥82)", len(fuzzy_pairs)],
            ["", ""],
            ["Lightsail ONLY (not on DO)", len(ls_only)],
            ["DO ONLY (not on Lightsail)", len(do_only)],
            ["", ""],
            ["Fuzzy name threshold", FUZZY_MIN],
            ["DO products CSV", str(DO_PRODUCTS)],
            ["Lightsail export JSON", str(LS_EXPORT)],
            ["Generated (UTC)", summary["generatedAt"]],
        ],
        [52, 14],
    )

    def pair_rows(pairs: list[tuple[Product, Product, float | None]], match_type: str):
        rows = []
        for item in pairs:
            if len(item) == 3:
                lp, dp, score = item
                rows.append(
                    [
                        match_type,
                        round(score, 1) if score is not None else "",
                        lp.slug,
                        dp.slug,
                        lp.name,
                        dp.name,
                        lp.woo_id or "",
                        dp.woo_id or "",
                        lp.variant_count,
                        dp.variant_count if hasattr(dp, "variant_count") else "",
                    ]
                )
            else:
                lp, dp = item
                rows.append(
                    [
                        match_type,
                        100.0,
                        lp.slug,
                        dp.slug,
                        lp.name,
                        dp.name,
                        lp.woo_id or "",
                        dp.woo_id or "",
                        lp.variant_count,
                        "",
                    ]
                )
        return rows

    # attach variant counts for DO in pairs
    do_var_count: dict[int, int] = {}
    with (ROOT / "data/compare/do_variants.csv").open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            pid = int(r["parent_id"])
            do_var_count[pid] = do_var_count.get(pid, 0) + 1

    def enrich_do_var_count(rows: list[list]) -> list[list]:
        out = []
        for row in rows:
            r = list(row)
            woo = r[7]
            if woo != "":
                r[9] = do_var_count.get(int(woo), 0) or (1 if woo else 0)
            out.append(r)
        return out

    headers = [
        "Match type",
        "Score",
        "Lightsail slug",
        "DO slug",
        "Lightsail name",
        "DO name",
        "Lightsail wooCommerceId",
        "DO id",
        "LS variant rows",
        "DO variant rows",
    ]

    em = wb.create_sheet("Exact Slug Matches")
    rows_slug = enrich_do_var_count(
        pair_rows([(a, b) for a, b in exact_slug_pairs], "exact_slug")  # type: ignore[arg-type]
    )
    write_sheet(em, headers, rows_slug, [14, 8, 28, 28, 40, 40, 14, 10, 12, 12])
    for ri in range(2, em.max_row + 1):
        em.cell(row=ri, column=1).fill = GREEN

    ew = wb.create_sheet("Exact Woo ID Matches")
    rows_woo = enrich_do_var_count(
        pair_rows([(a, b) for a, b in exact_woo_pairs], "exact_woo_id")  # type: ignore[arg-type]
    )
    write_sheet(ew, headers, rows_woo, [14, 8, 28, 28, 40, 40, 14, 10, 12, 12])
    for ri in range(2, ew.max_row + 1):
        ew.cell(row=ri, column=1).fill = GREEN

    fm = wb.create_sheet("Fuzzy Name Matches")
    rows_fuzzy = enrich_do_var_count(
        pair_rows([(a, b, s) for a, b, s in fuzzy_pairs], "fuzzy_name")
    )
    write_sheet(fm, headers, rows_fuzzy, [14, 8, 28, 28, 40, 40, 14, 10, 12, 12])
    for ri in range(2, fm.max_row + 1):
        fm.cell(row=ri, column=1).fill = BLUE

    lo = wb.create_sheet("Lightsail Only")
    write_sheet(
        lo,
        ["Lightsail slug", "Product name", "wooCommerceId", "Variant rows (info)"],
        [[p.slug, p.name, p.woo_id or "", p.variant_count] for p in sorted(ls_only, key=lambda x: x.name.lower())],
        [30, 44, 14, 14],
    )
    for ri in range(2, lo.max_row + 1):
        lo.cell(row=ri, column=1).fill = RED

    doo = wb.create_sheet("DO Only")
    write_sheet(
        doo,
        ["DO slug", "Product name", "DO id", "Variant rows (info)"],
        [
            [p.slug, p.name, p.woo_id or "", do_var_count.get(p.woo_id or 0, 0) or 1]
            for p in sorted(do_only, key=lambda x: x.name.lower())
        ],
        [30, 44, 14, 14],
    )
    for ri in range(2, doo.max_row + 1):
        doo.cell(row=ri, column=1).fill = YELLOW

    wb.save(OUT_XLSX)
    print(json.dumps(summary, indent=2))
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
