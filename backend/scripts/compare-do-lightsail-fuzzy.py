#!/usr/bin/env python3
"""
DO MySQL vs Lightsail Postgres — fuzzy product/variant name comparison.

Matches on **product name + variant name only** (not SKU/slug as primary key).
Shows Lightsail-only, DO-only, and fuzzy-matched pairs with variant balance.

Inputs (refresh in same session):
  data/compare/do_products.csv
  data/compare/do_variants.csv
  data/compare/lightsail-catalog-export.json

Output:
  data/compare/do-vs-lightsail-fuzzy.xlsx
  data/compare/do-vs-lightsail-fuzzy-summary.json

Usage:
  python3 backend/scripts/compare-do-lightsail-fuzzy.py
"""
from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

ROOT = Path(__file__).resolve().parents[2]
DO_PRODUCTS = ROOT / "data/compare/do_products.csv"
DO_VARIANTS = ROOT / "data/compare/do_variants.csv"
LS_EXPORT = ROOT / "data/compare/lightsail-catalog-export.json"
OUT_XLSX = ROOT / "data/compare/do-vs-lightsail-fuzzy.xlsx"
OUT_JSON = ROOT / "data/compare/do-vs-lightsail-fuzzy-summary.json"

PRODUCT_MATCH_MIN = 82
VARIANT_MATCH_MIN = 85

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN = PatternFill("solid", fgColor="E2EFDA")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
RED = PatternFill("solid", fgColor="FCE4D6")
BLUE = PatternFill("solid", fgColor="DDEBF7")


@dataclass
class VariantRow:
    name: str
    sku: str = ""


@dataclass
class ProductRow:
    key: str
    slug: str
    name: str
    woo_id: int | None = None
    variants: list[VariantRow] = field(default_factory=list)


def norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def norm_variant(s: str) -> str:
    s = norm_text(s)
    return re.sub(r"[\s|/·,–—\-]+", " ", s).strip()


def norm_sku(s: str) -> str:
    return norm_text(s).upper()


def parse_do_variant_name(attrs: str, title: str, product_name: str) -> str:
    if attrs:
        parts = []
        for seg in attrs.split(";"):
            if "=" in seg:
                parts.append(seg.split("=", 1)[1].strip())
        if parts:
            return " / ".join(parts)
    if title and " - " in title:
        tail = title.split(" - ", 1)[1].strip()
        if tail and norm_text(tail) != norm_text(product_name):
            return tail
    return "Standard"


def variant_exact(a: str, b: str) -> bool:
    na, nb = norm_variant(a), norm_variant(b)
    if not na and not nb:
        return True
    if na == nb:
        return True
    if na and nb and set(na.split()) == set(nb.split()):
        return True
    return False


def variant_score(a: str, b: str) -> float:
    na, nb = norm_variant(a), norm_variant(b)
    if not na and not nb:
        return 100.0
    if variant_exact(a, b):
        return 100.0
    return float(fuzz.token_sort_ratio(na, nb))


def product_score(a: ProductRow, b: ProductRow) -> float:
    if a.woo_id and b.woo_id and a.woo_id == b.woo_id:
        return 100.0
    if norm_text(a.slug) == norm_text(b.slug):
        return 99.0
    return float(fuzz.token_sort_ratio(norm_text(a.name), norm_text(b.name)))


def is_shop_product_slug(slug: str) -> bool:
    s = (slug or "").lower()
    return not (s.startswith("course-checkout-") or s.startswith("event-checkout-"))


def load_do_products() -> list[ProductRow]:
    products: dict[str, dict] = {}
    with DO_PRODUCTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            products[r["id"]] = r

    by_parent: dict[str, list] = {}
    with DO_VARIANTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            by_parent.setdefault(r["parent_id"], []).append(r)

    out: list[ProductRow] = []
    for pid, p in products.items():
        slug = p.get("slug") or ""
        if not is_shop_product_slug(slug):
            continue
        ptype = (p.get("product_type") or "").lower()
        vars_ = by_parent.get(pid, [])
        variants: list[VariantRow] = []
        if ptype == "simple" or not vars_:
            variants.append(VariantRow(name="Standard", sku=(p.get("sku") or "").strip()))
        else:
            for v in vars_:
                variants.append(
                    VariantRow(
                        name=parse_do_variant_name(v.get("attrs") or "", v.get("title") or "", p.get("name") or ""),
                        sku=(v.get("sku") or "").strip(),
                    )
                )
        out.append(
            ProductRow(
                key=f"do:{pid}",
                slug=slug,
                name=p.get("name") or "",
                variants=variants,
            )
        )
    return out


def load_ls_products() -> list[ProductRow]:
    data = json.loads(LS_EXPORT.read_text(encoding="utf-8"))
    grouped: dict[str, ProductRow] = {}
    for r in data["rows"]:
        slug = r.get("slug") or ""
        if not is_shop_product_slug(slug):
            continue
        if slug not in grouped:
            grouped[slug] = ProductRow(
                key=f"ls:{slug}",
                slug=slug,
                name=r.get("name") or "",
                woo_id=int(r["wooCommerceId"]) if r.get("wooCommerceId") else None,
                variants=[],
            )
        vname = (r.get("variantName") or "").strip() or "Standard"
        grouped[slug].variants.append(VariantRow(name=vname, sku=(r.get("sku") or "").strip()))
    return sorted(grouped.values(), key=lambda p: p.slug)


def greedy_pairs(
    left: list[ProductRow],
    right: list[ProductRow],
    score_fn,
    min_score: float,
) -> tuple[list[tuple[int, int, float]], list[int], list[int]]:
    candidates: list[tuple[float, int, int]] = []
    for li, lp in enumerate(left):
        for ri, rp in enumerate(right):
            s = score_fn(lp, rp)
            if s >= min_score:
                candidates.append((s, li, ri))
    candidates.sort(reverse=True)

    used_l: set[int] = set()
    used_r: set[int] = set()
    pairs: list[tuple[int, int, float]] = []
    for s, li, ri in candidates:
        if li in used_l or ri in used_r:
            continue
        used_l.add(li)
        used_r.add(ri)
        pairs.append((li, ri, s))

    unmatched_l = [i for i in range(len(left)) if i not in used_l]
    unmatched_r = [i for i in range(len(right)) if i not in used_r]
    return pairs, unmatched_l, unmatched_r


def match_variants(ls_vars: list[VariantRow], do_vars: list[VariantRow]) -> dict:
    if not ls_vars and not do_vars:
        return {"matched": [], "ls_only": [], "do_only": [], "balance": "both empty"}

    candidates: list[tuple[float, int, int, str]] = []
    for li, lv in enumerate(ls_vars):
        for di, dv in enumerate(do_vars):
            s = variant_score(lv.name, dv.name)
            if s >= VARIANT_MATCH_MIN or variant_exact(lv.name, dv.name):
                mtype = "exact" if variant_exact(lv.name, dv.name) else "fuzzy"
                candidates.append((s, li, di, mtype))
    candidates.sort(reverse=True)

    used_l: set[int] = set()
    used_d: set[int] = set()
    matched: list[dict] = []
    for s, li, di, mtype in candidates:
        if li in used_l or di in used_d:
            continue
        used_l.add(li)
        used_d.add(di)
        lv, dv = ls_vars[li], do_vars[di]
        matched.append(
            {
                "ls_variant": lv.name,
                "do_variant": dv.name,
                "ls_sku": lv.sku,
                "do_sku": dv.sku,
                "sku_match": norm_sku(lv.sku) == norm_sku(dv.sku) if lv.sku and dv.sku else "N/A",
                "match_type": mtype,
                "score": round(s, 1),
            }
        )

    ls_only = [ls_vars[i] for i in range(len(ls_vars)) if i not in used_l]
    do_only = [do_vars[i] for i in range(len(do_vars)) if i not in used_d]

    if len(matched) == len(ls_vars) == len(do_vars) and all(m["match_type"] == "exact" for m in matched):
        balance = "exact variant parity"
    elif len(ls_vars) == len(do_vars) == len(matched):
        balance = "same variant count (fuzzy names)"
    elif len(ls_vars) > len(do_vars):
        balance = f"Lightsail has {len(ls_only)} extra variant(s)"
    elif len(do_vars) > len(ls_vars):
        balance = f"DO has {len(do_only)} extra variant(s)"
    elif ls_only and do_only:
        balance = f"variant mismatch — LS +{len(ls_only)} / DO +{len(do_only)}"
    else:
        balance = "partial variant overlap"

    return {"matched": matched, "ls_only": ls_only, "do_only": do_only, "balance": balance}


def fmt_variants(variants: list[VariantRow]) -> str:
    if not variants:
        return "—"
    return " | ".join(f"{v.name}" + (f" [{v.sku}]" if v.sku else "") for v in variants)


def write_sheet(ws, headers: list[str], rows: list[list], col_widths: list[int] | None = None):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
    for ri, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=ri, column=col, value=val)
    for i, h in enumerate(headers, 1):
        w = col_widths[i - 1] if col_widths else min(48, max(12, len(h) + 2))
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> None:
    for p in (DO_PRODUCTS, DO_VARIANTS, LS_EXPORT):
        if not p.exists():
            raise SystemExit(f"Missing {p} — refresh live dumps first")

    ls_products = load_ls_products()
    do_products = load_do_products()

    pairs, ls_unmatched_idx, do_unmatched_idx = greedy_pairs(
        ls_products, do_products, product_score, PRODUCT_MATCH_MIN
    )

    ls_only = [ls_products[i] for i in ls_unmatched_idx]
    do_only = [do_products[i] for i in do_unmatched_idx]

    matched_products: list[dict] = []
    matched_variant_rows: list[list] = []

    for li, di, pscore in pairs:
        lp, dp = ls_products[li], do_products[di]
        vm = match_variants(lp.variants, dp.variants)
        matched_products.append(
            {
                "product_score": round(pscore, 1),
                "lightsail_slug": lp.slug,
                "do_slug": dp.slug,
                "lightsail_name": lp.name,
                "do_name": dp.name,
                "ls_variant_count": len(lp.variants),
                "do_variant_count": len(dp.variants),
                "matched_variant_count": len(vm["matched"]),
                "ls_only_variant_count": len(vm["ls_only"]),
                "do_only_variant_count": len(vm["do_only"]),
                "variant_balance": vm["balance"],
            }
        )
        for m in vm["matched"]:
            matched_variant_rows.append(
                [
                    lp.name,
                    dp.name,
                    round(pscore, 1),
                    m["ls_variant"],
                    m["do_variant"],
                    m["match_type"],
                    round(m["score"], 1),
                    m["ls_sku"],
                    m["do_sku"],
                    m["sku_match"],
                    vm["balance"],
                ]
            )
        for v in vm["ls_only"]:
            matched_variant_rows.append(
                [lp.name, dp.name, round(pscore, 1), v.name, "—", "ls_only", "—", v.sku, "—", "—", vm["balance"]]
            )
        for v in vm["do_only"]:
            matched_variant_rows.append(
                [lp.name, dp.name, round(pscore, 1), "—", v.name, "do_only", "—", "—", v.sku, "—", vm["balance"]]
            )

    summary = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "lightsail_products": len(ls_products),
        "do_products": len(do_products),
        "fuzzy_matched_products": len(matched_products),
        "lightsail_only_products": len(ls_only),
        "do_only_products": len(do_only),
        "product_match_threshold": PRODUCT_MATCH_MIN,
        "variant_match_threshold": VARIANT_MATCH_MIN,
        "inputs": {
            "do_products": str(DO_PRODUCTS),
            "do_variants": str(DO_VARIANTS),
            "lightsail_export": str(LS_EXPORT),
        },
        "outputs": {"xlsx": str(OUT_XLSX), "json": str(OUT_JSON)},
    }
    OUT_JSON.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    wb = Workbook()

    sm = wb.active
    sm.title = "Summary"
    write_sheet(
        sm,
        ["Metric", "Count"],
        [
            ["Lightsail ACTIVE storefront products", len(ls_products)],
            ["DO publish products", len(do_products)],
            ["Fuzzy matched product pairs", len(matched_products)],
            ["Lightsail only (no DO fuzzy match)", len(ls_only)],
            ["DO only (no Lightsail fuzzy match)", len(do_only)],
            ["Product name match threshold", PRODUCT_MATCH_MIN],
            ["Variant name match threshold", VARIANT_MATCH_MIN],
            ["Lightsail export", str(LS_EXPORT)],
            ["DO products CSV", str(DO_PRODUCTS)],
        ],
    )

    lo = wb.create_sheet("Lightsail Only")
    write_sheet(
        lo,
        [
            "Lightsail slug",
            "Product name",
            "Variant count",
            "Variants (name [SKU])",
        ],
        [
            [p.slug, p.name, len(p.variants), fmt_variants(p.variants)]
            for p in sorted(ls_only, key=lambda x: x.name.lower())
        ],
        [28, 42, 12, 80],
    )
    for ri in range(2, lo.max_row + 1):
        lo.cell(row=ri, column=1).fill = RED

    doo = wb.create_sheet("DO Only")
    write_sheet(
        doo,
        [
            "DO slug",
            "Product name",
            "Variant count",
            "Variants (name [SKU])",
        ],
        [
            [p.slug, p.name, len(p.variants), fmt_variants(p.variants)]
            for p in sorted(do_only, key=lambda x: x.name.lower())
        ],
        [28, 42, 12, 80],
    )
    for ri in range(2, doo.max_row + 1):
        doo.cell(row=ri, column=1).fill = RED

    mp = wb.create_sheet("Fuzzy Matched Products")
    mp_headers = [
        "Product match score",
        "Lightsail slug",
        "DO slug",
        "Lightsail product name",
        "DO product name",
        "LS variants",
        "DO variants",
        "Matched variants",
        "LS-only variants",
        "DO-only variants",
        "Variant balance",
    ]
    write_sheet(
        mp,
        mp_headers,
        [
            [
                m["product_score"],
                m["lightsail_slug"],
                m["do_slug"],
                m["lightsail_name"],
                m["do_name"],
                m["ls_variant_count"],
                m["do_variant_count"],
                m["matched_variant_count"],
                m["ls_only_variant_count"],
                m["do_only_variant_count"],
                m["variant_balance"],
            ]
            for m in sorted(matched_products, key=lambda x: x["lightsail_name"].lower())
        ],
        [14, 28, 28, 38, 38, 10, 10, 12, 12, 12, 36],
    )
    balance_col = mp_headers.index("Variant balance") + 1
    for ri in range(2, mp.max_row + 1):
        bal = str(mp.cell(row=ri, column=balance_col).value or "")
        if "exact" in bal.lower():
            fill = GREEN
        elif "extra" in bal.lower() or "mismatch" in bal.lower():
            fill = YELLOW
        else:
            fill = BLUE
        mp.cell(row=ri, column=balance_col).fill = fill

    mv = wb.create_sheet("Fuzzy Matched Variants")
    mv_headers = [
        "Lightsail product",
        "DO product",
        "Product score",
        "Lightsail variant",
        "DO variant",
        "Variant match type",
        "Variant score",
        "Lightsail SKU",
        "DO SKU",
        "SKU match",
        "Product variant balance",
    ]
    write_sheet(mv, mv_headers, matched_variant_rows, [36, 36, 12, 28, 28, 14, 12, 16, 16, 10, 36])
    type_col = mv_headers.index("Variant match type") + 1
    for ri in range(2, mv.max_row + 1):
        t = str(mv.cell(row=ri, column=type_col).value or "")
        if t == "exact":
            mv.cell(row=ri, column=type_col).fill = GREEN
        elif t == "fuzzy":
            mv.cell(row=ri, column=type_col).fill = BLUE
        elif t == "ls_only":
            mv.cell(row=ri, column=type_col).fill = YELLOW
        elif t == "do_only":
            mv.cell(row=ri, column=type_col).fill = RED

    wb.save(OUT_XLSX)

    print(json.dumps(summary, indent=2))
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_JSON}")


if __name__ == "__main__":
    main()
