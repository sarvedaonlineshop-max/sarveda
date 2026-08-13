#!/usr/bin/env python3
"""
Apply Fuzzy Match sheet DECISION column to Lightsail catalog (name / variant / SKU only).

Reads: data/compare/latest-inventory-fuzzy.xlsx  (sheet "Fuzzy Match")
Uses:  PUT /api/admin/products/xl-sheet on Lightsail (preserves qty + prices)

Usage:
  export ADMIN_EMAIL=you@sarveda.com ADMIN_PASSWORD='...'
  python3 backend/scripts/apply-fuzzy-match-decisions.py --dry-run
  python3 backend/scripts/apply-fuzzy-match-decisions.py --apply
"""
from __future__ import annotations

import argparse
import json
import os
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_API = "http://13.204.112.165"
DEFAULT_XLSX = Path("data/compare/latest-inventory-fuzzy.xlsx")
OUT_DIR = Path("data/compare/fuzzy-apply")


def resolve_targets(
    decision: str,
    sheet_name: str,
    sheet_variant: str,
    sheet_sku: str,
    db_name: str,
    db_variant: str,
    db_sku: str,
) -> tuple[str, str, str] | None:
    d = (decision or "").strip().lower()
    if not d or "will do later" in d:
        return None

    product = db_name
    variant = db_variant
    sku = (db_sku or sheet_sku).strip()

    if "adapt sheet product name" in d or "product name adapt from sheet" in d:
        product = sheet_name
    elif "adapt db product name" in d or "product name same" in d:
        product = db_name

    if "vairant adapt from sheet" in d or "variant adapt from sheet" in d:
        variant = sheet_variant
    elif (
        "vairant adapt from db" in d
        or "variant adapt from db" in d
        or "vairant same" in d
        or "variant same" in d
    ):
        variant = db_variant

    return product.strip(), (variant or "").strip(), sku


def http_json(method: str, url: str, token: str | None = None, body: dict | None = None) -> dict:
    data = None
    headers = {"Accept": "application/json"}
    if body is not None:
        data = json.dumps(body).encode()
        headers["Content-Type"] = "application/json"
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        raw = e.read().decode()
        raise RuntimeError(f"HTTP {e.code} {url}: {raw[:500]}") from e


def login(api: str, email: str, password: str) -> str:
    res = http_json(
        "POST",
        f"{api}/api/auth/login",
        body={"email": email.strip(), "password": password},
    )
    token = res.get("data", {}).get("token")
    if not token:
        raise RuntimeError(f"Login failed: {res}")
    return token


def load_decisions(xlsx: Path) -> list[dict]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    if "Fuzzy Match" not in wb.sheetnames:
        raise SystemExit(f'Missing "Fuzzy Match" sheet in {xlsx}')
    ws = wb["Fuzzy Match"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h or "").strip() for h in rows[0]]
    out: list[dict] = []
    for r in rows[1:]:
        rec = {headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))}
        if not str(rec.get("DECISION") or "").strip():
            continue
        out.append(rec)
    wb.close()
    return out


def variant_label_from_api(attrs) -> str:
    if not attrs:
        return ""
    parts = sorted(
        (a["attributeValue"]["attribute"]["slug"], a["attributeValue"]["value"])
        for a in attrs
        if a.get("attributeValue")
    )
    return " / ".join(v for _, v in parts)


def fetch_public_catalog(api: str) -> dict[str, dict]:
    page = 1
    slugs: list[str] = []
    while True:
        data = http_json("GET", f"{api.rstrip('/')}/api/products?limit=100&page={page}&status=ACTIVE")
        slugs.extend(i["slug"] for i in data["data"]["items"])
        pag = data["data"]["pagination"]
        if page >= pag["totalPages"]:
            break
        page += 1

    by_sku: dict[str, dict] = {}

    def one(slug: str) -> None:
        data = http_json("GET", f"{api.rstrip('/')}/api/products/{slug}")
        p = data["data"]["product"]
        if p.get("deletedAt") or p.get("status") != "ACTIVE" or p.get("catalogHidden"):
            return
        for v in p.get("variants") or []:
            if v.get("status") != "ACTIVE":
                continue
            sku = str(v["sku"]).strip()
            inv = v.get("inventory")
            qty = int(inv.get("onHand") or 0) if isinstance(inv, dict) else 0
            by_sku[sku.upper()] = {
                "productId": p["id"],
                "variantId": v["id"],
                "productName": p["name"],
                "variantName": variant_label_from_api(v.get("attributeValues")),
                "sku": sku,
                "qty": qty,
                "costInPaise": v.get("costInPaise"),
                "mrpInPaise": v.get("mrpInPaise") or 0,
                "saleInPaise": v.get("saleInPaise") or 0,
                "mrpUsdCents": v.get("mrpUsdCents"),
                "saleUsdCents": v.get("saleUsdCents"),
                "mrpAedFils": v.get("mrpAedFils"),
                "saleAedFils": v.get("saleAedFils"),
                "mrpGbpPence": v.get("mrpGbpPence"),
                "saleGbpPence": v.get("saleGbpPence"),
                "hsnCode": p.get("hsnCode") or "",
                "productStatus": p.get("status"),
                "variantStatus": v.get("status"),
            }

    with ThreadPoolExecutor(max_workers=12) as ex:
        futs = [ex.submit(one, s) for s in slugs]
        for fut in as_completed(futs):
            fut.result()
    return by_sku


def fetch_xl_catalog(api: str, token: str) -> dict[str, dict]:
    xl = http_json(
        "GET",
        f"{api.rstrip('/')}/api/admin/products/xl-sheet?status=ALL",
        token=token,
    )
    rows = xl["data"]["rows"]
    return {str(r["sku"]).strip().upper(): r for r in rows}


def build_plan(decisions: list[dict], by_sku: dict[str, dict]) -> tuple[list[dict], list[dict], list[dict]]:
    planned: list[dict] = []
    skipped: list[dict] = []
    missing: list[dict] = []

    for rec in decisions:
        decision = str(rec.get("DECISION") or "")
        sheet_name = str(rec.get("Sheet Product name") or "")
        db_name = str(rec.get("DB Product name") or "")
        sheet_variant = str(rec.get("Sheet Variant name") or "")
        db_variant = str(rec.get("DB Variant name") or "")
        sheet_sku = str(rec.get("Sheet SKU") or "").strip()
        db_sku = str(rec.get("DB SKU") or sheet_sku).strip()

        targets = resolve_targets(
            decision, sheet_name, sheet_variant, sheet_sku, db_name, db_variant, db_sku
        )
        if targets is None:
            skipped.append({**rec, "reason": "will do later / empty decision"})
            continue

        target_product, target_variant, target_sku = targets
        row = by_sku.get(db_sku.upper())
        if not row:
            missing.append({**rec, "reason": f"SKU {db_sku} not in catalog"})
            continue

        cur_product = str(row.get("productName") or "")
        cur_variant = str(row.get("variantName") or "")
        cur_sku = str(row.get("sku") or "")

        if (
            cur_product == target_product
            and cur_variant == target_variant
            and cur_sku.upper() == target_sku.upper()
        ):
            skipped.append({**rec, "reason": "already matches target"})
            continue

        updated = {**row, "productName": target_product, "variantName": target_variant, "sku": target_sku}
        planned.append(
            {
                "decision": decision,
                "db_sku": db_sku,
                "from": {"productName": cur_product, "variantName": cur_variant, "sku": cur_sku},
                "to": {"productName": target_product, "variantName": target_variant, "sku": target_sku},
                "row": updated,
            }
        )

    return planned, skipped, missing


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--api", default=os.environ.get("SARVEDA_API_BASE", DEFAULT_API))
    ap.add_argument("--email", default=os.environ.get("ADMIN_EMAIL", ""))
    ap.add_argument("--password", default=os.environ.get("ADMIN_PASSWORD", ""))
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    if not args.dry_run and not args.apply:
        args.dry_run = True

    decisions = load_decisions(args.xlsx)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    token = None
    if args.apply:
        if not args.email or not args.password:
            raise SystemExit("Set ADMIN_EMAIL and ADMIN_PASSWORD (or --email / --password) for --apply")
        token = login(args.api.rstrip("/"), args.email, args.password)
        print("Logged in as", args.email)
        print("Loading xl-sheet catalog...")
        by_sku = fetch_xl_catalog(args.api, token)
    else:
        print("Loading public catalog (dry-run)...")
        by_sku = fetch_public_catalog(args.api)

    planned, skipped, missing = build_plan(decisions, by_sku)

    plan_path = OUT_DIR / "apply-plan.json"
    plan_path.write_text(json.dumps(planned, indent=2) + "\n")
    (OUT_DIR / "skipped.json").write_text(json.dumps(skipped, indent=2) + "\n")
    (OUT_DIR / "missing.json").write_text(json.dumps(missing, indent=2) + "\n")

    print(f"Decisions read:      {len(decisions)}")
    print(f"Rows to update:      {len(planned)}")
    print(f"Skipped (no change): {len(skipped)}")
    print(f"Missing SKU:         {len(missing)}")
    print(f"Plan:                {plan_path}")

    for p in planned[:12]:
        t, f = p["to"], p["from"]
        print(f"\n  {p['db_sku']}:")
        if f["productName"] != t["productName"]:
            print(f"    product: {f['productName']!r} -> {t['productName']!r}")
        if f["variantName"] != t["variantName"]:
            print(f"    variant: {f['variantName']!r} -> {t['variantName']!r}")
        if f["sku"] != t["sku"]:
            print(f"    sku:     {f['sku']!r} -> {t['sku']!r}")
    if len(planned) > 12:
        print(f"\n  ... +{len(planned) - 12} more in apply-plan.json")

    if not args.apply:
        print("\nDry run only. Re-run with --apply to write to Lightsail.")
        return

    payload_rows = [p["row"] for p in planned]
    results = []
    batch_size = 100
    for i in range(0, len(payload_rows), batch_size):
        chunk = payload_rows[i : i + batch_size]
        res = http_json(
            "PUT",
            f"{args.api.rstrip('/')}/api/admin/products/xl-sheet",
            token=token,
            body={"rows": chunk},
        )
        results.append(res.get("data", res))
        print(f"Batch {i // batch_size + 1}: saved {len(chunk)} rows ->", res.get("data", res))

    (OUT_DIR / "apply-result.json").write_text(json.dumps(results, indent=2) + "\n")
    print("\nDone. Verify on admin Products XL View or re-run fuzzy compare.")


if __name__ == "__main__":
    main()
