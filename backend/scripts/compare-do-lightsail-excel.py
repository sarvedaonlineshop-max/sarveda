#!/usr/bin/env python3
"""
DO MySQL (sarveda.com) vs Lightsail Postgres (sarveda-demo.xyz) — product catalog Excel.

Fresh inputs required:
  data/compare/do_products.csv
  data/compare/do_variants.csv          (from dump_do_woo.py on DO server)
  data/compare/lightsail-catalog-export.json (from export-lightsail-catalog.ts on Lightsail)

Output:
  data/compare/do-vs-lightsail-catalog.xlsx

Usage:
  python3 backend/scripts/compare-do-lightsail-excel.py
"""
from __future__ import annotations

import csv
import json
import re
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DO_PRODUCTS = ROOT / "data/compare/do_products.csv"
DO_VARIANTS = ROOT / "data/compare/do_variants.csv"
LS_EXPORT = ROOT / "data/compare/lightsail-catalog-export.json"
OUT = ROOT / "data/compare/do-vs-lightsail-catalog.xlsx"

HEADERS = [
    "DO MySQL product name",
    "Lightsail name",
    "DO MySQL variant name",
    "Equivalent variant present (Lightsail)",
    "DO SKU",
    "Lightsail SKU",
    "SKU matches Lightsail (Yes/No)",
    "Prices (INR / USD / GBP) — DO vs Lightsail sale",
    "HSN code (DO / Lightsail)",
    "Overall conclusion",
    "What mismatches exactly",
]


def norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def norm_sku(s: str) -> str:
    return norm_text(s).upper()


def parse_do_variant_name(attrs: str, title: str, product_name: str) -> str:
    if attrs:
        parts = []
        for seg in attrs.split(";"):
            if "=" in seg:
                parts.append(seg.split("=", 1)[1].strip())
        if parts:
            return " / ".join(parts)
    if title and " - " in title:
        tail = title.split(" - ", 1)[1].strip()
        if tail and norm_text(tail) != norm_text(product_name):
            return tail
    return ""


def money(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except ValueError:
        return None


def fmt_price_zone(label: str, do_sale, do_reg, ls_sale, ls_mrp) -> str:
    do_s = money(do_sale)
    do_r = money(do_reg)
    ls_s = ls_sale
    parts = []
    if do_s is not None or ls_s is not None:
        do_txt = f"₹{do_s:.0f}" if do_s is not None else "—"
        ls_txt = f"₹{ls_s:.2f}" if ls_s is not None else "—"
        if label == "INR":
            parts.append(f"INR DO {do_txt} / LS {ls_txt}")
        elif label == "USD":
            do_txt = "—"  # not in DO variant dump
            ls_txt = f"${ls_s:.2f}" if ls_s is not None else "—"
            parts.append(f"USD DO {do_txt} / LS {ls_txt}")
        elif label == "GBP":
            do_txt = "—"
            ls_txt = f"£{ls_s:.2f}" if ls_s is not None else "—"
            parts.append(f"GBP DO {do_txt} / LS {ls_txt}")
    return " | ".join(parts)


def load_do_rows() -> list[dict]:
    products = {}
    with DO_PRODUCTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            products[r["id"]] = r

    by_parent: dict[str, list] = {}
    with DO_VARIANTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            by_parent.setdefault(r["parent_id"], []).append(r)

    rows = []
    for pid, p in products.items():
        ptype = (p.get("product_type") or "").lower()
        vars_ = by_parent.get(pid, [])
        if ptype == "simple" or not vars_:
            rows.append(
                {
                    "do_product_id": pid,
                    "do_slug": p.get("slug") or "",
                    "do_product_name": p.get("name") or "",
                    "do_variant_name": "Standard",
                    "do_sku": (p.get("sku") or "").strip(),
                    "do_hsn": (p.get("hsn") or "").strip(),
                    "do_sale_price": p.get("sale_price") or "",
                    "do_regular_price": p.get("regular_price") or "",
                }
            )
            continue
        for v in vars_:
            rows.append(
                {
                    "do_product_id": pid,
                    "do_slug": v.get("parent_slug") or p.get("slug") or "",
                    "do_product_name": p.get("name") or "",
                    "do_variant_name": parse_do_variant_name(
                        v.get("attrs") or "", v.get("title") or "", p.get("name") or ""
                    )
                    or "Standard",
                    "do_sku": (v.get("sku") or "").strip(),
                    "do_hsn": (p.get("hsn") or "").strip(),
                    "do_sale_price": v.get("sale_price") or "",
                    "do_regular_price": v.get("regular_price") or "",
                }
            )
    return rows


def load_lightsail() -> dict:
    data = json.loads(LS_EXPORT.read_text(encoding="utf-8"))
    rows = data["rows"]
    by_slug: dict[str, list] = {}
    by_sku: dict[str, dict] = {}
    by_wc_id: dict[int, list] = {}
    products_by_slug = {norm_text(k): v for k, v in data.get("bySlug", {}).items()}

    for r in rows:
        slug = r.get("slug") or ""
        by_slug.setdefault(norm_text(slug), []).append(r)
        sku = norm_sku(r.get("sku") or "")
        if sku:
            by_sku[sku] = r
        wc = r.get("wooCommerceId")
        if wc:
            by_wc_id.setdefault(int(wc), []).append(r)

    return {
        "rows": rows,
        "by_slug": by_slug,
        "by_sku": by_sku,
        "by_wc_id": by_wc_id,
        "products_by_slug": products_by_slug,
        "by_name": data.get("byName") or {},
        "meta": data,
    }


def find_ls_product(ls: dict, do_slug: str, do_name: str, do_product_id: str) -> tuple[str | None, list]:
    ns = norm_text(do_slug)
    if ns and ns in ls["by_slug"]:
        return ls["by_slug"][ns][0]["slug"], ls["by_slug"][ns]

    try:
        wc = int(do_product_id)
        if wc in ls["by_wc_id"]:
            grp = ls["by_wc_id"][wc]
            return grp[0]["slug"], grp
    except ValueError:
        pass

    nn = norm_text(do_name)
    name_hits = ls["by_name"].get(nn)
    if name_hits:
        slug = name_hits[0]
        return slug, ls["by_slug"].get(norm_text(slug), [])

    # fuzzy slug / name
    best = (0.0, None)
    seen_slugs = set()
    for r in ls["rows"]:
        slug = r["slug"]
        if slug in seen_slugs:
            continue
        seen_slugs.add(slug)
        for cand in (r["name"], slug.replace("-", " ")):
            score = SequenceMatcher(None, nn, norm_text(cand)).ratio()
            if score > best[0]:
                best = (score, slug)
    if best[0] >= 0.82 and best[1]:
        return best[1], ls["by_slug"].get(norm_text(best[1]), [])
    return None, []


def find_ls_variant(candidates: list, do_sku: str, do_variant_name: str) -> dict | None:
    if not candidates:
        return None
    ds = norm_sku(do_sku)
    if ds:
        for c in candidates:
            if norm_sku(c.get("sku") or "") == ds:
                return c
        if ds in {norm_sku(c.get("sku") or "") for c in candidates}:
            pass
        # global sku index handled by caller

    dv = norm_text(do_variant_name)
    for c in candidates:
        if norm_text(c.get("variantName") or "") == dv:
            return c
    # token subset match
    d_tokens = set(re.split(r"[\s/]+", dv))
    best = (0.0, None)
    for c in candidates:
        cv = norm_text(c.get("variantName") or "")
        c_tokens = set(re.split(r"[\s/]+", cv))
        if not d_tokens or not c_tokens:
            continue
        overlap = len(d_tokens & c_tokens) / max(len(d_tokens), len(c_tokens))
        if overlap > best[0]:
            best = (overlap, c)
    if best[0] >= 0.5:
        return best[1]
    if len(candidates) == 1 and (dv in ("standard", "") or not dv):
        return candidates[0]
    return None


def compare_prices(do_sale, do_reg, ls_row) -> list[str]:
    issues = []
    if not ls_row or not ls_row.get("prices"):
        return ["Lightsail variant/prices missing"]
    ls_p = ls_row["prices"]
    do_inr = money(do_sale) if money(do_sale) is not None else money(do_reg)
    ls_inr = ls_p["inr"]["sale"]
    if do_inr is not None and ls_inr is not None and abs(do_inr - ls_inr) > 1.0:
        issues.append(f"INR sale DO ₹{do_inr:.0f} vs LS ₹{ls_inr:.2f}")
    return issues


def build_row(do: dict, ls: dict) -> dict:
    ls_slug, candidates = find_ls_product(ls, do["do_slug"], do["do_product_name"], do["do_product_id"])
    ls_var = None
    if ls_slug:
        ls_var = find_ls_variant(candidates, do["do_sku"], do["do_variant_name"])
        if not ls_var and do["do_sku"]:
            ls_var = ls["by_sku"].get(norm_sku(do["do_sku"]))

    ls_name = ls_var["name"] if ls_var else (candidates[0]["name"] if candidates else "")
    ls_variant = ls_var["variantName"] if ls_var else ""
    ls_sku = ls_var["sku"] if ls_var else ""

    sku_match = "N/A"
    if do["do_sku"] and ls_sku:
        sku_match = "Yes" if norm_sku(do["do_sku"]) == norm_sku(ls_sku) else "No"
    elif not do["do_sku"] and ls_sku:
        sku_match = "N/A (DO empty)"
    elif not do["do_sku"] and not ls_sku:
        sku_match = "N/A (both empty)"

    price_str = ""
    if ls_var and ls_var.get("prices"):
        p = ls_var["prices"]
        price_str = " | ".join(
            [
                f"INR DO ₹{money(do['do_sale_price']) or money(do['do_regular_price']) or '—'} / LS ₹{p['inr']['sale']:.2f}"
                if p["inr"]["sale"] is not None
                else "INR DO — / LS —",
                f"USD DO — / LS ${p['usd']['sale']:.2f}" if p["usd"]["sale"] is not None else "USD DO — / LS —",
                f"GBP DO — / LS £{p['gbp']['sale']:.2f}" if p["gbp"]["sale"] is not None else "GBP DO — / LS —",
            ]
        )
    else:
        do_inr = money(do["do_sale_price"]) or money(do["do_regular_price"])
        price_str = f"INR DO ₹{do_inr:.0f} / LS — | USD DO — / LS — | GBP DO — / LS —" if do_inr else "—"

    do_hsn = do["do_hsn"]
    ls_hsn = (ls_var or {}).get("hsnCode") or (candidates[0]["hsnCode"] if candidates else "") or ""
    hsn_str = f"DO: {do_hsn or '—'} / LS: {ls_hsn or '—'}"

    mismatches = []
    if not ls_slug:
        mismatches.append("Product missing on Lightsail")
    elif norm_text(ls_name) != norm_text(do["do_product_name"]):
        mismatches.append(f"Product name differs (LS: {ls_name})")
    if ls_slug and not ls_var:
        mismatches.append("Variant missing on Lightsail")
    elif ls_var and norm_text(ls_variant) != norm_text(do["do_variant_name"]):
        mismatches.append(f"Variant label differs (LS: {ls_variant})")
    if sku_match == "No":
        mismatches.append(f"SKU DO {do['do_sku'] or '—'} vs LS {ls_sku or '—'}")
    mismatches.extend(compare_prices(do["do_sale_price"], do["do_regular_price"], ls_var))
    if do_hsn and ls_hsn and norm_text(do_hsn) != norm_text(ls_hsn):
        mismatches.append(f"HSN DO {do_hsn} vs LS {ls_hsn}")
    elif do_hsn and not ls_hsn:
        mismatches.append("HSN missing on Lightsail")

    if not ls_slug:
        conclusion = "Missing on Lightsail"
    elif not ls_var:
        conclusion = "Product found; variant missing"
    elif not mismatches:
        conclusion = "Match"
    elif sku_match.startswith("N/A") and not [m for m in mismatches if "INR" in m or "HSN" in m or "missing" in m.lower() or "differs" in m.lower()]:
        conclusion = "Match (DO SKU empty on Woo)"
    elif sku_match == "No" and len(mismatches) == 1:
        conclusion = "SKU mismatch only"
    else:
        conclusion = "Partial mismatch"

    return {
        "DO MySQL product name": do["do_product_name"],
        "Lightsail name": ls_name or "—",
        "DO MySQL variant name": do["do_variant_name"],
        "Equivalent variant present (Lightsail)": ls_variant or "—",
        "DO SKU": do["do_sku"],
        "Lightsail SKU": ls_sku,
        "SKU matches Lightsail (Yes/No)": sku_match,
        "Prices (INR / USD / GBP) — DO vs Lightsail sale": price_str,
        "HSN code (DO / Lightsail)": hsn_str,
        "Overall conclusion": conclusion,
        "What mismatches exactly": "; ".join(mismatches) if mismatches else "—",
        "_slug_do": do["do_slug"],
        "_slug_ls": ls_slug or "",
    }


def autosize(ws):
    for i, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(48, max(14, len(h) + 2))


def main() -> None:
    for p in (DO_PRODUCTS, DO_VARIANTS, LS_EXPORT):
        if not p.exists():
            raise SystemExit(f"Missing {p} — refresh live dumps first")

    do_rows = load_do_rows()
    ls = load_lightsail()
    compared = [build_row(d, ls) for d in do_rows]

    wb = Workbook()
    ws = wb.active
    ws.title = "DO vs Lightsail"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(wrap_text=True, vertical="top")

    green = PatternFill("solid", fgColor="E2EFDA")
    red = PatternFill("solid", fgColor="FCE4D6")
    yellow = PatternFill("solid", fgColor="FFF2CC")

    for ri, row in enumerate(compared, 2):
        for col, h in enumerate(HEADERS, 1):
            ws.cell(row=ri, column=col, value=row.get(h, ""))
        con = row["Overall conclusion"]
        fill = green if con == "Match" else red if con == "Missing on Lightsail" else yellow
        ws.cell(row=ri, column=HEADERS.index("Overall conclusion") + 1).fill = fill

    autosize(ws)

    # Summary sheet
    sm = wb.create_sheet("Summary")
    from collections import Counter

    counts = Counter(r["Overall conclusion"] for r in compared)
    sm.append(["Metric", "Count"])
    sm.append(["DO publish variant rows", len(compared)])
    sm.append(["Lightsail product export", ls["meta"].get("productCount")])
    sm.append(["Lightsail variant rows", ls["meta"].get("variantRowCount")])
    for k, v in counts.most_common():
        sm.append([k, v])
    sm.append([])
    sm.append(["DO fetch", str(DO_PRODUCTS)])
    sm.append(["Lightsail export", str(LS_EXPORT)])
    sm.append(["Generated", OUT.name])

    # LS-only products (rough)
    matched_ls_slugs = {norm_text(r["_slug_ls"]) for r in compared if r["_slug_ls"]}
    ls_only = []
    seen = set()
    for r in ls["rows"]:
        ns = norm_text(r["slug"])
        if ns in matched_ls_slugs or ns in seen:
            continue
        seen.add(ns)
        ls_only.append(r)
    if ls_only:
        lo = wb.create_sheet("Lightsail only")
        lo.append(["Lightsail slug", "Name", "SKU", "Variant"])
        for r in ls_only:
            lo.append([r["slug"], r["name"], r.get("sku"), r.get("variantName")])

    wb.save(OUT)
    print(f"DO publish rows: {len(compared)}")
    print(json.dumps(dict(counts), indent=2))
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
