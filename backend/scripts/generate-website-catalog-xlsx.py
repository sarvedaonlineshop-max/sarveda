#!/usr/bin/env python3
"""
Export live website catalog (ACTIVE, not catalogHidden) for Zoho alignment.

Rules:
  - Product list + all prices + stock + GST/HSN → DB (website) is source of truth.
  - SKU → Zoho master when variant_id matches "Match Yes"; else DB SKU.
  - INR / UK / USD / Dirams always show DB sale prices.
  - If Zoho match exists and Zoho price ≠ DB price, cell is highlighted (fix Zoho).

Dirams (AED) = INR ÷ AED_INR rate (default 22.6) — DB has no AED field yet.

Usage:
  python3 backend/scripts/generate-website-catalog-xlsx.py
  python3 backend/scripts/generate-website-catalog-xlsx.py \\
    --zoho-master data/sarveda-sku-zoho-db-master.xlsx \\
    --output data/website-catalog-zoho.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_API = "http://13.206.192.106:5000"
DEFAULT_HSN = "9205"
DEFAULT_ZOHO_MASTER = "data/sarveda-sku-zoho-db-master.xlsx"
DEFAULT_SYNC_REPORT = "data/sarveda-zoho-db-sync-report.xlsx"
DEFAULT_AED_INR = 22.6
UNTRACKED_STOCK = 999
PRICE_MISMATCH_FILL = PatternFill("solid", fgColor="FFF2CC")

GST_RATES: dict[str, int] = {
    "standard": 18,
    "gst18": 18,
    "gst12": 12,
    "gst-5": 5,
    "gst-zero-rate": 0,
}

PRICE_FIELDS = (
    ("inr", "INR", "indian_price_inr", 4),
    ("uk", "UK", "gbp_price", 5),
    ("usd", "USD", "us_price_usd", 6),
)


def gst_rate_percent(tax_class: str | None) -> int:
    if not tax_class:
        return GST_RATES["standard"]
    key = tax_class.strip().lower()
    if key == "gst18":
        return 18
    return GST_RATES.get(key, GST_RATES["standard"])


def money_minor(minor: int | None) -> float | None:
    if minor is None:
        return None
    return round(minor / 100, 2)


def dirhams_from_inr(inr: float | None, aed_inr: float) -> float | None:
    if inr is None:
        return None
    return round(float(inr) / aed_inr, 2)


def fetch_json(url: str, timeout: int = 60) -> dict:
    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.load(resp)
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {exc.code} for {url}: {body[:400]}") from exc


def list_website_products(api_base: str) -> list[dict]:
    items: list[dict] = []
    page = 1
    while True:
        url = f"{api_base.rstrip('/')}/api/products?status=ACTIVE&limit=100&page={page}"
        payload = fetch_json(url)
        batch = payload.get("data", {}).get("items", [])
        if not batch:
            break
        items.extend(batch)
        pagination = payload.get("data", {}).get("pagination", {})
        total_pages = pagination.get("totalPages", page)
        if page >= total_pages:
            break
        page += 1
    return sorted(items, key=lambda p: (p.get("name") or "").lower())


def fetch_product_detail(api_base: str, slug: str) -> dict:
    url = f"{api_base.rstrip('/')}/api/products/{urllib.parse.quote(slug)}"
    payload = fetch_json(url)
    product = payload.get("data", {}).get("product")
    if not product:
        raise RuntimeError(f"Product not found: {slug}")
    if product.get("catalogHidden"):
        raise RuntimeError(f"Product is catalogHidden: {slug}")
    if product.get("status") != "ACTIVE":
        raise RuntimeError(f"Product is not ACTIVE: {slug}")
    return product


def fetch_all_products(api_base: str, slugs: list[str], workers: int = 8) -> list[dict]:
    products: list[dict] = []
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(fetch_product_detail, api_base, slug): slug for slug in slugs}
        for future in as_completed(futures):
            products.append(future.result())
    return sorted(products, key=lambda p: (p.get("name") or "").lower())


def read_sheet_rows(path: Path, sheet: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise RuntimeError(f"Sheet {sheet!r} not found in {path}")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for row in rows[1:]:
        if not any(c is not None and c != "" for c in row):
            continue
        out.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return out


def load_zoho_matches(path: Path) -> dict[str, dict]:
    rows = read_sheet_rows(path, "Match Yes")
    out: dict[str, dict] = {}
    for row in rows:
        vid = str(row.get("variant_id") or "").strip()
        if vid:
            out[vid] = row
    return out


def variant_label(variant: dict) -> str:
    attrs = variant.get("attributeValues") or []
    if not attrs:
        return "Standard"
    return " / ".join(row["attributeValue"]["value"] for row in attrs)


def as_number(val) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", ""))
    except ValueError:
        return None


def prices_match(db_val: float | None, zoho_val: float | None, tol: float = 0.01) -> bool:
    if db_val is None and zoho_val is None:
        return True
    if db_val is None or zoho_val is None:
        return False
    return abs(float(db_val) - float(zoho_val)) <= tol


def db_prices(variant: dict) -> dict[str, float | None]:
    return {
        "inr": money_minor(variant.get("saleInPaise")),
        "uk": money_minor(variant.get("saleGbpPence")),
        "usd": money_minor(variant.get("saleUsdCents")),
    }


def zoho_prices(zoho: dict) -> dict[str, float | None]:
    return {
        "inr": as_number(zoho.get("indian_price_inr")),
        "uk": as_number(zoho.get("gbp_price")),
        "usd": as_number(zoho.get("us_price_usd")),
    }


def resolve_sku(variant: dict, zoho: dict | None) -> tuple[str, str]:
    """Returns (sku, sku_source)."""
    if zoho:
        sku = clean_export_sku(str(zoho.get("zoho_sku") or zoho.get("db_sku") or "").strip())
        if sku:
            return sku, "Zoho"
    return clean_export_sku(str(variant.get("sku") or "").strip()), "DB"


def clean_export_sku(sku: str) -> str:
    if not sku:
        return ""
    if sku.lower().startswith("woo-var-"):
        return ""
    return sku


def strip_parent_variant(var_name: str, parent_name: str) -> str:
    var_name = (var_name or "").strip()
    parent_name = (parent_name or "").strip()
    if not var_name:
        return ""
    if parent_name and var_name.lower().startswith(parent_name.lower()):
        rest = var_name[len(parent_name) :].lstrip(" -–—|")
        return rest or var_name
    if " - " in var_name:
        return var_name.split(" - ", 1)[-1].strip()
    return var_name


def build_missing_rows(
    website_variant_ids: set[str],
    sync_report_path: Path | None,
    zoho_matches: dict[str, dict] | None,
    aed_inr: float,
) -> list[dict]:
    """Rows not on the live website: DB-only + Woo draft/orphan not imported."""
    if not sync_report_path or not sync_report_path.is_file():
        return []

    rows: list[dict] = []
    zoho_matches = zoho_matches or {}

    def append_row(**kwargs) -> None:
        inr = as_number(kwargs.get("inr"))
        rows.append(
            {
                "missing_reason": kwargs.get("missing_reason", ""),
                "name": kwargs.get("name", ""),
                "variant_name": kwargs.get("variant_name", ""),
                "sku": clean_export_sku(str(kwargs.get("sku") or "")),
                "inr": inr,
                "uk": as_number(kwargs.get("uk")),
                "usd": as_number(kwargs.get("usd")),
                "dirams": dirhams_from_inr(inr, aed_inr),
                "source": kwargs.get("source", ""),
                "variant_id": kwargs.get("variant_id", ""),
                "woo_id": kwargs.get("woo_id", ""),
                "slug": kwargs.get("slug", ""),
                "notes": kwargs.get("notes", ""),
            }
        )

    # 1) In DB but not on live website (DRAFT / hidden / internal)
    for row in read_sheet_rows(sync_report_path, "All DB Variants"):
        vid = str(row.get("variant_id") or "").strip()
        if not vid or vid in website_variant_ids:
            continue
        zoho = zoho_matches.get(vid)
        sku = ""
        if zoho:
            sku = clean_export_sku(str(zoho.get("zoho_sku") or zoho.get("db_sku") or ""))
        if not sku:
            sku = clean_export_sku(str(row.get("db_sku") or ""))
        append_row(
            missing_reason="In DB but not on live website",
            name=row.get("product_name") or "",
            variant_name=row.get("attrs") or "Standard",
            sku=sku,
            inr=row.get("inr_sale"),
            uk=row.get("gbp_sale"),
            usd=row.get("usd_sale"),
            source="DB",
            variant_id=vid,
            slug=row.get("slug") or "",
            notes="Product likely DRAFT, catalogHidden, or internal checkout SKU",
        )

    # 2) Woo draft variations (never imported)
    for row in read_sheet_rows(sync_report_path, "WC Draft Variations"):
        parent = str(row.get("wc_parent_name") or "")
        var_name = strip_parent_variant(str(row.get("wc_var_name") or ""), parent) or row.get("wc_var_name") or ""
        append_row(
            missing_reason="Woo draft variation (not imported to DB)",
            name=parent,
            variant_name=var_name,
            sku=row.get("wc_var_sku"),
            source="WooCommerce",
            woo_id=str(row.get("wc_var_woo_id") or ""),
            notes=f"parent_woo_id={row.get('wc_parent_woo_id', '')}",
        )

    # 3) Woo orphan variations (parent missing / skipped import)
    for row in read_sheet_rows(sync_report_path, "WC Orphan Variations"):
        parent = str(row.get("wc_parent_name") or "Missing parent")
        var_name = strip_parent_variant(str(row.get("wc_var_name") or ""), parent) or row.get("wc_var_name") or ""
        append_row(
            missing_reason="Woo orphan variation (not imported to DB)",
            name=parent,
            variant_name=var_name,
            sku=row.get("wc_var_sku"),
            source="WooCommerce",
            woo_id=str(row.get("wc_var_woo_id") or ""),
            notes=str(row.get("import_status") or row.get("wc_parent_status") or ""),
        )

    # 4) Woo draft parent products with no DB row
    for row in read_sheet_rows(sync_report_path, "WC Draft Products"):
        if str(row.get("in_db") or "").strip().upper() == "YES":
            continue
        append_row(
            missing_reason="Woo draft parent product (not in DB)",
            name=row.get("wc_name") or "",
            variant_name="(parent product — no variants imported)",
            sku=row.get("wc_sku"),
            source="WooCommerce",
            woo_id=str(row.get("wc_woo_id") or ""),
            notes=f"type={row.get('wc_type', '')}; status={row.get('wc_status', '')}",
        )

    rows.sort(key=lambda r: (r["missing_reason"], str(r["name"]).lower(), str(r["variant_name"]).lower()))
    return rows


def write_missing_sheet(wb, missing_rows: list[dict]) -> None:
    ws = wb.create_sheet("Missing from Website")
    headers = [
        "Missing reason",
        "Name",
        "Variant Name",
        "SKU",
        "INR",
        "UK",
        "USD",
        "Dirams",
        "Source",
        "variant_id",
        "woo_id",
        "slug",
        "notes",
    ]
    for col, title in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=title)
    style_header(ws, 1, len(headers))

    for i, row in enumerate(missing_rows, start=2):
        ws.cell(row=i, column=1, value=row.get("missing_reason"))
        ws.cell(row=i, column=2, value=row.get("name"))
        ws.cell(row=i, column=3, value=row.get("variant_name"))
        ws.cell(row=i, column=4, value=row.get("sku") or None)
        ws.cell(row=i, column=5, value=row.get("inr"))
        ws.cell(row=i, column=6, value=row.get("uk"))
        ws.cell(row=i, column=7, value=row.get("usd"))
        ws.cell(row=i, column=8, value=row.get("dirams"))
        ws.cell(row=i, column=9, value=row.get("source"))
        ws.cell(row=i, column=10, value=row.get("variant_id"))
        ws.cell(row=i, column=11, value=row.get("woo_id"))
        ws.cell(row=i, column=12, value=row.get("slug"))
        ws.cell(row=i, column=13, value=row.get("notes"))

    widths = [36, 34, 28, 18, 10, 10, 10, 10, 14, 38, 12, 24, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def stock_on_hand(variant: dict) -> int | str:
    inv = variant.get("inventory") or {}
    on_hand = inv.get("onHand")
    if on_hand is None:
        return ""
    if on_hand >= UNTRACKED_STOCK:
        return "In stock"
    return int(on_hand)


def style_header(ws, row: int, cols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_price_cell(
    ws,
    row: int,
    col: int,
    db_val: float | None,
    zoho_val: float | None,
    has_zoho: bool,
) -> bool:
    cell = ws.cell(row=row, column=col, value=db_val)
    mismatched = has_zoho and not prices_match(db_val, zoho_val)
    if mismatched:
        cell.fill = PRICE_MISMATCH_FILL
    return mismatched


def write_workbook(
    products: list[dict],
    output: Path,
    zoho_matches: dict[str, dict] | None = None,
    aed_inr: float = DEFAULT_AED_INR,
    sync_report_path: Path | None = None,
) -> dict[str, int]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Website Catalog"

    total_cols = 11
    ws.merge_cells("A1:I1")
    ws["A1"] = "Product"
    ws["J1"] = "HSN Code"
    ws["K1"] = "GST %"
    style_header(ws, 1, total_cols)

    headers = [
        "Name",
        "Variant Name",
        "SKU",
        "INR",
        "UK",
        "USD",
        "Dirams",
        "stock",
        "Included In website",
        "HSN Code",
        "GST %",
    ]
    for col, title in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=title)
    style_header(ws, 2, total_cols)

    stats = {
        "products": 0,
        "variant_rows": 0,
        "zoho_sku_rows": 0,
        "db_sku_rows": 0,
        "price_mismatches": 0,
    }
    mismatch_rows: list[dict] = []
    website_variant_ids: set[str] = set()

    row_idx = 3
    for product in products:
        name = product.get("name") or ""
        hsn = (product.get("hsnCode") or "").strip() or DEFAULT_HSN
        gst_pct = gst_rate_percent(product.get("taxClass"))
        variants = [v for v in (product.get("variants") or []) if v.get("status") == "ACTIVE"]
        if not variants:
            continue

        stats["products"] += 1
        start_row = row_idx
        for variant in variants:
            vid = str(variant.get("id") or "").strip()
            website_variant_ids.add(vid)
            zoho = (zoho_matches or {}).get(vid)
            has_zoho = zoho is not None
            db_p = db_prices(variant)
            z_p = zoho_prices(zoho) if zoho else {}

            ws.cell(row=row_idx, column=2, value=variant_label(variant))

            sku, sku_source = resolve_sku(variant, zoho)
            ws.cell(row=row_idx, column=3, value=sku or None)
            if sku_source == "Zoho":
                stats["zoho_sku_rows"] += 1
            else:
                stats["db_sku_rows"] += 1

            for key, label, _zoho_key, col in PRICE_FIELDS:
                if write_price_cell(ws, row_idx, col, db_p[key], z_p.get(key), has_zoho):
                    stats["price_mismatches"] += 1
                    mismatch_rows.append(
                        {
                            "product": name,
                            "variant": variant_label(variant),
                            "sku": sku,
                            "field": label,
                            "db_value": db_p[key],
                            "zoho_value": z_p.get(key),
                        }
                    )

            ws.cell(row=row_idx, column=7, value=dirhams_from_inr(db_p["inr"], aed_inr))
            ws.cell(row=row_idx, column=8, value=stock_on_hand(variant))
            ws.cell(row=row_idx, column=9, value="YES")

            stats["variant_rows"] += 1
            row_idx += 1

        end_row = row_idx - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=10, end_row=end_row, end_column=10)
        ws.merge_cells(start_row=start_row, start_column=11, end_row=end_row, end_column=11)
        for col, value in ((1, name), (10, hsn), (11, f"{gst_pct}%")):
            cell = ws.cell(row=start_row, column=col, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Price mismatches sheet
    mm_ws = wb.create_sheet("Price Mismatches")
    mm_headers = ["product", "variant", "sku", "field", "db_value", "zoho_value"]
    for col, title in enumerate(mm_headers, start=1):
        mm_ws.cell(row=1, column=col, value=title)
    style_header(mm_ws, 1, len(mm_headers))
    for i, row in enumerate(mismatch_rows, start=2):
        for col, key in enumerate(mm_headers, start=1):
            mm_ws.cell(row=i, column=col, value=row[key])
    mm_ws.freeze_panes = "A2"
    for col, width in enumerate([34, 28, 18, 8, 12, 12], start=1):
        mm_ws.column_dimensions[get_column_letter(col)].width = width

    missing_rows = build_missing_rows(website_variant_ids, sync_report_path, zoho_matches, aed_inr)
    write_missing_sheet(wb, missing_rows)
    missing_by_reason: dict[str, int] = {}
    for row in missing_rows:
        reason = str(row.get("missing_reason") or "")
        missing_by_reason[reason] = missing_by_reason.get(reason, 0) + 1
    stats["missing_rows"] = len(missing_rows)
    stats["missing_by_reason"] = missing_by_reason

    summary = wb.create_sheet("Summary", 0)
    summary["A1"] = "Metric"
    summary["B1"] = "Value"
    style_header(summary, 1, 2)
    summary_rows = [
        ("Website products (ACTIVE)", stats["products"]),
        ("Variant rows on website", stats["variant_rows"]),
        ("Missing from website (sheet 2)", stats.get("missing_rows", 0)),
        ("SKU from Zoho (matched)", stats["zoho_sku_rows"]),
        ("SKU from DB (no Zoho match)", stats["db_sku_rows"]),
        ("Price cells highlighted (DB ≠ Zoho)", stats["price_mismatches"]),
        ("Prices source of truth", "DB (website)"),
        ("SKU source when matched", "Zoho"),
        ("Dirams formula", f"INR ÷ {aed_inr} (AED)"),
        ("Highlighted cells", "Yellow = fix price in Zoho"),
        ("Woo original sellable SKUs (approx)", 1069),
    ]
    for reason, count in sorted(stats.get("missing_by_reason", {}).items()):
        summary_rows.append((f"  ↳ {reason}", count))
    for i, (label, val) in enumerate(summary_rows, start=2):
        summary.cell(row=i, column=1, value=label)
        summary.cell(row=i, column=2, value=val)
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 28

    widths = [34, 28, 18, 10, 10, 10, 10, 12, 18, 12, 8]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A3"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return stats


def main() -> int:
    parser = argparse.ArgumentParser(description="Export website catalog to Zoho-alignment xlsx")
    parser.add_argument("--api-base", default=DEFAULT_API, help="Backend API base URL")
    parser.add_argument("--output", default="data/website-catalog-zoho.xlsx", help="Output xlsx path")
    parser.add_argument(
        "--zoho-master",
        default=DEFAULT_ZOHO_MASTER,
        help="SKU master xlsx with Match Yes sheet (blank to skip Zoho SKU)",
    )
    parser.add_argument(
        "--sync-report",
        default=DEFAULT_SYNC_REPORT,
        help="Sync report xlsx for Missing from Website sheet (blank to skip)",
    )
    parser.add_argument("--aed-inr", type=float, default=DEFAULT_AED_INR, help="INR per 1 AED for Dirams")
    parser.add_argument("--sample", type=int, default=0, help="Export only first N products")
    parser.add_argument("--product-slug", default="", help="Export a single product slug")
    parser.add_argument("--workers", type=int, default=8, help="Parallel API fetch workers")
    args = parser.parse_args()

    zoho_matches: dict[str, dict] | None = None
    if args.zoho_master.strip():
        master_path = Path(args.zoho_master)
        if not master_path.is_file():
            print(f"Zoho master not found: {master_path}", file=sys.stderr)
            return 1
        zoho_matches = load_zoho_matches(master_path)
        print(f"Loaded Match Yes rows: {len(zoho_matches)}")

    sync_report_path: Path | None = None
    if args.sync_report.strip():
        sync_path = Path(args.sync_report)
        if not sync_path.is_file():
            print(f"Sync report not found: {sync_path}", file=sys.stderr)
            return 1
        sync_report_path = sync_path

    if args.product_slug:
        products = [fetch_product_detail(args.api_base, args.product_slug)]
    elif args.sample > 0:
        listed = list_website_products(args.api_base)
        if not listed:
            print("No ACTIVE products found.", file=sys.stderr)
            return 1
        slugs = [p["slug"] for p in listed[: args.sample]]
        products = fetch_all_products(args.api_base, slugs, workers=args.workers)
    else:
        listed = list_website_products(args.api_base)
        if not listed:
            print("No ACTIVE products found.", file=sys.stderr)
            return 1
        slugs = [p["slug"] for p in listed]
        print(f"Fetching {len(slugs)} products from API…")
        products = fetch_all_products(args.api_base, slugs, workers=args.workers)

    output = Path(args.output)
    stats = write_workbook(
        products,
        output,
        zoho_matches,
        aed_inr=args.aed_inr,
        sync_report_path=sync_report_path,
    )

    print(f"Wrote {output}")
    print(
        f"Products: {stats['products']} | Website variants: {stats['variant_rows']} | "
        f"Missing: {stats.get('missing_rows', 0)} | "
        f"Zoho SKU: {stats['zoho_sku_rows']} | DB SKU: {stats['db_sku_rows']} | "
        f"Price mismatches: {stats['price_mismatches']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
