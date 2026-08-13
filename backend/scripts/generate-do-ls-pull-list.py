#!/usr/bin/env python3
"""
Build DO→Lightsail media pull list from fuzzy compare + user MATCHED ONES sheet.

Reads:
  data/compare/do-vs-lightsail-fuzzy.xlsx  (incl. user sheet "MATCHED ONES")
  data/compare/do_products.csv
  data/compare/do_variants.csv
  data/compare/lightsail-catalog-export.json

Writes:
  data/compare/do-ls-pull-list.xlsx

Usage:
  python3 backend/scripts/generate-do-ls-pull-list.py
"""
from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

ROOT = Path(__file__).resolve().parents[2]
FUZZY_XLSX = ROOT / "data/compare/do-vs-lightsail-fuzzy.xlsx"
DO_PRODUCTS = ROOT / "data/compare/do_products.csv"
DO_VARIANTS = ROOT / "data/compare/do_variants.csv"
LS_EXPORT = ROOT / "data/compare/lightsail-catalog-export.json"
OUT = ROOT / "data/compare/do-ls-pull-list.xlsx"

VARIANT_MATCH_MIN = 85
HEADERS = [
    "Lightsail Product name",
    "DO DB product name",
    "Lightsail variant name",
    "DO variant name",
    "Lightsail SKU",
    "DO SKU",
    "Note on mismatch",
    "Match source",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN = PatternFill("solid", fgColor="E2EFDA")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GRAY = PatternFill("solid", fgColor="EDEDED")


@dataclass
class VariantRow:
    name: str
    sku: str = ""


@dataclass
class ProductRow:
    slug: str
    name: str
    variants: list[VariantRow] = field(default_factory=list)


@dataclass
class PullRow:
    ls_product: str
    do_product: str
    ls_variant: str
    do_variant: str
    ls_sku: str
    do_sku: str
    note: str
    source: str

    def key(self) -> tuple:
        return (
            norm_text(self.ls_product),
            norm_variant(self.ls_variant),
            norm_text(self.do_product),
            norm_variant(self.do_variant),
            norm_sku(self.ls_sku),
        )


def norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def norm_variant(s: str) -> str:
    s = norm_text(s)
    return re.sub(r"[\s|/·,–—\-]+", " ", s).strip()


def norm_sku(s: str) -> str:
    return norm_text(s).upper()


def variant_exact(a: str, b: str) -> bool:
    na, nb = norm_variant(a), norm_variant(b)
    if na == nb:
        return True
    if na and nb and set(na.split()) == set(nb.split()):
        return True
    return False


def variant_score(a: str, b: str) -> float:
    if variant_exact(a, b):
        return 100.0
    return float(fuzz.token_sort_ratio(norm_variant(a), norm_variant(b)))


def is_copper_water_bottle(name: str, slug: str = "") -> bool:
    n = norm_text(name)
    s = (slug or "").lower()
    if "tongue cleaner" in n:
        return False
    if "bedroom jar" in n:
        return False
    if "copper bottle" in n or "copper water" in n:
        return True
    if s.startswith("copper-bottle") or "copper-water" in s:
        return True
    if "chakras copper" in n and "bottle" in n:
        return True
    if "artistically designed copper" in n:
        return True
    if "diamond groove copper" in n:
        return True
    if "engraved-copper" in s or "printed-copper" in s:
        return True
    return False


def parse_do_variant_name(attrs: str, title: str, product_name: str) -> str:
    if attrs:
        parts = []
        for seg in attrs.split(";"):
            if "=" in seg:
                parts.append(seg.split("=", 1)[1].strip())
        if parts:
            return " / ".join(parts)
    if title and " - " in title:
        tail = title.split(" - ", 1)[1].strip()
        if tail and norm_text(tail) != norm_text(product_name):
            return tail
    return "Standard"


def load_do_by_name() -> dict[str, ProductRow]:
    products: dict[str, dict] = {}
    with DO_PRODUCTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            products[r["id"]] = r

    by_parent: dict[str, list] = {}
    with DO_VARIANTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            by_parent.setdefault(r["parent_id"], []).append(r)

    out: dict[str, ProductRow] = {}
    for pid, p in products.items():
        name = p.get("name") or ""
        slug = p.get("slug") or ""
        ptype = (p.get("product_type") or "").lower()
        vars_ = by_parent.get(pid, [])
        variants: list[VariantRow] = []
        if ptype == "simple" or not vars_:
            variants.append(VariantRow(name="Standard", sku=(p.get("sku") or "").strip()))
        else:
            for v in vars_:
                variants.append(
                    VariantRow(
                        name=parse_do_variant_name(v.get("attrs") or "", v.get("title") or "", name),
                        sku=(v.get("sku") or "").strip(),
                    )
                )
        out[norm_text(name)] = ProductRow(slug=slug, name=name, variants=variants)
    return out


def load_ls_by_name() -> dict[str, ProductRow]:
    data = json.loads(LS_EXPORT.read_text(encoding="utf-8"))
    grouped: dict[str, ProductRow] = {}
    for r in data["rows"]:
        name = r.get("name") or ""
        slug = r.get("slug") or ""
        key = norm_text(name)
        if key not in grouped:
            grouped[key] = ProductRow(slug=slug, name=name, variants=[])
        vname = (r.get("variantName") or "").strip() or "Standard"
        grouped[key].variants.append(VariantRow(name=vname, sku=(r.get("sku") or "").strip()))
    return grouped


def parse_matched_ones() -> tuple[list[tuple[str, str]], dict[str, str], set[str], set[str]]:
    """Returns manual_pairs, exclude_reason_by_key, excluded_ls_names, excluded_do_names."""
    wb = load_workbook(FUZZY_XLSX, read_only=True, data_only=True)
    ws = wb["MATCHED ONES"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    manual_pairs: list[tuple[str, str]] = []
    exclude: dict[str, str] = {}
    excluded_ls: set[str] = set()
    excluded_do: set[str] = set()
    in_new_section = False

    for do_name, ls_note in rows:
        do_name = (do_name or "").strip()
        ls_note = (ls_note or "").strip()
        if do_name.startswith("---") or do_name.startswith("---"):
            in_new_section = True
            if ls_note:
                excluded_ls.add(norm_text(ls_note.split("(")[0].strip()))
                exclude[f"ls:{norm_text(ls_note)}"] = "New Lightsail-only product (ignore for DO pull)"
            continue
        if in_new_section:
            if ls_note:
                excluded_ls.add(norm_text(ls_note))
                exclude[f"ls:{norm_text(ls_note)}"] = "New Lightsail-only product (ignore for DO pull)"
            continue
        if not do_name:
            continue

        note_lower = ls_note.lower()
        if "ignore for now" in note_lower:
            excluded_do.add(norm_text(do_name))
            exclude[f"do:{norm_text(do_name)}"] = "Copper bottles — ignore for now"
            continue
        if "not selling" in note_lower:
            excluded_do.add(norm_text(do_name))
            exclude[f"do:{norm_text(do_name)}"] = "Not selling on new website"
            continue
        if "split into different products" in note_lower or "artistically designed copper" in do_name.lower():
            excluded_do.add(norm_text(do_name))
            exclude[f"do:{norm_text(do_name)}"] = "Copper bottles split on Lightsail — ignore for now"
            continue
        if "named as" in note_lower and "7 chakra" in note_lower:
            manual_pairs.append((do_name, "7 Chakras Yoga Mats"))
            continue
        if ls_note:
            manual_pairs.append((do_name, ls_note))

    wb.close()
    return manual_pairs, exclude, excluded_ls, excluded_do


def find_product(by_name: dict[str, ProductRow], name: str) -> ProductRow | None:
    key = norm_text(name)
    if key in by_name:
        return by_name[key]
    best = (0.0, None)
    for k, p in by_name.items():
        s = float(fuzz.token_sort_ratio(key, k))
        if s > best[0]:
            best = (s, p)
    return best[1] if best[0] >= 88 else None


def sku_note(ls_sku: str, do_sku: str) -> str:
    if not ls_sku and not do_sku:
        return ""
    if not ls_sku or not do_sku:
        return "SKU missing on one side"
    if norm_sku(ls_sku) != norm_sku(do_sku):
        return f"SKU differs (LS {ls_sku} vs DO {do_sku})"
    return ""


def match_variants_for_pair(
    ls: ProductRow,
    do: ProductRow,
    source: str,
) -> list[PullRow]:
    rows: list[PullRow] = []
    used_d: set[int] = set()

    for lv in ls.variants:
        best_i, best_s = None, 0.0
        for i, dv in enumerate(do.variants):
            if i in used_d:
                continue
            s = variant_score(lv.name, dv.name)
            if s > best_s:
                best_s, best_i = s, i
        if best_i is not None and (best_s >= VARIANT_MATCH_MIN or variant_exact(lv.name, do.variants[best_i].name)):
            dv = do.variants[best_i]
            used_d.add(best_i)
            mtype = "exact variant name" if variant_exact(lv.name, dv.name) else f"fuzzy variant name ({best_s:.0f}%)"
            notes = [mtype]
            sn = sku_note(lv.sku, dv.sku)
            if sn:
                notes.append(sn)
            rows.append(
                PullRow(
                    ls_product=ls.name,
                    do_product=do.name,
                    ls_variant=lv.name,
                    do_variant=dv.name,
                    ls_sku=lv.sku,
                    do_sku=dv.sku,
                    note="; ".join(notes),
                    source=source,
                )
            )
        else:
            rows.append(
                PullRow(
                    ls_product=ls.name,
                    do_product=do.name,
                    ls_variant=lv.name,
                    do_variant="—",
                    ls_sku=lv.sku,
                    do_sku="",
                    note="DO variant not found for this Lightsail variant",
                    source=source,
                )
            )

    for i, dv in enumerate(do.variants):
        if i in used_d:
            continue
        rows.append(
            PullRow(
                ls_product=ls.name,
                do_product=do.name,
                ls_variant="—",
                do_variant=dv.name,
                ls_sku="",
                do_sku=dv.sku,
                note="Extra DO variant — not on Lightsail",
                source=source,
            )
        )
    return rows


def load_fuzzy_variant_rows() -> list[PullRow]:
    wb = load_workbook(FUZZY_XLSX, read_only=True, data_only=True)
    ws = wb["Fuzzy Matched Variants"]
    out: list[PullRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        ls_p, do_p, _score, ls_v, do_v, mtype, vscore, ls_sku, do_sku, sku_match, balance = row[:11]
        ls_v = ls_v or "—"
        do_v = do_v or "—"
        ls_sku = ls_sku or ""
        do_sku = do_sku or ""

        notes: list[str] = []
        if mtype == "exact":
            notes.append("Variant names match")
        elif mtype == "fuzzy":
            notes.append(f"Variant names fuzzy match ({vscore}%)")
        elif mtype == "ls_only":
            notes.append("Lightsail variant only — no DO equivalent")
        elif mtype == "do_only":
            notes.append("DO variant only — not on Lightsail")
        if sku_match == "No":
            notes.append(f"SKU differs (LS {ls_sku} vs DO {do_sku})")
        elif sku_match == "N/A" and (ls_sku or do_sku):
            notes.append("SKU missing on DO or Lightsail")
        if balance and str(balance) not in ("exact variant parity",):
            notes.append(str(balance))

        out.append(
            PullRow(
                ls_product=str(ls_p),
                do_product=str(do_p),
                ls_variant=str(ls_v),
                do_variant=str(do_v),
                ls_sku=str(ls_sku) if ls_sku else "",
                do_sku=str(do_sku) if do_sku else "",
                note="; ".join(notes) if notes else "Ready for media pull",
                source="fuzzy",
            )
        )
    wb.close()
    return out


def should_exclude_row(row: PullRow, excluded_ls: set[str], excluded_do: set[str], ls_by_name: dict, do_by_name: dict) -> str | None:
    if is_copper_water_bottle(row.ls_product) or is_copper_water_bottle(row.do_product):
        return "Copper water bottles — ignore for now"
    if norm_text(row.ls_product) in excluded_ls:
        return "New Lightsail-only product"
    if norm_text(row.do_product) in excluded_do:
        return "Excluded DO product (not selling / ignore)"
    ls = ls_by_name.get(norm_text(row.ls_product))
    do = do_by_name.get(norm_text(row.do_product))
    if ls and is_copper_water_bottle(ls.name, ls.slug):
        return "Copper water bottles — ignore for now"
    if do and is_copper_water_bottle(do.name, do.slug):
        return "Copper water bottles — ignore for now"
    return None


def write_sheet(ws, headers: list[str], rows: list[list], fills: dict[int, PatternFill] | None = None):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
    for ri, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=col, value=val)
            if fills and col in fills:
                cell.fill = fills[col]
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(52, max(14, len(h) + 2))


def main() -> None:
    for p in (FUZZY_XLSX, DO_PRODUCTS, DO_VARIANTS, LS_EXPORT):
        if not p.exists():
            raise SystemExit(f"Missing {p}")

    manual_pairs, exclude_meta, excluded_ls, excluded_do = parse_matched_ones()
    ls_by_name = load_ls_by_name()
    do_by_name = load_do_by_name()

    # Also exclude LS copper bottle products by slug/name scan
    for p in ls_by_name.values():
        if is_copper_water_bottle(p.name, p.slug):
            excluded_ls.add(norm_text(p.name))

    pull: dict[tuple, PullRow] = {}
    excluded_rows: list[list] = []
    ls_products_with_fuzzy: set[str] = set()

    def add_row(row: PullRow, reason_override: str | None = None):
        reason = reason_override or should_exclude_row(row, excluded_ls, excluded_do, ls_by_name, do_by_name)
        if reason:
            excluded_rows.append(
                [
                    row.ls_product,
                    row.do_product,
                    row.ls_variant,
                    row.do_variant,
                    row.ls_sku,
                    row.do_sku,
                    row.note,
                    row.source,
                    reason,
                ]
            )
            return
        pull[row.key()] = row
        if row.source == "fuzzy":
            ls_products_with_fuzzy.add(norm_text(row.ls_product))

    # 1) Fuzzy matched variant rows (bulk)
    for row in load_fuzzy_variant_rows():
        add_row(row)

    # 2) Manual pairs — only when Lightsail product is not already in fuzzy pull list
    for do_name, ls_name in manual_pairs:
        ls_p = find_product(ls_by_name, ls_name)
        if ls_p and norm_text(ls_p.name) in ls_products_with_fuzzy:
            continue
        do_p = find_product(do_by_name, do_name)
        if not do_p:
            excluded_rows.append([ls_name, do_name, "—", "—", "", "", "", "manual", f"DO product not found: {do_name}"])
            continue
        if not ls_p:
            excluded_rows.append([ls_name, do_name, "—", "—", "", "", "", "manual", f"Lightsail product not found: {ls_name}"])
            continue
        for row in match_variants_for_pair(ls_p, do_p, "manual (MATCHED ONES)"):
            add_row(row)

    final = sorted(pull.values(), key=lambda r: (r.ls_product.lower(), r.ls_variant.lower()))

    wb = Workbook()
    pl = wb.active
    pl.title = "Pull List"
    pl_rows = [
        [r.ls_product, r.do_product, r.ls_variant, r.do_variant, r.ls_sku, r.do_sku, r.note, r.source]
        for r in final
    ]
    write_sheet(pl, HEADERS, pl_rows)

    note_col = HEADERS.index("Note on mismatch") + 1
    for ri in range(2, pl.max_row + 1):
        note = str(pl.cell(row=ri, column=note_col).value or "")
        if "only" in note.lower() or "not found" in note.lower() or "differs" in note.lower() or "extra" in note.lower():
            pl.cell(row=ri, column=note_col).fill = YELLOW
        elif note.startswith("Variant names match") or note == "Ready for media pull":
            pl.cell(row=ri, column=note_col).fill = GREEN

    ex = wb.create_sheet("Excluded")
    ex_headers = HEADERS + ["Exclude reason"]
    write_sheet(ex, ex_headers, excluded_rows)
    reason_col = len(ex_headers)
    for ri in range(2, ex.max_row + 1):
        ex.cell(row=ri, column=reason_col).fill = GRAY

    sm = wb.create_sheet("Summary")
    matched_variants = sum(1 for r in final if r.ls_variant != "—" and r.do_variant != "—")
    ready_pull = sum(1 for r in final if r.do_variant != "—" and "only" not in r.note.lower())
    sm_data = [
        ["Generated at (UTC)", datetime.now(timezone.utc).isoformat()],
        ["Pull list variant rows", len(final)],
        ["Rows with both LS + DO variant", matched_variants],
        ["Rows ready for DO media pull", ready_pull],
        ["Excluded rows", len(excluded_rows)],
        ["Manual pairs from MATCHED ONES", len(manual_pairs)],
        ["Copper water bottles", "Excluded per user"],
        ["New LS-only products", "Excluded per MATCHED ONES sheet"],
        ["Not selling on new site", "Excluded per MATCHED ONES sheet"],
    ]
    write_sheet(sm, ["Metric", "Value"], sm_data)

    wb.save(OUT)
    print(f"Pull list rows: {len(final)}")
    print(f"Excluded rows: {len(excluded_rows)}")
    print(f"Both-side variant matches: {matched_variants}")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
