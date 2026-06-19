#!/usr/bin/env python3
"""
Build a master SKU alignment workbook from data/sarveda-zoho-db-sync-report.xlsx.

Includes: all DB variants, Zoho-only SKUs, WC draft variations, WC orphan variations.

Usage:
  python3 backend/scripts/generate-sku-master-report.py
  python3 backend/scripts/generate-sku-master-report.py --input data/sarveda-zoho-db-sync-report.xlsx --output data/sarveda-sku-zoho-db-master.xlsx
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def read_sheet(wb: openpyxl.Workbook, name: str) -> list[dict]:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for row in rows[1:]:
        if not any(c is not None and c != "" for c in row):
            continue
        out.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return out


def norm_sku(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def fmt_price(val) -> str | float | int | None:
    if val is None or val == "":
        return ""
    if isinstance(val, (int, float)):
        return val
    try:
        return float(str(val).replace(",", ""))
    except ValueError:
        return str(val)


def strip_parent_from_var_name(var_name: str, parent_name: str) -> str:
    var_name = (var_name or "").strip()
    parent_name = (parent_name or "").strip()
    if not var_name:
        return ""
    if parent_name and var_name.lower().startswith(parent_name.lower()):
        rest = var_name[len(parent_name) :].lstrip(" -–—|")
        return rest or var_name
    if " - " in var_name:
        return var_name.split(" - ", 1)[-1].strip()
    return var_name


def zoho_variant_hint(zoho_name: str, product_name: str) -> str:
    zn = (zoho_name or "").strip()
    if not zn:
        return ""
    if " | " in zn:
        part = zn.split(" | ")[-1].strip()
        if part:
            return part
    pn = (product_name or "").strip()
    if zn.lower() == pn.lower():
        return ""
    cleaned = zn
    for prefix in ("Sarveda ", "sarveda "):
        if cleaned.startswith(prefix):
            cleaned = cleaned[len(prefix) :].strip()
    if pn and cleaned.lower().startswith(pn.lower()):
        rest = cleaned[len(pn) :].lstrip(" -–—|")
        if rest:
            return rest
    if cleaned.lower() != pn.lower():
        return cleaned
    return ""


def build_wc_indexes(wc_rows: list[dict]) -> tuple[dict[str, dict], dict[str, dict]]:
    by_sku: dict[str, dict] = {}
    by_woo_id: dict[str, dict] = {}
    for row in wc_rows:
        sku = norm_sku(row.get("wc_var_sku"))
        woo_id = str(row.get("wc_var_woo_id") or "").strip()
        if sku:
            by_sku[sku] = row
        if woo_id:
            by_woo_id[woo_id] = row
    return by_sku, by_woo_id


def derive_variant_name(
    *,
    product_name: str,
    db_sku: str,
    attrs,
    zoho_name: str = "",
    wc_by_sku: dict[str, dict] | None = None,
    wc_by_woo_id: dict[str, dict] | None = None,
) -> tuple[str, str]:
    """
    Return (variant_name, source).
    DB has no variant title column — attrs is usually empty. We derive from WC / Zoho.
    """
    wc_by_sku = wc_by_sku or {}
    wc_by_woo_id = wc_by_woo_id or {}

    if attrs not in (None, "", "[]", []):
        if isinstance(attrs, str) and attrs.strip().startswith("["):
            return attrs, "db_attrs"
        return str(attrs), "db_attrs"

    db_sku = norm_sku(db_sku)
    wc_row = wc_by_sku.get(db_sku)
    if not wc_row:
        m = re.match(r"^woo-var-(\d+)$", db_sku, re.I)
        if m:
            wc_row = wc_by_woo_id.get(m.group(1))

    if wc_row:
        label = strip_parent_from_var_name(
            str(wc_row.get("wc_var_name") or ""),
            str(wc_row.get("wc_parent_name") or ""),
        )
        pn_lower = (product_name or "").strip().lower()
        if label and label.lower() != pn_lower:
            return label, "woocommerce"

    zoho_hint = zoho_variant_hint(zoho_name, product_name)
    if zoho_hint:
        return zoho_hint, "zoho_name"

    if wc_row:
        label = strip_parent_from_var_name(
            str(wc_row.get("wc_var_name") or ""),
            str(wc_row.get("wc_parent_name") or ""),
        )
        if label:
            return label, "woocommerce"

    pn = (product_name or "").strip()
    if pn:
        return "(single variant — no size/type in DB yet)", "placeholder"
    return "", "none"


def sku_match(db_sku: str, zoho_sku: str) -> str:
    if not db_sku or not zoho_sku:
        return "No"
    return "Yes" if db_sku.strip().upper() == zoho_sku.strip().upper() else "No"


def autosize(ws, max_width: int = 48) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        width = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            width = max(width, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = max(width, 10)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input",
        default="data/sarveda-zoho-db-sync-report.xlsx",
        help="Source sync report xlsx",
    )
    parser.add_argument(
        "--output",
        default="data/sarveda-sku-zoho-db-master.xlsx",
        help="Output master workbook path",
    )
    args = parser.parse_args()

    root = Path(__file__).resolve().parents[2]
    input_path = root / args.input
    output_path = root / args.output

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    all_db = read_sheet(wb, "All DB Variants")
    matched = read_sheet(wb, "Zoho-DB Matched")
    needs_update = read_sheet(wb, "Zoho Needs SKU Update")
    zoho_unmatched = read_sheet(wb, "Zoho Unmatched")
    zoho_all = read_sheet(wb, "All Zoho SKUs")
    wc_draft_vars = read_sheet(wb, "WC Draft Variations")
    wc_orphans = read_sheet(wb, "WC Orphan Variations")
    wc_draft_products = read_sheet(wb, "WC Draft Products")
    wc_vs_db = read_sheet(wb, "WC vs DB Products")
    wc_published = read_sheet(wb, "WC Published Variations")
    wb.close()

    wc_all = wc_published + wc_draft_vars + wc_orphans
    wc_by_sku, wc_by_woo_id = build_wc_indexes(wc_all)

    matched_by_db_sku = {norm_sku(r.get("db_sku")): r for r in matched}
    needs_by_variant_id = {str(r.get("variant_id")): r for r in needs_update}
    zoho_by_sku = {norm_sku(r.get("zoho_sku")): r for r in zoho_all}
    db_skus = {norm_sku(r.get("db_sku")) for r in all_db if norm_sku(r.get("db_sku"))}
    needs_zoho_skus = {norm_sku(r.get("zoho_sku")) for r in needs_update if norm_sku(r.get("zoho_sku"))}
    wc_product_status = {str(r.get("wc_woo_id")): r for r in wc_vs_db}

    master_rows: list[dict] = []

    # --- 1. All DB variants (ACTIVE, DRAFT, ARCHIVED, woo placeholders) ---
    for row in all_db:
        db_sku = norm_sku(row.get("db_sku"))
        vid = str(row.get("variant_id") or "")
        product_name = row.get("product_name") or ""
        zoho_sku = ""
        zoho_name = row.get("zoho_name") or ""

        if str(row.get("zoho_sku_match", "")).upper() == "YES":
            zoho_sku = db_sku
            m = matched_by_db_sku.get(db_sku)
            if m:
                zoho_name = m.get("zoho_name") or zoho_name
        elif vid in needs_by_variant_id:
            nu = needs_by_variant_id[vid]
            zoho_sku = norm_sku(nu.get("zoho_sku"))
            zoho_name = nu.get("zoho_name") or zoho_name
        elif db_sku in zoho_by_sku:
            zoho_sku = db_sku
            zoho_name = zoho_by_sku[db_sku].get("zoho_name") or zoho_name

        variant_name, variant_source = derive_variant_name(
            product_name=product_name,
            db_sku=db_sku,
            attrs=row.get("attrs"),
            zoho_name=zoho_name,
            wc_by_sku=wc_by_sku,
            wc_by_woo_id=wc_by_woo_id,
        )

        master_rows.append(
            {
                "record_source": "DB",
                "product_status": "ACTIVE/IN_DB",
                "product_name": product_name,
                "variant_name": variant_name,
                "variant_name_source": variant_source,
                "db_sku": db_sku,
                "zoho_product_name": zoho_name,
                "zoho_sku": zoho_sku,
                "indian_price_inr": fmt_price(row.get("inr_sale")),
                "us_price_usd": fmt_price(row.get("usd_sale")),
                "gbp_price": fmt_price(row.get("gbp_sale")),
                "zoho_db_sku_match": sku_match(db_sku, zoho_sku),
                "slug": row.get("slug") or "",
                "variant_id": vid,
                "is_woo_placeholder_sku": row.get("is_woo_placeholder") or "",
                "notes": "",
            }
        )

    used_zoho_skus = {norm_sku(r["zoho_sku"]) for r in master_rows if norm_sku(r["zoho_sku"])}

    # --- 2. Zoho SKUs with no DB variant row (not exact-matched, not on needs-update) ---
    for row in zoho_unmatched:
        zoho_sku = norm_sku(row.get("zoho_sku"))
        if not zoho_sku or zoho_sku in used_zoho_skus or zoho_sku in needs_zoho_skus:
            continue
        zoho_name = row.get("zoho_name") or ""
        master_rows.append(
            {
                "record_source": "Zoho only",
                "product_status": "ZOHO_ONLY",
                "product_name": "",
                "variant_name": zoho_name,
                "variant_name_source": "zoho_only",
                "db_sku": "",
                "zoho_product_name": zoho_name,
                "zoho_sku": zoho_sku,
                "indian_price_inr": "",
                "us_price_usd": "",
                "gbp_price": "",
                "zoho_db_sku_match": "No",
                "slug": "",
                "variant_id": "",
                "is_woo_placeholder_sku": "",
                "notes": f"in_wc_csv={row.get('in_wc_csv', '')}",
            }
        )
        used_zoho_skus.add(zoho_sku)

    # --- 3. WC draft variations (parent often draft / not imported to DB) ---
    for row in wc_draft_vars:
        wc_sku = norm_sku(row.get("wc_var_sku"))
        if wc_sku and wc_sku in db_skus:
            continue
        parent_name = row.get("wc_parent_name") or ""
        var_name = row.get("wc_var_name") or ""
        zoho_sku = wc_sku if wc_sku in zoho_by_sku else ""
        zoho_name = zoho_by_sku[zoho_sku].get("zoho_name", "") if zoho_sku else ""
        parent_id = str(row.get("wc_parent_woo_id") or "")
        parent_meta = wc_product_status.get(parent_id, {})
        var_label, var_src = derive_variant_name(
            product_name=parent_name,
            db_sku=wc_sku or f"wc-var-{row.get('wc_var_woo_id', '')}",
            attrs=None,
            zoho_name=zoho_name,
            wc_by_sku=wc_by_sku,
            wc_by_woo_id=wc_by_woo_id,
        )
        if not var_label:
            var_label = strip_parent_from_var_name(var_name, parent_name) or var_name
            var_src = "woocommerce"
        master_rows.append(
            {
                "record_source": "WC draft variation",
                "product_status": parent_meta.get("wc_status") or "Draft",
                "product_name": parent_name,
                "variant_name": var_label,
                "variant_name_source": var_src,
                "db_sku": wc_sku or f"wc-var-{row.get('wc_var_woo_id', '')}",
                "zoho_product_name": zoho_name,
                "zoho_sku": zoho_sku,
                "indian_price_inr": "",
                "us_price_usd": "",
                "gbp_price": "",
                "zoho_db_sku_match": sku_match(wc_sku, zoho_sku) if wc_sku and zoho_sku else "No",
                "slug": "",
                "variant_id": "",
                "is_woo_placeholder_sku": "WC_DRAFT",
                "notes": f"wc_var_id={row.get('wc_var_woo_id', '')}; parent_wc_id={parent_id}",
            }
        )

    # --- 4. WC orphan variations (skipped at import) ---
    for row in wc_orphans:
        wc_sku = norm_sku(row.get("wc_var_sku"))
        parent_name = row.get("wc_parent_name") or row.get("wc_parent_woo_id") or ""
        var_name = row.get("wc_var_name") or ""
        zoho_sku = wc_sku if wc_sku and wc_sku in zoho_by_sku else ""
        zoho_name = zoho_by_sku[zoho_sku].get("zoho_name", "") if zoho_sku else ""
        var_label = strip_parent_from_var_name(var_name, parent_name if isinstance(parent_name, str) else "") or var_name
        master_rows.append(
            {
                "record_source": "WC orphan variation",
                "product_status": row.get("wc_parent_status") or "ORPHAN",
                "product_name": parent_name if parent_name else "Missing parent",
                "variant_name": var_label,
                "variant_name_source": "woocommerce",
                "db_sku": wc_sku or f"wc-orphan-{row.get('wc_var_woo_id', '')}",
                "zoho_product_name": zoho_name,
                "zoho_sku": zoho_sku,
                "indian_price_inr": "",
                "us_price_usd": "",
                "gbp_price": "",
                "zoho_db_sku_match": sku_match(wc_sku, zoho_sku) if wc_sku and zoho_sku else "No",
                "slug": "",
                "variant_id": "",
                "is_woo_placeholder_sku": "WC_ORPHAN",
                "notes": row.get("import_status") or "",
            }
        )

    # --- 5. WC draft products without any DB product row ---
    for row in wc_draft_products:
        if str(row.get("in_db", "")).upper() == "YES":
            continue
        wc_sku = norm_sku(row.get("wc_sku"))
        product_name = row.get("wc_name") or ""
        zoho_sku = wc_sku if wc_sku and wc_sku in zoho_by_sku else ""
        zoho_name = zoho_by_sku[zoho_sku].get("zoho_name", "") if zoho_sku else ""
        master_rows.append(
            {
                "record_source": "WC draft product",
                "product_status": row.get("wc_status") or "Draft",
                "product_name": product_name,
                "variant_name": "(parent product — no variants imported)",
                "variant_name_source": "woocommerce",
                "db_sku": wc_sku or f"wc-product-{row.get('wc_woo_id', '')}",
                "zoho_product_name": zoho_name,
                "zoho_sku": zoho_sku,
                "indian_price_inr": "",
                "us_price_usd": "",
                "gbp_price": "",
                "zoho_db_sku_match": sku_match(wc_sku, zoho_sku) if wc_sku and zoho_sku else "No",
                "slug": "",
                "variant_id": "",
                "is_woo_placeholder_sku": "WC_DRAFT_PARENT",
                "notes": f"wc_product_id={row.get('wc_woo_id', '')}; type={row.get('wc_type', '')}",
            }
        )

    # Sort: matched first, then product name
    match_rank = {"Yes": 0, "No": 1}
    master_rows.sort(
        key=lambda r: (
            match_rank.get(r["zoho_db_sku_match"], 2),
            str(r["product_name"]).lower(),
            str(r["variant_name"]).lower(),
            str(r["db_sku"]).lower(),
        )
    )

    # --- Write workbook ---
    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "SKU Master"

    headers = [
        "product_name",
        "variant_name",
        "variant_name_source",
        "db_sku",
        "zoho_product_name",
        "zoho_sku",
        "indian_price_inr",
        "us_price_usd",
        "gbp_price",
        "zoho_db_sku_match",
        "record_source",
        "product_status",
        "slug",
        "variant_id",
        "is_woo_placeholder_sku",
        "notes",
    ]

    header_fill = PatternFill("solid", fgColor="1E3A2F")
    header_font = Font(color="FFFFFF", bold=True)
    yes_fill = PatternFill("solid", fgColor="DCFCE7")
    no_fill = PatternFill("solid", fgColor="FEE2E2")

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in master_rows:
        ws.append([row.get(h, "") for h in headers])

    for r_idx in range(2, ws.max_row + 1):
        match_cell = ws.cell(r_idx, headers.index("zoho_db_sku_match") + 1)
        fill = yes_fill if match_cell.value == "Yes" else no_fill
        for c_idx in range(1, len(headers) + 1):
            ws.cell(r_idx, c_idx).fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    autosize(ws)

    # Summary sheet
    summary = out_wb.create_sheet("Summary", 0)
    yes_count = sum(1 for r in master_rows if r["zoho_db_sku_match"] == "Yes")
    no_count = sum(1 for r in master_rows if r["zoho_db_sku_match"] == "No")
    by_source: dict[str, int] = {}
    for r in master_rows:
        by_source[r["record_source"]] = by_source.get(r["record_source"], 0) + 1

    summary_rows = [
        ("Total rows", len(master_rows)),
        ("Zoho & DB SKU match = Yes", yes_count),
        ("Zoho & DB SKU match = No", no_count),
        ("DB variants", by_source.get("DB", 0)),
        ("Zoho only (no DB row)", by_source.get("Zoho only", 0)),
        ("WC draft variations", by_source.get("WC draft variation", 0)),
        ("WC orphan variations", by_source.get("WC orphan variation", 0)),
        ("WC draft products (no DB)", by_source.get("WC draft product", 0)),
        ("Source file", str(input_path.name)),
    ]
    summary.append(("Metric", "Value"))
    for k, v in summary_rows:
        summary.append((k, v))
    autosize(summary)

    # Action queues
    def add_sheet(title: str, rows: list[dict]) -> None:
        sh = out_wb.create_sheet(title)
        sh.append(headers)
        for cell in sh[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            sh.append([row.get(h, "") for h in headers])
        sh.freeze_panes = "A2"
        autosize(sh)

    add_sheet("Match Yes", [r for r in master_rows if r["zoho_db_sku_match"] == "Yes"])
    add_sheet("Match No", [r for r in master_rows if r["zoho_db_sku_match"] == "No"])
    add_sheet(
        "Needs SKU rename",
        [
            r
            for r in master_rows
            if r["record_source"] == "DB"
            and r["zoho_db_sku_match"] == "No"
            and norm_sku(r["zoho_sku"])
        ],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)

    print(f"Wrote {output_path}")
    print(f"  Total rows: {len(master_rows)}")
    print(f"  SKU match Yes: {yes_count}")
    print(f"  SKU match No:  {no_count}")
    for src, n in sorted(by_source.items()):
        print(f"  {src}: {n}")


if __name__ == "__main__":
    main()
