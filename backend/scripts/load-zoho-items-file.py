#!/usr/bin/env python3
"""Read Zoho Items export (.xls or .xlsx) → JSON catalog on stdout."""
from __future__ import annotations

import json
import sys
from pathlib import Path


def normalize_hsn(raw) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.lower() in ("none", "nan"):
        return None
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s.split(".")[0] if s else None


def load_xls(path: Path) -> tuple[list[dict], dict]:
    import xlrd

    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_index(0)
    headers = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
    sku_i = headers.index("SKU")
    hsn_i = next(i for i, h in enumerate(headers) if "HSN" in h.upper() or "SAC" in h.upper())
    name_i = next(i for i, h in enumerate(headers) if h.lower() in ("item name", "name"))
    out = []
    for r in range(1, sh.nrows):
        sku = str(sh.cell_value(r, sku_i)).strip()
        if not sku or sku.lower() == "none":
            continue
        hsn = normalize_hsn(sh.cell_value(r, hsn_i))
        if not hsn:
            continue
        name = str(sh.cell_value(r, name_i)).strip()
        out.append({"sku": sku, "name": name, "hsn": hsn})
    return out, {"total_rows": sh.nrows - 1, "with_sku_and_hsn": len(out)}


def load_xlsx(path: Path) -> tuple[list[dict], dict]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    sku_i = headers.index("SKU")
    hsn_i = next(i for i, h in enumerate(headers) if "HSN" in h.upper() or "SAC" in h.upper())
    name_i = next(i for i, h in enumerate(headers) if h.lower() in ("item name", "name"))
    out = []
    for row in rows[1:]:
        if not row:
            continue
        sku = str(row[sku_i]).strip() if sku_i < len(row) and row[sku_i] is not None else ""
        if not sku or sku.lower() == "none":
            continue
        hsn = normalize_hsn(row[hsn_i] if hsn_i < len(row) else None)
        if not hsn:
            continue
        name = str(row[name_i]).strip() if name_i < len(row) and row[name_i] is not None else ""
        out.append({"sku": sku, "name": name, "hsn": hsn})
    return out, {"total_rows": len(rows) - 1, "with_sku_and_hsn": len(out)}


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: load-zoho-items-file.py <path.xls|xlsx>", file=sys.stderr)
        sys.exit(1)
    path = Path(sys.argv[1]).expanduser().resolve()
    if not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        sys.exit(1)
    suffix = path.suffix.lower()
    if suffix == ".xls":
        catalog, stats = load_xls(path)
    elif suffix in (".xlsx", ".xlsm"):
        catalog, stats = load_xlsx(path)
    else:
        print(f"Unsupported format: {suffix}", file=sys.stderr)
        sys.exit(1)
    print(json.dumps({"catalog": catalog, "source": str(path), "stats": stats}, ensure_ascii=False))


if __name__ == "__main__":
    main()
