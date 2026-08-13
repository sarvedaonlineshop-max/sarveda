#!/usr/bin/env python3
"""
Fuzzy compare team inventory xlsx vs live Lightsail catalog.

Outputs multi-sheet xlsx:
  - Exact Match
  - Pending (text diff on same SKU — excludes rows your DECISION already resolved)
  - Deferred (DECISION = will do later)
  - Sheet Only
  - DB Only

Re-use DECISION column from a prior compare file (--decisions) so accepted
decisions (keep DB / already applied) drop out of Pending.

Usage:
  python3 backend/scripts/fuzzy-compare-latest-inventory.py \\
    --xlsx data/latest_inventory.xlsx \\
    --out data/compare/latest-inventory-fuzzy.xlsx \\
    --decisions data/compare/latest-inventory-fuzzy.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

API_DEFAULT = "http://13.204.112.165"

# Accepted done — never show in Pending (team confirmed / manual apply)
MANUALLY_DONE_SKUS = frozenset(
    s.upper()
    for s in (
        "MI-SB-GAB-SET3",
        "MI-GRO",
        "MI-KL",
        "MI-JE-S",
        "MI-NF-T",
        "MI-NF-D",
        "MI-NF-S-L",
        "MI-NF-S-M",
        "MI-OC-S",
    )
)

HEADERS = [
    "DECISION",
    "Sheet Product name",
    "DB Product name",
    "Sheet Variant name",
    "DB Variant name",
    "Sheet SKU",
    "DB SKU",
    "Match score",
    "Match type",
    "Notes",
    "Pending reason",
]


@dataclass
class Row:
    product: str
    variant: str
    sku: str


def variant_match(a: str, b: str) -> bool:
    na, nb = norm_variant(a), norm_variant(b)
    if na == nb:
        return True
    if na and nb and set(na.split()) == set(nb.split()):
        return True
    return False


def rows_exact(s: Row, d: Row) -> bool:
    return (
        norm_text(s.product) == norm_text(d.product)
        and variant_match(s.variant, d.variant)
        and norm_sku(s.sku) == norm_sku(d.sku)
    )


def norm_text(s: str) -> str:
    return " ".join(str(s or "").lower().split())


def norm_sku(s: str) -> str:
    return " ".join(str(s or "").strip().split()).upper()


def norm_variant(s: str) -> str:
    s = norm_text(s)
    return re.sub(r"[\s|/·,–—\-]+", " ", s).strip()


def resolve_targets(
    decision: str,
    sheet_name: str,
    sheet_variant: str,
    sheet_sku: str,
    db_name: str,
    db_variant: str,
    db_sku: str,
) -> tuple[str, str, str] | None:
    """Same rules as apply-fuzzy-match-decisions.py."""
    d = (decision or "").strip().lower()
    if not d or "will do later" in d:
        return None

    product = db_name
    variant = db_variant
    sku = (db_sku or sheet_sku).strip()

    if (
        "adapt sheet product name" in d
        or "product name adapt from sheet" in d
        or "adapt db product name" in d
    ):
        product = sheet_name
    elif "product name same" in d:
        product = db_name

    if "vairant adapt from sheet" in d or "variant adapt from sheet" in d:
        variant = sheet_variant
    elif (
        "vairant adapt from db" in d
        or "variant adapt from db" in d
        or "vairant same" in d
        or "variant same" in d
    ):
        variant = db_variant

    return product.strip(), (variant or "").strip(), sku


def load_decisions_by_sku(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}

    json_sidecar = path.parent / "fuzzy-decisions-by-sku.json"
    if json_sidecar.exists():
        out.update({k.upper(): v for k, v in json.loads(json_sidecar.read_text()).items()})

    if not path.exists():
        return out
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "Fuzzy Match" if "Fuzzy Match" in wb.sheetnames else "Pending"
    if sheet_name not in wb.sheetnames:
        wb.close()
        return out
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return {}
    headers = [str(h or "").strip() for h in rows[0]]
    for r in rows[1:]:
        rec = {headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))}
        decision = str(rec.get("DECISION") or "").strip()
        if not decision:
            continue
        sku = str(rec.get("DB SKU") or rec.get("Sheet SKU") or "").strip().upper()
        if sku:
            out[sku] = decision
    wb.close()
    return out


def split_by_decision(fuzzy: list[dict], decisions_by_sku: dict[str, str]) -> tuple[list[dict], list[dict], int]:
    pending: list[dict] = []
    deferred: list[dict] = []
    resolved = 0

    for row in fuzzy:
        sku = norm_sku(str(row.get("DB SKU") or row.get("Sheet SKU") or ""))
        decision = decisions_by_sku.get(sku, "")
        row = {**row, "DECISION": decision}

        if sku in MANUALLY_DONE_SKUS:
            resolved += 1
            continue

        if not decision:
            row["Pending reason"] = "No DECISION yet"
            pending.append(row)
            continue

        if "will do later" in decision.lower():
            row["Pending reason"] = "Deferred by you"
            deferred.append(row)
            continue

        targets = resolve_targets(
            decision,
            str(row.get("Sheet Product name") or ""),
            str(row.get("Sheet Variant name") or ""),
            str(row.get("Sheet SKU") or ""),
            str(row.get("DB Product name") or ""),
            str(row.get("DB Variant name") or ""),
            str(row.get("DB SKU") or ""),
        )
        if targets is None:
            row["Pending reason"] = "Deferred by you"
            deferred.append(row)
            continue

        target_product, target_variant, target_sku = targets
        db_product = str(row.get("DB Product name") or "").strip()
        db_variant = str(row.get("DB Variant name") or "").strip()
        db_sku = str(row.get("DB SKU") or "").strip()

        if (
            db_product == target_product
            and db_variant == target_variant
            and norm_sku(db_sku) == norm_sku(target_sku)
        ):
            resolved += 1
            continue

        changes: list[str] = []
        if db_product != target_product:
            changes.append(f"product -> {target_product!r}")
        if db_variant != target_variant:
            changes.append(f"variant -> {target_variant!r}")
        if norm_sku(db_sku) != norm_sku(target_sku):
            changes.append(f"sku -> {target_sku!r}")
        row["Pending reason"] = "Apply pending: " + "; ".join(changes)
        pending.append(row)

    return pending, deferred, resolved


def load_orphan_plan() -> dict | None:
    plan_path = Path(__file__).resolve().parents[2] / "data/compare/sheet-db-only-plan.json"
    if not plan_path.exists():
        return None
    return json.loads(plan_path.read_text())


def apply_orphan_plan(
    sheet_only: list[dict], db_only: list[dict], plan: dict | None
) -> tuple[list[dict], list[dict], list[dict]]:
    if not plan:
        return sheet_only, db_only, []

    rename_to = {norm_sku(x["toSku"]) for x in plan.get("rename_skus", [])}
    rename_from = {norm_sku(x["fromSku"]) for x in plan.get("rename_skus", [])}
    create_skus = {
        norm_sku(x["sku"])
        for x in plan.get("create_variants", []) + plan.get("create_products", [])
    }
    draft_skus = {norm_sku(x["sku"]) for x in plan.get("draft_variants", [])}

    actions: list[dict] = []
    for x in plan.get("rename_skus", []):
        actions.append(
            {
                "Action": "rename_sku",
                "SKU": x["toSku"],
                "Details": f"{x['fromSku']} -> {x['toSku']}",
                "Reason": x.get("reason", ""),
            }
        )
    for x in plan.get("create_variants", []):
        actions.append(
            {
                "Action": "create_variant",
                "SKU": x["sku"],
                "Details": x.get("productSlug", ""),
                "Reason": x.get("reason", ""),
            }
        )
    for x in plan.get("create_products", []):
        actions.append(
            {
                "Action": "create_product",
                "SKU": x["sku"],
                "Details": x.get("productName", ""),
                "Reason": x.get("reason", ""),
            }
        )
    for x in plan.get("draft_variants", []):
        actions.append(
            {
                "Action": "draft_variant",
                "SKU": x["sku"],
                "Details": x.get("productName", ""),
                "Reason": x.get("reason", ""),
            }
        )

    filtered_sheet = [
        r
        for r in sheet_only
        if norm_sku(str(r.get("Sheet SKU") or "")) not in rename_to
        and norm_sku(str(r.get("Sheet SKU") or "")) not in create_skus
    ]
    filtered_db = [
        r
        for r in db_only
        if norm_sku(str(r.get("DB SKU") or "")) not in rename_from
        and norm_sku(str(r.get("DB SKU") or "")) not in draft_skus
    ]
    return filtered_sheet, filtered_db, actions


def write_actions_sheet(ws, rows: list[dict]) -> None:
    headers = ["Action", "SKU", "Details", "Reason"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for ri, row in enumerate(rows, start=2):
        for ci, h in enumerate(headers, start=1):
            ws.cell(row=ri, column=ci, value=row.get(h, ""))
    for i, w in enumerate([18, 20, 42, 48], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def fuzzy_score(sheet: Row, db: Row) -> float:
    if norm_sku(sheet.sku) == norm_sku(db.sku):
        name_s = fuzz.token_set_ratio(sheet.product, db.product)
        var_s = fuzz.token_set_ratio(sheet.variant or "", db.variant or "") if (sheet.variant or db.variant) else 100.0
        return min(100.0, 40 + name_s * 0.35 + var_s * 0.25)

    name_s = fuzz.token_set_ratio(sheet.product, db.product)
    var_s = fuzz.token_set_ratio(sheet.variant or "", db.variant or "")
    sku_s = fuzz.ratio(norm_sku(sheet.sku), norm_sku(db.sku))
    combo = name_s * 0.45 + var_s * 0.35 + sku_s * 0.20
    if name_s >= 92 and var_s >= 85:
        combo = max(combo, 88.0)
    return combo


def load_inventory(path: Path) -> list[Row]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Website Catalog"] if "Website Catalog" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_i = next(i for i, r in enumerate(rows[:20]) if r and str(r[0] or "").strip() == "Name")
    out: list[Row] = []
    cur = ""
    for r in rows[header_i + 1 :]:
        if not r:
            continue
        if r[0] is not None and str(r[0]).strip():
            cur = str(r[0]).strip()
        sku = "" if len(r) < 3 or r[2] is None else str(r[2]).strip()
        if not sku:
            continue
        variant = "" if r[1] is None else str(r[1]).strip()
        out.append(Row(cur, variant, sku))
    wb.close()
    return out


def fetch_json(url: str):
    for attempt in range(5):
        try:
            with urllib.request.urlopen(url, timeout=60) as resp:
                return json.loads(resp.read().decode())
        except urllib.error.HTTPError as e:
            if e.code == 429 and attempt < 4:
                time.sleep(1.5 * (attempt + 1))
                continue
            raise
        finally:
            time.sleep(0.08)


def variant_label(attrs) -> str:
    if not attrs:
        return ""
    parts = sorted(
        (a["attributeValue"]["attribute"]["slug"], a["attributeValue"]["value"])
        for a in attrs
        if a.get("attributeValue")
    )
    return " / ".join(v for _, v in parts)


def fetch_lightsail_rows(api: str) -> list[Row]:
    page = 1
    slugs: list[str] = []
    while True:
        data = fetch_json(f"{api}/api/products?limit=100&page={page}&status=ACTIVE")
        items = data["data"]["items"]
        slugs.extend(i["slug"] for i in items)
        pag = data["data"]["pagination"]
        if page >= pag["totalPages"]:
            break
        page += 1

    rows: list[Row] = []

    def one(slug: str) -> list[Row]:
        data = fetch_json(f"{api}/api/products/{slug}")
        p = data["data"]["product"]
        if p.get("deletedAt") or p.get("status") != "ACTIVE" or p.get("catalogHidden"):
            return []
        name = p["name"]
        out: list[Row] = []
        for v in p.get("variants") or []:
            if v.get("status") != "ACTIVE":
                continue
            out.append(Row(name, variant_label(v.get("attributeValues")), v["sku"]))
        return out

    with ThreadPoolExecutor(max_workers=12) as ex:
        for fut in as_completed(ex.submit(one, s) for s in slugs):
            rows.extend(fut.result())
    return rows


def classify_match(sheet: Row, db: Row) -> tuple[str, str]:
    if rows_exact(sheet, db):
        return "exact", "All three fields match"

    notes: list[str] = []
    if norm_sku(sheet.sku) != norm_sku(db.sku):
        notes.append("SKU differs")
    if norm_text(sheet.product) != norm_text(db.product):
        notes.append("Product name differs")
    if not variant_match(sheet.variant, db.variant):
        notes.append("Variant name differs")
    if norm_sku(sheet.sku) == norm_sku(db.sku):
        return "sku_exact_text_diff", "; ".join(notes)
    return "fuzzy", "; ".join(notes) if notes else "Fuzzy paired"


def match_rows(sheet_rows: list[Row], db_rows: list[Row], *, fuzzy_threshold: float = 82.0):
    exact: list[dict] = []
    fuzzy: list[dict] = []
    used_db: set[int] = set()
    used_sheet: set[int] = set()

    db_by_sku: dict[str, list[int]] = {}
    for i, r in enumerate(db_rows):
        db_by_sku.setdefault(norm_sku(r.sku), []).append(i)

    # Pass 1: exact triple (name + variant + SKU)
    for si, s in enumerate(sheet_rows):
        for di, d in enumerate(db_rows):
            if di in used_db or si in used_sheet:
                continue
            if rows_exact(s, d):
                used_sheet.add(si)
                used_db.add(di)
                exact.append(
                    {
                        "DECISION": "",
                        "Sheet Product name": s.product,
                        "DB Product name": d.product,
                        "Sheet Variant name": s.variant,
                        "DB Variant name": d.variant,
                        "Sheet SKU": s.sku,
                        "DB SKU": d.sku,
                        "Match score": 100,
                        "Match type": "exact",
                        "Notes": "All three fields match",
                        "Pending reason": "",
                    }
                )
                break

    # Pass 2: same SKU, text differs
    for si, s in enumerate(sheet_rows):
        if si in used_sheet:
            continue
        candidates = [di for di in db_by_sku.get(norm_sku(s.sku), []) if di not in used_db]
        if len(candidates) == 1:
            di = candidates[0]
            d = db_rows[di]
            used_sheet.add(si)
            used_db.add(di)
            mtype, note = classify_match(s, d)
            fuzzy.append(
                {
                    "DECISION": "",
                    "Sheet Product name": s.product,
                    "DB Product name": d.product,
                    "Sheet Variant name": s.variant,
                    "DB Variant name": d.variant,
                    "Sheet SKU": s.sku,
                    "DB SKU": d.sku,
                    "Match score": round(fuzzy_score(s, d), 1),
                    "Match type": mtype,
                    "Notes": note,
                    "Pending reason": "",
                }
            )

    # Pass 3: greedy fuzzy on remaining
    pairs: list[tuple[float, int, int]] = []
    for si, s in enumerate(sheet_rows):
        if si in used_sheet:
            continue
        for di, d in enumerate(db_rows):
            if di in used_db:
                continue
            sc = fuzzy_score(s, d)
            if sc >= fuzzy_threshold:
                pairs.append((sc, si, di))
    pairs.sort(reverse=True)

    for sc, si, di in pairs:
        if si in used_sheet or di in used_db:
            continue
        s, d = sheet_rows[si], db_rows[di]
        used_sheet.add(si)
        used_db.add(di)
        mtype, note = classify_match(s, d)
        fuzzy.append(
            {
                "DECISION": "",
                "Sheet Product name": s.product,
                "DB Product name": d.product,
                "Sheet Variant name": s.variant,
                "DB Variant name": d.variant,
                "Sheet SKU": s.sku,
                "DB SKU": d.sku,
                "Match score": round(sc, 1),
                "Match type": mtype,
                "Notes": note,
                "Pending reason": "",
            }
        )

    sheet_only = [
        {
            "DECISION": "",
            "Sheet Product name": sheet_rows[si].product,
            "DB Product name": "",
            "Sheet Variant name": sheet_rows[si].variant,
            "DB Variant name": "",
            "Sheet SKU": sheet_rows[si].sku,
            "DB SKU": "",
            "Match score": "",
            "Match type": "sheet_only",
            "Notes": "No fuzzy match on Lightsail",
            "Pending reason": "Create on Lightsail",
        }
        for si in range(len(sheet_rows))
        if si not in used_sheet
    ]

    db_only = [
        {
            "DECISION": "",
            "Sheet Product name": "",
            "DB Product name": db_rows[di].product,
            "Sheet Variant name": "",
            "DB Variant name": db_rows[di].variant,
            "Sheet SKU": "",
            "DB SKU": db_rows[di].sku,
            "Match score": "",
            "Match type": "db_only",
            "Notes": "Not on team sheet",
            "Pending reason": "Add to sheet or hide",
        }
        for di in range(len(db_rows))
        if di not in used_db
    ]

    return exact, fuzzy, sheet_only, db_only


def write_sheet(ws, rows: list[dict]) -> None:
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for ri, row in enumerate(rows, start=2):
        for ci, h in enumerate(HEADERS, start=1):
            ws.cell(row=ri, column=ci, value=row.get(h, ""))
    widths = [36, 38, 38, 32, 32, 18, 18, 12, 16, 40, 44]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=Path("data/latest_inventory.xlsx"))
    ap.add_argument("--out", type=Path, default=Path("data/compare/latest-inventory-fuzzy.xlsx"))
    ap.add_argument("--api", default=API_DEFAULT)
    ap.add_argument("--threshold", type=float, default=82.0)
    ap.add_argument(
        "--decisions",
        type=Path,
        default=None,
        help="Prior fuzzy xlsx with DECISION column (default: --out if it exists)",
    )
    args = ap.parse_args()

    decisions_path = args.decisions
    if decisions_path is None and args.out.exists():
        decisions_path = args.out

    print("Loading sheet...")
    sheet_rows = load_inventory(args.xlsx)
    print(f"Fetching Lightsail ({args.api})...")
    db_rows = fetch_lightsail_rows(args.api)
    print(f"Sheet: {len(sheet_rows)} | DB: {len(db_rows)}")

    exact, fuzzy_raw, sheet_only, db_only = match_rows(
        sheet_rows, db_rows, fuzzy_threshold=args.threshold
    )

    decisions_by_sku: dict[str, str] = {}
    if decisions_path:
        decisions_by_sku = load_decisions_by_sku(decisions_path)
        print(f"Loaded {len(decisions_by_sku)} decisions from {decisions_path}")

    pending, deferred, resolved = split_by_decision(fuzzy_raw, decisions_by_sku)
    print(f"Fuzzy raw: {len(fuzzy_raw)} | Resolved by decision: {resolved} | Pending: {len(pending)} | Deferred: {len(deferred)}")

    orphan_plan = load_orphan_plan()
    sheet_only, db_only, orphan_actions = apply_orphan_plan(sheet_only, db_only, orphan_plan)
    if orphan_plan:
        s = orphan_plan.get("summary", {})
        print(
            f"Orphan plan: rename={s.get('rename_skus', 0)} create_variants={s.get('create_variants', 0)} "
            f"create_products={s.get('create_products', 0)} draft={s.get('draft_variants', 0)} | "
            f"Sheet Only after plan: {len(sheet_only)} | DB Only after plan: {len(db_only)}"
        )

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ("Exact Match", exact),
        ("Pending", pending),
        ("Deferred", deferred),
        ("Actions", orphan_actions),
        ("Sheet Only", sheet_only),
        ("DB Only", db_only),
    ]
    for title, rows in sheets:
        ws = wb.create_sheet(title)
        if title == "Actions":
            write_actions_sheet(ws, rows)
        else:
            write_sheet(ws, rows)

    summary = wb.create_sheet("Summary", 0)
    summary["A1"] = "Metric"
    summary["B1"] = "Count"
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    metrics = [
        ("Sheet rows (with SKU)", len(sheet_rows)),
        ("Lightsail ACTIVE variant rows", len(db_rows)),
        ("Exact Match", len(exact)),
        ("Fuzzy raw (before decisions)", len(fuzzy_raw)),
        ("Resolved by your DECISION", resolved),
        ("Pending (needs action)", len(pending)),
        ("Deferred (will do later)", len(deferred)),
        ("Planned actions", len(orphan_actions)),
        ("Sheet Only (unmatched, after plan)", len(sheet_only)),
        ("DB Only (unmatched, after plan)", len(db_only)),
        ("Fuzzy threshold", args.threshold),
        ("Decisions file", str(decisions_path) if decisions_path else ""),
        ("Source API", args.api),
    ]
    for i, (k, v) in enumerate(metrics, start=2):
        summary.cell(row=i, column=1, value=k)
        summary.cell(row=i, column=2, value=v)
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 20

    wb.save(args.out)

    print()
    print(f"Exact Match:  {len(exact)}")
    print(f"Pending:      {len(pending)}")
    print(f"Deferred:     {len(deferred)}")
    print(f"Resolved:     {resolved} (dropped from Pending)")
    print(f"Actions:      {len(orphan_actions)} (see Actions sheet)")
    print(f"Sheet Only:   {len(sheet_only)}")
    print(f"DB Only:      {len(db_only)}")
    print(f"Wrote: {args.out}")


if __name__ == "__main__":
    main()
