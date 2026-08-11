#!/usr/bin/env python3
"""
Cross-check Inventory sheet (Name, Variant Name, SKU) against current DB variants.

Usage (from repo root, with DATABASE_URL in backend/.env):
  cd backend && python3 ../backend/scripts/compare-inventory-abc-to-db.py

Optional:
  python3 backend/scripts/compare-inventory-abc-to-db.py \\
    --xlsx "data/Inventory count 09 Aug Price, and SKU.xlsx" \\
    --out-dir data/compare/inventory-09aug-abc
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    psycopg2 = None  # type: ignore


def load_database_url() -> str:
    env_path = Path(__file__).resolve().parents[2] / "backend" / ".env"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            if line.startswith("DATABASE_URL="):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    url = os.environ.get("DATABASE_URL", "").strip()
    if not url:
        raise SystemExit("DATABASE_URL not found in backend/.env or environment")
    return url


def norm_text(s: str) -> str:
    return " ".join(str(s or "").lower().split())


def norm_sku(s: str) -> str:
    return " ".join(str(s or "").strip().split()).upper()


def norm_variant(s: str) -> str:
    s = norm_text(s)
    s = re.sub(r"[\s|/·,]+", " ", s).strip()
    return s


def variant_match(a: str, b: str) -> bool:
    na, nb = norm_variant(a), norm_variant(b)
    if na == nb:
        return True
    if na and nb and set(na.split()) == set(nb.split()):
        return True
    return False


def load_inventory(path: Path) -> list[tuple[str, str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Website Catalog"] if "Website Catalog" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_i = next(i for i, r in enumerate(rows[:20]) if r and str(r[0] or "").strip() == "Name")
    out: list[tuple[str, str, str]] = []
    cur = ""
    for r in rows[header_i + 1 :]:
        if not r:
            continue
        if r[0] is not None and str(r[0]).strip():
            cur = str(r[0]).strip()
        sku = "" if len(r) < 3 or r[2] is None else str(r[2]).strip()
        if not sku:
            continue
        variant = "" if r[1] is None else str(r[1]).strip()
        out.append((cur, variant, sku))
    wb.close()
    return out


def fetch_db_rows(database_url: str) -> list[tuple[str, str, str]]:
    """Return (product_name, variant_label, sku) for non-deleted products."""
    if psycopg2 is None:
        raise SystemExit("psycopg2 is required: pip install psycopg2-binary")

    sql = """
    SELECT
      p.name AS product_name,
      coalesce((
        SELECT string_agg(av.value, ' / ' ORDER BY pa.slug)
        FROM "VariantAttributeValue" vv
        JOIN "AttributeValue" av ON av.id = vv."attributeValueId"
        JOIN "ProductAttribute" pa ON pa.id = av."attributeId"
        WHERE vv."variantId" = v.id
      ), '') AS variant_name,
      v.sku AS sku
    FROM "ProductVariant" v
    JOIN "Product" p ON p.id = v."productId"
    WHERE p."deletedAt" IS NULL
    ORDER BY p.name, v.sku
    """
    conn = psycopg2.connect(database_url)
    try:
        with conn.cursor() as cur:
            cur.execute(sql)
            return [(str(a), str(b or ""), str(c)) for a, b, c in cur.fetchall()]
    finally:
        conn.close()


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=Path("data/Inventory count 09 Aug Price, and SKU.xlsx"),
    )
    ap.add_argument("--out-dir", type=Path, default=Path("data/compare/inventory-09aug-abc"))
    args = ap.parse_args()
    args.out_dir.mkdir(parents=True, exist_ok=True)

    inv = load_inventory(args.xlsx)
    db = fetch_db_rows(load_database_url())

    by_sku: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
    by_nv: dict[tuple[str, str], list[tuple[str, str, str]]] = defaultdict(list)
    for name, variant, sku in db:
        by_sku[norm_sku(sku)].append((name, variant, sku))
        by_nv[(norm_text(name), norm_variant(variant))].append((name, variant, sku))

    exact_rows: list[dict] = []
    sku_alone_rows: list[dict] = []
    nv_rows: list[dict] = []
    no_match_rows: list[dict] = []

    exact = sku_alone = nv_match = 0
    sku_hit = 0

    for name, variant, sku in inv:
        ns = norm_sku(sku)
        entries = by_sku.get(ns, [])
        nv_hits = by_nv.get((norm_text(name), norm_variant(variant)), [])
        # also tolerant variant token match under same product name
        if not nv_hits:
            nv_hits = [
                row
                for rows in by_nv.values()
                for row in rows
                if norm_text(row[0]) == norm_text(name) and variant_match(variant, row[1])
            ]

        if entries:
            sku_hit += 1
            if any(
                norm_text(db_name) == norm_text(name) and variant_match(variant, db_var)
                for db_name, db_var, _ in entries
            ):
                exact += 1
                exact_rows.append(
                    {
                        "sheet_name": name,
                        "sheet_variant": variant,
                        "sheet_sku": sku,
                        "db_name": entries[0][0],
                        "db_variant": entries[0][1],
                        "db_sku": entries[0][2],
                    }
                )
            else:
                sku_alone += 1
                sku_alone_rows.append(
                    {
                        "sheet_name": name,
                        "sheet_variant": variant,
                        "sheet_sku": sku,
                        "db_name": entries[0][0],
                        "db_variant": entries[0][1],
                        "db_sku": entries[0][2],
                    }
                )
        elif nv_hits:
            nv_match += 1
            nv_rows.append(
                {
                    "sheet_name": name,
                    "sheet_variant": variant,
                    "sheet_sku": sku,
                    "db_name": nv_hits[0][0],
                    "db_variant": nv_hits[0][1],
                    "db_sku": nv_hits[0][2],
                }
            )
        else:
            # name+variant may still match even when sku matched differently — count nv separately below
            no_match_rows.append(
                {"sheet_name": name, "sheet_variant": variant, "sheet_sku": sku}
            )

    # Name+variant matches across whole sheet (including those that also matched SKU)
    nv_any = 0
    for name, variant, sku in inv:
        hits = [
            row
            for rows in by_nv.values()
            for row in rows
            if norm_text(row[0]) == norm_text(name) and variant_match(variant, row[1])
        ]
        if hits:
            nv_any += 1

    def write_csv(path: Path, rows: list[dict]) -> None:
        if not rows:
            path.write_text("")
            return
        with path.open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)

    write_csv(args.out_dir / "01-exact-name-variant-sku.csv", exact_rows)
    write_csv(args.out_dir / "02-sku-alone.csv", sku_alone_rows)
    write_csv(args.out_dir / "03-name-variant-sku-differs.csv", nv_rows)
    write_csv(args.out_dir / "04-no-sku-no-name-variant.csv", no_match_rows)

    summary = {
        "inventory_rows_with_sku": len(inv),
        "db_variant_rows": len(db),
        "exact_name_variant_sku": exact,
        "sku_alone_name_or_variant_differ": sku_alone,
        "sku_found_total": sku_hit,
        "name_and_variant_match_sku_missing_or_different": nv_match,
        "name_and_variant_match_any": nv_any,
        "no_sku_and_no_name_variant": len(no_match_rows),
    }
    (args.out_dir / "summary.json").write_text(
        __import__("json").dumps(summary, indent=2) + "\n"
    )

    print("Inventory rows (with SKU):", summary["inventory_rows_with_sku"])
    print("DB variant rows:", summary["db_variant_rows"])
    print()
    print("Exact match (Name + Variant Name + SKU):", summary["exact_name_variant_sku"])
    print("SKU alone (SKU matches, name and/or variant differ):", summary["sku_alone_name_or_variant_differ"])
    print("Name + Variant match (SKU missing/different on that row bucket):", summary["name_and_variant_match_sku_missing_or_different"])
    print("Name + Variant match (any, including exact):", summary["name_and_variant_match_any"])
    print()
    print("Wrote:", args.out_dir)


if __name__ == "__main__":
    main()
