#!/usr/bin/env python3
"""
DO vs Lightsail — pending launch list (6-column side-by-side format).

Columns (only 6):
  DO Product name | LS Product name | DO Variant name | LS Variant name | DO SKU | LS SKU

Sheets:
  Summary
  Exact Match        — product pair + variant row fully aligned
  Fuzzy Match        — product pair matched, variant name/SKU differs
  DO Only            — product on DO only (all variant rows)
  LS Only            — product on LS only (all variant rows)

Product match: exact slug → exact woo id → fuzzy name (≥82).
Within pair: variant match by SKU first, then fuzzy/exact variant name.

Usage:
  python3 backend/scripts/compare-do-ls-pending-6col.py
"""
from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

ROOT = Path(__file__).resolve().parents[2]
DO_PRODUCTS = ROOT / "data/compare/do_products.csv"
DO_VARIANTS = ROOT / "data/compare/do_variants.csv"
LS_EXPORT = ROOT / "data/compare/lightsail-catalog-export.json"
OUT_XLSX = ROOT / "data/compare/pending-do-vs-ls-launch.xlsx"
OUT_JSON = ROOT / "data/compare/pending-do-vs-ls-launch-summary.json"

FUZZY_MIN = 82.0
VARIANT_FUZZY_MIN = 85.0

DO_ONLY_EXCLUDE = {"printed-copper-water-bottles"}
LS_ONLY_EXCLUDE = {"copper-bottle-blue-tranquillity-meditation"}

COLS_6 = [
    "DO Product name",
    "LS Product name",
    "DO Variant name",
    "LS Variant name",
    "DO SKU",
    "LS SKU",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN = PatternFill("solid", fgColor="E2EFDA")
BLUE = PatternFill("solid", fgColor="DDEBF7")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
RED = PatternFill("solid", fgColor="FCE4D6")


@dataclass
class VariantRow:
    name: str
    sku: str


@dataclass
class ProductRow:
    key: str
    slug: str
    name: str
    woo_id: int | None
    variants: list[VariantRow] = field(default_factory=list)


def norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def norm_sku(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", (s or "").strip().upper())


def norm_variant(s: str) -> str:
    s = norm_text(s)
    return re.sub(r"[\s|/·,–—\-]+", " ", s).strip()


def variant_exact(a: str, b: str) -> bool:
    na, nb = norm_variant(a), norm_variant(b)
    if not na and not nb:
        return True
    if na == nb:
        return True
    return bool(na and nb and set(na.split()) == set(nb.split()))


def variant_score(a: str, b: str) -> float:
    na, nb = norm_variant(a), norm_variant(b)
    if not na and not nb:
        return 100.0
    if variant_exact(a, b):
        return 100.0
    return float(fuzz.token_sort_ratio(na, nb))


def product_score(a: ProductRow, b: ProductRow) -> float:
    if a.woo_id and b.woo_id and a.woo_id == b.woo_id:
        return 100.0
    if norm_text(a.slug) == norm_text(b.slug):
        return 99.0
    return float(fuzz.token_sort_ratio(norm_text(a.name), norm_text(b.name)))


def is_shop_slug(slug: str) -> bool:
    s = (slug or "").lower()
    return not (s.startswith("course-checkout-") or s.startswith("event-checkout-"))


def parse_do_variant_name(attrs: str, title: str, product_name: str) -> str:
    if attrs:
        parts = []
        for seg in attrs.split(";"):
            if "=" in seg:
                parts.append(seg.split("=", 1)[1].strip())
        if parts:
            return " / ".join(parts)
    if title and " - " in title:
        tail = title.split(" - ", 1)[1].strip()
        if tail and norm_text(tail) != norm_text(product_name):
            return tail
    return ""


def load_do_products() -> list[ProductRow]:
    products: dict[str, dict] = {}
    with DO_PRODUCTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            products[r["id"]] = r

    by_parent: dict[str, list] = {}
    with DO_VARIANTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            by_parent.setdefault(r["parent_id"], []).append(r)

    out: list[ProductRow] = []
    for pid, p in products.items():
        slug = (p.get("slug") or "").strip()
        if not is_shop_slug(slug):
            continue
        name = (p.get("name") or "").strip()
        ptype = (p.get("product_type") or "").lower()
        vars_ = by_parent.get(pid, [])
        variants: list[VariantRow] = []
        if ptype == "simple" or not vars_:
            variants.append(VariantRow(name="", sku=(p.get("sku") or "").strip()))
        else:
            for v in vars_:
                variants.append(
                    VariantRow(
                        name=parse_do_variant_name(v.get("attrs") or "", v.get("title") or "", name),
                        sku=(v.get("sku") or "").strip(),
                    )
                )
        out.append(
            ProductRow(
                key=f"do:{pid}",
                slug=slug,
                name=name,
                woo_id=int(pid),
                variants=variants,
            )
        )
    return out


def load_ls_products() -> list[ProductRow]:
    data = json.loads(LS_EXPORT.read_text(encoding="utf-8"))
    grouped: dict[str, ProductRow] = {}
    for r in data["rows"]:
        slug = (r.get("slug") or "").strip()
        if not is_shop_slug(slug):
            continue
        if slug not in grouped:
            grouped[slug] = ProductRow(
                key=f"ls:{slug}",
                slug=slug,
                name=(r.get("name") or "").strip(),
                woo_id=int(r["wooCommerceId"]) if r.get("wooCommerceId") else None,
                variants=[],
            )
        grouped[slug].variants.append(
            VariantRow(name=(r.get("variantName") or "").strip(), sku=(r.get("sku") or "").strip())
        )
    return sorted(grouped.values(), key=lambda p: p.slug)


def match_products(do_products: list[ProductRow], ls_products: list[ProductRow]):
    ls_by_slug = {norm_text(p.slug): p for p in ls_products if p.slug}
    do_by_slug = {norm_text(p.slug): p for p in do_products if p.slug}
    ls_by_woo = {p.woo_id: p for p in ls_products if p.woo_id}
    do_by_woo = {p.woo_id: p for p in do_products if p.woo_id}

    used_ls: set[str] = set()
    used_do: set[str] = set()
    pairs: list[tuple[ProductRow, ProductRow, str, float]] = []

    for lp in ls_products:
        ns = norm_text(lp.slug)
        if ns and ns in do_by_slug:
            dp = do_by_slug[ns]
            if dp.key not in used_do and lp.key not in used_ls:
                pairs.append((lp, dp, "exact_slug", 99.0))
                used_ls.add(lp.key)
                used_do.add(dp.key)

    for lp in ls_products:
        if lp.key in used_ls:
            continue
        if lp.woo_id and lp.woo_id in do_by_woo:
            dp = do_by_woo[lp.woo_id]
            if dp.key not in used_do:
                pairs.append((lp, dp, "exact_woo_id", 100.0))
                used_ls.add(lp.key)
                used_do.add(dp.key)

    ls_rem = [p for p in ls_products if p.key not in used_ls]
    do_rem = [p for p in do_products if p.key not in used_do]

    candidates: list[tuple[float, int, int]] = []
    for li, lp in enumerate(ls_rem):
        for di, dp in enumerate(do_rem):
            s = product_score(lp, dp)
            if s >= FUZZY_MIN:
                candidates.append((s, li, di))
    candidates.sort(reverse=True)

    used_lr: set[int] = set()
    used_dr: set[int] = set()
    for s, li, di in candidates:
        if li in used_lr or di in used_dr:
            continue
        used_lr.add(li)
        used_dr.add(di)
        pairs.append((ls_rem[li], do_rem[di], "fuzzy_name", s))
        used_ls.add(ls_rem[li].key)
        used_do.add(do_rem[di].key)

    ls_only = [p for p in ls_products if p.key not in used_ls]
    do_only = [p for p in do_products if p.key not in used_do]
    return pairs, ls_only, do_only


def match_variants_in_pair(lp: ProductRow, dp: ProductRow) -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    """Returns exact_rows, fuzzy_rows, ls_only_rows, do_only_rows (6-col dicts)."""
    exact: list[dict] = []
    fuzzy: list[dict] = []
    ls_only: list[dict] = []
    do_only: list[dict] = []

    used_d: set[int] = set()
    used_l: set[int] = set()

    def mk(dv: VariantRow | None, lv: VariantRow | None) -> dict:
        return {
            "DO Product name": dp.name,
            "LS Product name": lp.name,
            "DO Variant name": dv.name if dv else "",
            "LS Variant name": lv.name if lv else "",
            "DO SKU": dv.sku if dv else "",
            "LS SKU": lv.sku if lv else "",
        }

    # SKU match first
    ls_by_sku: dict[str, list[int]] = {}
    for i, lv in enumerate(lp.variants):
        k = norm_sku(lv.sku)
        if k:
            ls_by_sku.setdefault(k, []).append(i)

    for di, dv in enumerate(dp.variants):
        k = norm_sku(dv.sku)
        if not k:
            continue
        cands = [i for i in ls_by_sku.get(k, []) if i not in used_l]
        if len(cands) == 1:
            li = cands[0]
            lv = lp.variants[li]
            used_d.add(di)
            used_l.add(li)
            row = mk(dv, lv)
            if variant_exact(dv.name, lv.name) and norm_text(dp.name) == norm_text(lp.name):
                exact.append(row)
            else:
                fuzzy.append(row)

    # Greedy variant fuzzy on remainder
    cands: list[tuple[float, int, int]] = []
    for di, dv in enumerate(dp.variants):
        if di in used_d:
            continue
        for li, lv in enumerate(lp.variants):
            if li in used_l:
                continue
            sc = variant_score(dv.name, lv.name)
            if sc >= VARIANT_FUZZY_MIN:
                cands.append((sc, di, li))
    cands.sort(reverse=True)
    for sc, di, li in cands:
        if di in used_d or li in used_l:
            continue
        used_d.add(di)
        used_l.add(li)
        dv, lv = dp.variants[di], lp.variants[li]
        row = mk(dv, lv)
        if variant_exact(dv.name, lv.name) and norm_sku(dv.sku) == norm_sku(lv.sku):
            exact.append(row)
        else:
            fuzzy.append(row)

    for li, lv in enumerate(lp.variants):
        if li not in used_l:
            ls_only.append(mk(None, lv))
    for di, dv in enumerate(dp.variants):
        if di not in used_d:
            do_only.append(mk(dv, None))

    return exact, fuzzy, ls_only, do_only


def product_rows_side(product: ProductRow, side: str) -> list[dict]:
    rows: list[dict] = []
    for v in product.variants:
        if side == "do":
            rows.append(
                {
                    "DO Product name": product.name,
                    "LS Product name": "",
                    "DO Variant name": v.name,
                    "LS Variant name": "",
                    "DO SKU": v.sku,
                    "LS SKU": "",
                }
            )
        else:
            rows.append(
                {
                    "DO Product name": "",
                    "LS Product name": product.name,
                    "DO Variant name": "",
                    "LS Variant name": v.name,
                    "DO SKU": "",
                    "LS SKU": v.sku,
                }
            )
    return rows


def write_sheet(ws, headers: list[str], rows: list[dict], widths: list[int] | None = None):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
    for ri, row in enumerate(rows, 2):
        for col, h in enumerate(headers, 1):
            ws.cell(row=ri, column=col, value=row.get(h, ""))
    for i, h in enumerate(headers, 1):
        w = widths[i - 1] if widths else min(44, max(14, len(h) + 2))
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> None:
    for p in (DO_PRODUCTS, DO_VARIANTS, LS_EXPORT):
        if not p.exists():
            raise SystemExit(f"Missing {p} — refresh live dumps first")

    do_products = load_do_products()
    ls_products = load_ls_products()
    pairs, ls_only_prods, do_only_prods = match_products(do_products, ls_products)

    exact_all: list[dict] = []
    fuzzy_all: list[dict] = []
    do_only_all: list[dict] = []
    ls_only_all: list[dict] = []

    for lp, dp, _mt, _sc in pairs:
        ex, fz, lso, doo = match_variants_in_pair(lp, dp)
        exact_all.extend(ex)
        fuzzy_all.extend(fz)
        fuzzy_all.extend(lso)
        fuzzy_all.extend(doo)

    do_only_adj = [p for p in do_only_prods if p.slug not in DO_ONLY_EXCLUDE]
    ls_only_adj = [p for p in ls_only_prods if p.slug not in LS_ONLY_EXCLUDE]

    for p in do_only_prods:
        if p.slug not in DO_ONLY_EXCLUDE:
            do_only_all.extend(product_rows_side(p, "do"))
    for p in ls_only_prods:
        if p.slug not in LS_ONLY_EXCLUDE:
            ls_only_all.extend(product_rows_side(p, "ls"))

    summary = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "do_products": len(do_products),
        "ls_products": len(ls_products),
        "matched_product_pairs": len(pairs),
        "exact_match_variant_rows": len(exact_all),
        "fuzzy_match_variant_rows": len(fuzzy_all),
        "do_only_product_rows": len(do_only_all),
        "ls_only_product_rows": len(ls_only_all),
        "do_only_products_raw": len(do_only_prods),
        "ls_only_products_raw": len(ls_only_prods),
        "do_only_products_adjusted": len(do_only_adj),
        "ls_only_products_adjusted": len(ls_only_adj),
        "columns": COLS_6,
        "outputs": {"xlsx": str(OUT_XLSX), "json": str(OUT_JSON)},
    }
    OUT_JSON.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    wb = Workbook()
    sm = wb.active
    sm.title = "Summary"
    write_sheet(
        sm,
        ["Metric", "Count"],
        [
            {"Metric": "DO publish products", "Count": len(do_products)},
            {"Metric": "LS ACTIVE products", "Count": len(ls_products)},
            {"Metric": "Matched product pairs", "Count": len(pairs)},
            {"Metric": "Exact Match variant rows", "Count": len(exact_all)},
            {"Metric": "Fuzzy Match variant rows", "Count": len(fuzzy_all)},
            {"Metric": "DO Only variant rows", "Count": len(do_only_all)},
            {"Metric": "LS Only variant rows", "Count": len(ls_only_all)},
            {"Metric": "DO Only products (adjusted)", "Count": len(do_only_adj)},
            {"Metric": "LS Only products (adjusted)", "Count": len(ls_only_adj)},
            {"Metric": "Generated UTC", "Count": summary["generatedAt"]},
        ],
        [44, 14],
    )

    widths = [36, 36, 28, 28, 18, 18]
    ex = wb.create_sheet("Exact Match")
    write_sheet(ex, COLS_6, exact_all, widths)

    fz = wb.create_sheet("Fuzzy Match")
    write_sheet(fz, COLS_6, fuzzy_all, widths)

    do_ws = wb.create_sheet("DO Only")
    write_sheet(do_ws, COLS_6, do_only_all, widths)

    ls_ws = wb.create_sheet("LS Only")
    write_sheet(ls_ws, COLS_6, ls_only_all, widths)

    wb.save(OUT_XLSX)
    print(json.dumps(summary, indent=2))
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
