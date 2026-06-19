#!/usr/bin/env python3
"""
Export full WooCommerce catalog from WXR XML (products + variations).
No Zoho. SKU blank when empty or starts with "woo" (seed placeholders).

Usage:
  python3 backend/scripts/generate-woocommerce-catalog-report.py
  python3 backend/scripts/generate-woocommerce-catalog-report.py \\
    --products data/sarveda.WordPress.2026-05-29-products.xml \\
    --variations data/variations.xml \\
    --output data/sarveda-woocommerce-catalog.xlsx
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CDATA_RE = re.compile(r"<(\w+)><!\[CDATA\[([\s\S]*?)\]\]></\1>")
PLAIN_RE = re.compile(r"<(\w+)>([^<]*)</\1>")
META_RE = re.compile(
    r"<wp:meta_key><!\[CDATA\[([^\]]+)\]\]></wp:meta_key>\s*"
    r"<wp:meta_value><!\[CDATA\[([\s\S]*?)\]\]></wp:meta_value>"
)


def read_items(path: Path) -> list[str]:
    text = path.read_text(encoding="utf-8", errors="replace")
    return text.split("<item>")[1:]


def cdata(tag: str, block: str) -> str:
    m = re.search(rf"<{re.escape(tag)}><!\[CDATA\[([\s\S]*?)\]\]></{re.escape(tag)}>", block)
    if m:
        return m.group(1).strip()
    m = re.search(rf"<{re.escape(tag)}>([^<]*)</{re.escape(tag)}>", block)
    return (m.group(1).strip() if m else "")


def parse_meta(block: str) -> dict[str, str]:
    return {m.group(1): m.group(2) for m in META_RE.finditer(block)}


def parse_decimal(val: str | None) -> float | None:
    if val is None:
        return None
    s = str(val).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def sale_or_regular(meta: dict[str, str], sale_key: str, reg_key: str) -> float | None:
    return parse_decimal(meta.get(sale_key)) or parse_decimal(meta.get(reg_key))


def display_sku(raw: str | None) -> str:
    s = (raw or "").strip()
    if not s:
        return ""
    if s.lower().startswith("woo"):
        return ""
    return s


def strip_parent_variant(var_title: str, parent_name: str) -> str:
    var_title = (var_title or "").strip()
    parent_name = (parent_name or "").strip()
    if not var_title:
        return ""
    if parent_name and var_title.lower().startswith(parent_name.lower()):
        rest = var_title[len(parent_name) :].lstrip(" -–—|")
        return rest or var_title
    if " - " in var_title:
        return var_title.split(" - ", 1)[-1].strip()
    return var_title


def autosize(ws, max_width: int = 52) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        width = max(
            (min(len(str(c.value or "")) + 2, max_width) for c in col_cells if c.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = max(width, 10)


def write_sheet(wb: openpyxl.Workbook, title: str, headers: list[str], rows: list[dict]) -> None:
    ws = wb.create_sheet(title)
    header_fill = PatternFill("solid", fgColor="1E3A2F")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(["" if row.get(h) is None else row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    autosize(ws)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--products",
        default="data/sarveda.WordPress.2026-05-29-products.xml",
    )
    parser.add_argument("--variations", default="data/variations.xml")
    parser.add_argument("--output", default="data/sarveda-woocommerce-catalog.xlsx")
    args = parser.parse_args()

    root = Path(__file__).resolve().parents[2]
    products_path = root / args.products
    variations_path = root / args.variations
    output_path = root / args.output

    products: dict[str, dict] = {}
    for block in read_items(products_path):
        if "<wp:post_type><![CDATA[product]]></wp:post_type>" not in block:
            continue
        pid = cdata("wp:post_id", block)
        if not pid:
            continue
        meta = parse_meta(block)
        ptype = ""
        if 'nicename="variable"' in block:
            ptype = "variable"
        elif 'nicename="simple"' in block:
            ptype = "simple"
        elif 'nicename="grouped"' in block:
            ptype = "grouped"
        else:
            ptype = meta.get("_product_type") or "unknown"

        products[pid] = {
            "wc_product_id": pid,
            "product_name": cdata("title", block),
            "wc_status": cdata("wp:status", block),
            "product_type": ptype,
            "raw_sku": meta.get("_sku", ""),
            "slug": cdata("wp:post_name", block),
            "indian_price_inr": sale_or_regular(meta, "_india_sale_price", "_india_regular_price"),
            "us_price_usd": sale_or_regular(
                meta, "_dollars-zone_sale_price", "_dollars-zone_regular_price"
            ),
            "gbp_price": sale_or_regular(meta, "_zone-1_sale_price", "_zone-1_regular_price")
            or sale_or_regular(meta, "_zone-2_sale_price", "_zone-2_regular_price"),
        }

    catalog_rows: list[dict] = []

    # Simple / non-variable products as single rows
    for p in products.values():
        if p["product_type"] != "simple":
            continue
        catalog_rows.append(
            {
                "product_name": p["product_name"],
                "variant_name": "(simple product)",
                "sku": display_sku(p["raw_sku"]),
                "raw_sku_hidden": p["raw_sku"] or "",
                "indian_price_inr": p["indian_price_inr"] or "",
                "us_price_usd": p["us_price_usd"] or "",
                "gbp_price": p["gbp_price"] or "",
                "wc_status": p["wc_status"],
                "product_type": p["product_type"],
                "record_type": "simple_product",
                "wc_product_id": p["wc_product_id"],
                "wc_variation_id": "",
                "wc_parent_id": "",
                "is_orphan": "No",
                "slug": p["slug"],
            }
        )

    orphan_count = 0
    for block in read_items(variations_path):
        if "<wp:post_type><![CDATA[product_variation]]></wp:post_type>" not in block:
            continue
        vid = cdata("wp:post_id", block)
        parent_id = cdata("wp:post_parent", block)
        meta = parse_meta(block)
        raw_sku = meta.get("_sku", "")
        var_title = cdata("title", block)
        excerpt = cdata("excerpt:encoded", block)

        parent = products.get(parent_id)
        is_orphan = parent is None
        if is_orphan:
            orphan_count += 1
            product_name = f"(missing parent #{parent_id})"
            parent_name = ""
            wc_status = "orphan"
            product_type = "variation"
            slug = ""
        else:
            product_name = parent["product_name"]
            parent_name = parent["product_name"]
            wc_status = parent["wc_status"]
            product_type = parent["product_type"]
            slug = parent["slug"]

        variant_name = strip_parent_variant(var_title, parent_name)
        if (not variant_name or variant_name.lower() == (parent_name or "").lower()) and excerpt:
            parts = []
            clean = excerpt.replace("<p>", "").replace("</p>", "")
            for part in clean.split(","):
                part = part.strip()
                if ":" in part:
                    parts.append(part.split(":", 1)[1].strip())
                elif part:
                    parts.append(part)
            if parts:
                variant_name = ", ".join(parts)
        if not variant_name:
            variant_name = var_title or "(variation)"

        catalog_rows.append(
            {
                "product_name": product_name,
                "variant_name": variant_name,
                "sku": display_sku(raw_sku),
                "raw_sku_hidden": raw_sku or "",
                "indian_price_inr": sale_or_regular(meta, "_india_sale_price", "_india_regular_price")
                or "",
                "us_price_usd": sale_or_regular(
                    meta, "_dollars-zone_sale_price", "_dollars-zone_regular_price"
                )
                or "",
                "gbp_price": sale_or_regular(meta, "_zone-1_sale_price", "_zone-1_regular_price")
                or sale_or_regular(meta, "_zone-2_sale_price", "_zone-2_regular_price")
                or "",
                "wc_status": wc_status,
                "product_type": product_type,
                "record_type": "orphan_variation" if is_orphan else "variation",
                "wc_product_id": parent_id if not is_orphan else "",
                "wc_variation_id": vid,
                "wc_parent_id": parent_id,
                "is_orphan": "Yes" if is_orphan else "No",
                "slug": slug,
            }
        )

    catalog_rows.sort(
        key=lambda r: (
            str(r["product_name"]).lower(),
            str(r["variant_name"]).lower(),
            str(r["wc_variation_id"]),
        )
    )

    product_summary = []
    for p in sorted(products.values(), key=lambda x: x["product_name"].lower()):
        var_count = sum(
            1
            for r in catalog_rows
            if r["wc_product_id"] == p["wc_product_id"] and r["record_type"] == "variation"
        )
        product_summary.append(
            {
                "product_name": p["product_name"],
                "wc_product_id": p["wc_product_id"],
                "wc_status": p["wc_status"],
                "product_type": p["product_type"],
                "parent_sku": display_sku(p["raw_sku"]),
                "variation_count_in_xml": var_count if p["product_type"] == "variable" else 0,
                "slug": p["slug"],
            }
        )

    headers = [
        "product_name",
        "variant_name",
        "sku",
        "indian_price_inr",
        "us_price_usd",
        "gbp_price",
        "wc_status",
        "product_type",
        "record_type",
        "is_orphan",
        "wc_product_id",
        "wc_variation_id",
        "slug",
        "raw_sku_hidden",
    ]

    public_headers = [
        "product_name",
        "variant_name",
        "sku",
        "indian_price_inr",
        "us_price_usd",
        "gbp_price",
        "wc_status",
        "product_type",
        "record_type",
        "is_orphan",
        "wc_product_id",
        "wc_variation_id",
        "slug",
    ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Summary", 0)
    published = [r for r in catalog_rows if r["wc_status"] == "publish"]
    draft = [r for r in catalog_rows if r["wc_status"] == "draft"]
    orphans = [r for r in catalog_rows if r["is_orphan"] == "Yes"]

    summary_rows = [
        ("WooCommerce products (parent)", len(products)),
        ("Published products", sum(1 for p in products.values() if p["wc_status"] == "publish")),
        ("Draft products", sum(1 for p in products.values() if p["wc_status"] == "draft")),
        ("Variable products", sum(1 for p in products.values() if p["product_type"] == "variable")),
        ("Simple products", sum(1 for p in products.values() if p["product_type"] == "simple")),
        ("Total catalog rows (variants + simple)", len(catalog_rows)),
        ("Variation rows", sum(1 for r in catalog_rows if r["record_type"] == "variation")),
        ("Orphan variation rows", len(orphans)),
        ("Simple product rows", sum(1 for r in catalog_rows if r["record_type"] == "simple_product")),
        ("Rows with human SKU", sum(1 for r in catalog_rows if r["sku"])),
        ("Rows with blank SKU", sum(1 for r in catalog_rows if not r["sku"])),
        ("Products XML", products_path.name),
        ("Variations XML", variations_path.name),
    ]
    summary.append(("Metric", "Value"))
    for k, v in summary_rows:
        summary.append((k, v))
    autosize(summary)

    write_sheet(wb, "All WC Catalog", public_headers, catalog_rows)
    write_sheet(wb, "Published", public_headers, published)
    write_sheet(wb, "Draft", public_headers, draft)
    write_sheet(wb, "Orphans", public_headers, orphans)
    write_sheet(
        wb,
        "Products (170)",
        list(product_summary[0].keys()) if product_summary else ["product_name"],
        product_summary,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Wrote {output_path}")
    print(f"  Products: {len(products)}")
    print(f"  Catalog rows: {len(catalog_rows)} (variations + simple)")
    print(f"  Orphans: {len(orphans)}")
    print(f"  Published rows: {len(published)} | Draft rows: {len(draft)}")


if __name__ == "__main__":
    main()
