#!/usr/bin/env python3
"""
DO vs Lightsail — pending launch list (post copper-bottle split).

Sheets:
  Summary
  Exact Match OK          — slug/woo match, same variant count
  Exact Match Var Mismatch  — slug/woo match, variant count differs
  Fuzzy Match             — name fuzzy match (≥82), with variant counts
  DO Only                 — publish on sarveda.com, no LS product match
  LS Only                 — ACTIVE on demo, no DO product match
  Pending All             — union of mismatches + one-sided (action list)

Adjustments (copper artistic split, Aug 2026):
  - DO printed-copper-water-bottles → not counted DO-only (5 designs on LS)
  - LS copper-bottle-blue-tranquillity-meditation → not counted LS-only (variant of DO umbrella)

Inputs (refresh same session):
  data/compare/do_products.csv
  data/compare/do_variants.csv
  data/compare/lightsail-catalog-export.json

Usage:
  python3 backend/scripts/compare-do-ls-pending-launch.py
"""
from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

ROOT = Path(__file__).resolve().parents[2]
DO_PRODUCTS = ROOT / "data/compare/do_products.csv"
DO_VARIANTS = ROOT / "data/compare/do_variants.csv"
LS_EXPORT = ROOT / "data/compare/lightsail-catalog-export.json"
OUT_XLSX = ROOT / "data/compare/pending-do-vs-ls-launch.xlsx"
OUT_JSON = ROOT / "data/compare/pending-do-vs-ls-launch-summary.json"

FUZZY_MIN = 82

# Resolved by copper split fix — exclude from "only" pending counts
DO_ONLY_EXCLUDE_SLUGS = {"printed-copper-water-bottles"}
LS_ONLY_EXCLUDE_SLUGS = {"copper-bottle-blue-tranquillity-meditation"}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN = PatternFill("solid", fgColor="E2EFDA")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
RED = PatternFill("solid", fgColor="FCE4D6")
BLUE = PatternFill("solid", fgColor="DDEBF7")
ORANGE = PatternFill("solid", fgColor="FCE4D6")


@dataclass
class Product:
    source: str
    key: str
    slug: str
    name: str
    woo_id: int | None
    variant_count: int = 0


def norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def is_shop_slug(slug: str) -> bool:
    s = (slug or "").lower()
    return not (s.startswith("course-checkout-") or s.startswith("event-checkout-"))


def load_do_var_counts() -> dict[int, int]:
    counts: dict[int, int] = {}
    with DO_VARIANTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            pid = int(r["parent_id"])
            counts[pid] = counts.get(pid, 0) + 1
    return counts


def load_do(do_var: dict[int, int]) -> list[Product]:
    out: list[Product] = []
    with DO_PRODUCTS.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (r.get("status") or "").lower() != "publish":
                continue
            slug = (r.get("slug") or "").strip()
            if not is_shop_slug(slug):
                continue
            pid = int(r["id"])
            ptype = (r.get("product_type") or "").lower()
            vc = do_var.get(pid, 0) if ptype == "variable" else 1
            if ptype != "variable":
                vc = 1
            elif vc == 0:
                vc = 1
            out.append(
                Product(
                    source="do",
                    key=f"do:{pid}",
                    slug=slug,
                    name=(r.get("name") or "").strip(),
                    woo_id=pid,
                    variant_count=vc,
                )
            )
    return out


def load_ls() -> list[Product]:
    data = json.loads(LS_EXPORT.read_text(encoding="utf-8"))
    grouped: dict[str, Product] = {}
    for r in data["rows"]:
        slug = (r.get("slug") or "").strip()
        if not is_shop_slug(slug):
            continue
        if slug not in grouped:
            grouped[slug] = Product(
                source="ls",
                key=f"ls:{slug}",
                slug=slug,
                name=(r.get("name") or "").strip(),
                woo_id=int(r["wooCommerceId"]) if r.get("wooCommerceId") else None,
                variant_count=0,
            )
        grouped[slug].variant_count += 1
    return sorted(grouped.values(), key=lambda p: p.slug)


def fuzzy_score(a: Product, b: Product) -> float:
    if a.woo_id and b.woo_id and a.woo_id == b.woo_id:
        return 100.0
    if norm_text(a.slug) == norm_text(b.slug):
        return 99.0
    return float(fuzz.token_sort_ratio(norm_text(a.name), norm_text(b.name)))


def match_products(do_products: list[Product], ls_products: list[Product]):
    ls_by_slug = {norm_text(p.slug): p for p in ls_products if p.slug}
    do_by_slug = {norm_text(p.slug): p for p in do_products if p.slug}
    ls_by_woo = {p.woo_id: p for p in ls_products if p.woo_id}
    do_by_woo = {p.woo_id: p for p in do_products if p.woo_id}

    used_ls: set[str] = set()
    used_do: set[str] = set()
    exact_slug: list[tuple[Product, Product, str]] = []
    exact_woo: list[tuple[Product, Product, str]] = []

    for lp in ls_products:
        ns = norm_text(lp.slug)
        if not ns or ns not in do_by_slug:
            continue
        dp = do_by_slug[ns]
        if dp.key in used_do or lp.key in used_ls:
            continue
        exact_slug.append((lp, dp, "exact_slug"))
        used_ls.add(lp.key)
        used_do.add(dp.key)

    for lp in ls_products:
        if lp.key in used_ls:
            continue
        if lp.woo_id and lp.woo_id in do_by_woo:
            dp = do_by_woo[lp.woo_id]
            if dp.key in used_do:
                continue
            exact_woo.append((lp, dp, "exact_woo_id"))
            used_ls.add(lp.key)
            used_do.add(dp.key)

    ls_rem = [p for p in ls_products if p.key not in used_ls]
    do_rem = [p for p in do_products if p.key not in used_do]

    candidates: list[tuple[float, int, int]] = []
    for li, lp in enumerate(ls_rem):
        for di, dp in enumerate(do_rem):
            s = fuzzy_score(lp, dp)
            if s >= FUZZY_MIN:
                candidates.append((s, li, di))
    candidates.sort(reverse=True)

    used_lr: set[int] = set()
    used_dr: set[int] = set()
    fuzzy: list[tuple[Product, Product, float, str]] = []
    for s, li, di in candidates:
        if li in used_lr or di in used_dr:
            continue
        used_lr.add(li)
        used_dr.add(di)
        fuzzy.append((ls_rem[li], do_rem[di], s, "fuzzy_name"))
        used_ls.add(ls_rem[li].key)
        used_do.add(do_rem[di].key)

    ls_only = [p for p in ls_products if p.key not in used_ls]
    do_only = [p for p in do_products if p.key not in used_do]
    return exact_slug, exact_woo, fuzzy, ls_only, do_only


def write_sheet(ws, headers: list[str], rows: list[list], widths: list[int] | None = None):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
    for ri, row in enumerate(rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=ri, column=col, value=val)
    for i, h in enumerate(headers, 1):
        w = widths[i - 1] if widths else min(48, max(12, len(h) + 2))
        ws.column_dimensions[get_column_letter(i)].width = w


def pair_row(match_type: str, score: float, lp: Product, dp: Product) -> list:
    delta = lp.variant_count - dp.variant_count
    status = "OK" if delta == 0 else "MISMATCH"
    return [
        match_type,
        round(score, 1),
        status,
        lp.slug,
        dp.slug,
        lp.name,
        dp.name,
        lp.woo_id or "",
        dp.woo_id or "",
        lp.variant_count,
        dp.variant_count,
        delta,
    ]


def main() -> None:
    for p in (DO_PRODUCTS, DO_VARIANTS, LS_EXPORT):
        if not p.exists():
            raise SystemExit(f"Missing {p} — refresh live dumps first")

    do_var = load_do_var_counts()
    do_products = load_do(do_var)
    ls_products = load_ls()
    exact_slug, exact_woo, fuzzy, ls_only, do_only = match_products(do_products, ls_products)

    all_exact = [(a, b, t, 99.0 if t == "exact_slug" else 100.0) for a, b, t in exact_slug]
    all_exact += [(a, b, t, 100.0) for a, b, t in exact_woo]
    exact_ok = [pair_row(t, s, lp, dp) for lp, dp, t, s in all_exact if lp.variant_count == dp.variant_count]
    exact_mismatch = [pair_row(t, s, lp, dp) for lp, dp, t, s in all_exact if lp.variant_count != dp.variant_count]
    fuzzy_rows = [pair_row(t, s, lp, dp) for lp, dp, s, t in fuzzy]

    do_only_rows = [
        [p.slug, p.name, p.woo_id or "", p.variant_count, "YES" if p.slug in DO_ONLY_EXCLUDE_SLUGS else ""]
        for p in sorted(do_only, key=lambda x: x.name.lower())
    ]
    ls_only_rows = [
        [p.slug, p.name, p.woo_id or "", p.variant_count, "YES" if p.slug in LS_ONLY_EXCLUDE_SLUGS else ""]
        for p in sorted(ls_only, key=lambda x: x.name.lower())
    ]

    do_only_adj = [p for p in do_only if p.slug not in DO_ONLY_EXCLUDE_SLUGS]
    ls_only_adj = [p for p in ls_only if p.slug not in LS_ONLY_EXCLUDE_SLUGS]

    pending: list[list] = []
    for row in exact_mismatch:
        pending.append(["exact_variant_mismatch", *row])
    for row in fuzzy_rows:
        pending.append(["fuzzy_match", *row])
    for p in do_only_adj:
        pending.append(
            [
                "do_only",
                "",
                "DO ONLY",
                "",
                "",
                p.slug,
                "",
                p.name,
                "",
                p.woo_id or "",
                "",
                p.variant_count,
                "",
            ]
        )
    for p in ls_only_adj:
        pending.append(
            [
                "ls_only",
                "",
                "LS ONLY",
                "",
                p.slug,
                "",
                p.name,
                "",
                p.woo_id or "",
                p.variant_count,
                "",
                "",
                "",
            ]
        )

    summary = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "do_products_publish": len(do_products),
        "lightsail_products_active": len(ls_products),
        "exact_match_variant_ok": len(exact_ok),
        "exact_match_variant_mismatch": len(exact_mismatch),
        "fuzzy_match_pairs": len(fuzzy_rows),
        "do_only_raw": len(do_only),
        "ls_only_raw": len(ls_only),
        "do_only_adjusted": len(do_only_adj),
        "ls_only_adjusted": len(ls_only_adj),
        "pending_action_items": len(pending),
        "copper_split_excluded": {
            "do_only": sorted(DO_ONLY_EXCLUDE_SLUGS),
            "ls_only": sorted(LS_ONLY_EXCLUDE_SLUGS),
        },
        "inputs": {
            "do_products": str(DO_PRODUCTS),
            "do_variants": str(DO_VARIANTS),
            "lightsail_export": str(LS_EXPORT),
        },
        "outputs": {"xlsx": str(OUT_XLSX), "json": str(OUT_JSON)},
    }
    OUT_JSON.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    wb = Workbook()
    sm = wb.active
    sm.title = "Summary"
    write_sheet(
        sm,
        ["Metric", "Count"],
        [
            ["DO sarveda.com publish products", len(do_products)],
            ["Lightsail ACTIVE products", len(ls_products)],
            ["", ""],
            ["Exact match — variants OK", len(exact_ok)],
            ["Exact match — variant COUNT mismatch", len(exact_mismatch)],
            ["Fuzzy name match pairs", len(fuzzy_rows)],
            ["", ""],
            ["DO ONLY (raw)", len(do_only)],
            ["LS ONLY (raw)", len(ls_only)],
            ["DO ONLY (adjusted, excl. copper umbrella)", len(do_only_adj)],
            ["LS ONLY (adjusted, excl. blue tranquillity)", len(ls_only_adj)],
            ["", ""],
            ["TOTAL pending action rows", len(pending)],
            ["Generated UTC", summary["generatedAt"]],
        ],
        [48, 12],
    )

    headers = [
        "Match type",
        "Score",
        "Variant status",
        "LS slug",
        "DO slug",
        "LS name",
        "DO name",
        "LS wooId",
        "DO id",
        "LS variants",
        "DO variants",
        "Delta (LS-DO)",
    ]

    ok_ws = wb.create_sheet("Exact Match OK")
    write_sheet(ok_ws, headers, exact_ok, [14, 8, 12, 28, 28, 36, 36, 10, 10, 10, 10, 8])
    for ri in range(2, ok_ws.max_row + 1):
        ok_ws.cell(row=ri, column=3).fill = GREEN

    mm_ws = wb.create_sheet("Exact Match Var Mismatch")
    write_sheet(mm_ws, headers, exact_mismatch, [14, 8, 12, 28, 28, 36, 36, 10, 10, 10, 10, 8])
    for ri in range(2, mm_ws.max_row + 1):
        mm_ws.cell(row=ri, column=3).fill = YELLOW

    fz_ws = wb.create_sheet("Fuzzy Match")
    write_sheet(fz_ws, headers, fuzzy_rows, [14, 8, 12, 28, 28, 36, 36, 10, 10, 10, 10, 8])
    for ri in range(2, fz_ws.max_row + 1):
        fz_ws.cell(row=ri, column=1).fill = BLUE

    do_ws = wb.create_sheet("DO Only")
    write_sheet(
        do_ws,
        ["DO slug", "DO name", "DO id", "DO variants", "Copper split exclude"],
        do_only_rows,
        [32, 44, 10, 12, 18],
    )
    for ri in range(2, do_ws.max_row + 1):
        if do_ws.cell(row=ri, column=5).value == "YES":
            do_ws.cell(row=ri, column=1).fill = GREEN
        else:
            do_ws.cell(row=ri, column=1).fill = YELLOW

    ls_ws = wb.create_sheet("LS Only")
    write_sheet(
        ls_ws,
        ["LS slug", "LS name", "wooCommerceId", "LS variants", "Copper split exclude"],
        ls_only_rows,
        [32, 44, 14, 12, 18],
    )
    for ri in range(2, ls_ws.max_row + 1):
        if ls_ws.cell(row=ri, column=5).value == "YES":
            ls_ws.cell(row=ri, column=1).fill = GREEN
        else:
            ls_ws.cell(row=ri, column=1).fill = RED

    pend_headers = ["Pending category"] + headers
    pend_ws = wb.create_sheet("Pending All")
    write_sheet(pend_ws, pend_headers, pending, [18, 14, 8, 12, 28, 28, 36, 36, 10, 10, 10, 10, 8])

    wb.save(OUT_XLSX)
    print(json.dumps(summary, indent=2))
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
