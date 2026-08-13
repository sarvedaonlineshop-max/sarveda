#!/usr/bin/env python3
"""
Build reconcile plan from latest-inventory-fuzzy.xlsx Sheet Only + DB Only tabs.

Outputs: data/compare/sheet-db-only-plan.json

Usage:
  python3 backend/scripts/build-sheet-db-only-plan.py
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
FUZZY = ROOT / "data/compare/latest-inventory-fuzzy.xlsx"
MASTER = ROOT / "data/latest_inventory.xlsx"
OUT = ROOT / "data/compare/sheet-db-only-plan.json"

# Sheet SKU -> existing Lightsail SKU (same variant; rename DB to sheet)
SKU_RENAME_TO_SHEET = {
    "CB-BM": "CB-AD-BM",
    "CB-BM-B": "CB-AD-BM-B",
    "CB-BM-.5": "CB-AD-BM-.5",
    "CB-BM-B-.5": "CB-AD-BM-.5-B",
    "CB-C-V": "CB-CDG-V",
    "ME-CZ-C-R": "ME-CZ-C-RP",
}

# Sheet-only SKU -> add variant on this product slug (not a new product)
# Slugs match stage2 rename: zafu-zabuton-combo-plain / zafu-zabuton-combo-lotus-embroidery
CREATE_VARIANT_ON = {
    "ME-Z-Zn-DG": "zafu-zabuton-combo-plain",
    "ME-Z-Zn-EM-L-NB": "zafu-zabuton-combo-lotus-embroidery",
    "ME-Z-Zn-EM-L-DG": "zafu-zabuton-combo-lotus-embroidery",
    "ME-CZ-C-LG": "crescent-zafu-cushion-compact-buck-wheat",
    "Me-CZ-W-S": "crescent-zafu-cushion-wide-cotton",
    "Me-CZ-W-DG": "crescent-zafu-cushion-wide-cotton",
    "MI-SB-SS-GO-7": "sacred-symbols-singing-bowls",
}

# Truly new simple products (create on Lightsail)
CREATE_PRODUCT_SKUS = {
    "MI-SB-DDT-5.5",
    "MI-SK-DD",
    "MI-SK-SM",
    "MI-SK-ME",
    "MI-SK-LE",
    "MI-SF",
    "MI-CW",
    "MI-OF",
    "MI-MT-BL",
    "MI-GO-NG",
}


def norm_sku(s: str) -> str:
    return " ".join(str(s or "").strip().split()).upper()


def load_fuzzy_sheet(name: str) -> list[dict]:
    wb = load_workbook(FUZZY, read_only=True, data_only=True)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h or "").strip() for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(r):
            continue
        out.append({headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))})
    wb.close()
    return out


def load_master_rows() -> dict[str, dict]:
    wb = load_workbook(MASTER, read_only=True, data_only=True)
    ws = wb["Website Catalog"]
    rows = list(ws.iter_rows(values_only=True))
    header_i = next(i for i, r in enumerate(rows[:20]) if r and str(r[0] or "").strip() == "Name")
    headers = [str(c or "").strip() for c in rows[header_i]]
    by_sku: dict[str, dict] = {}
    cur = ""
    for r in rows[header_i + 1 :]:
        if not r:
            continue
        if r[0] is not None and str(r[0]).strip():
            cur = str(r[0]).strip()
        sku = "" if len(r) < 3 or r[2] is None else str(r[2]).strip()
        if not sku:
            continue
        rec = {"productName": cur, "variantName": "" if r[1] is None else str(r[1]).strip(), "sku": sku}
        for i, h in enumerate(headers):
            if i < len(r):
                rec[h] = r[i]
        by_sku[norm_sku(sku)] = rec
    wb.close()
    return by_sku


def main() -> None:
    sheet_only = load_fuzzy_sheet("Sheet Only")
    db_only = load_fuzzy_sheet("DB Only")
    master = load_master_rows()
    master_skus = set(master.keys())

    rename_skus = []
    resolved_sheet = []
    create_variants = []
    create_products = []
    draft_skus = []

    sheet_by_sku = {norm_sku(str(r.get("Sheet SKU") or "")): r for r in sheet_only if r.get("Sheet SKU")}
    db_only_skus = {norm_sku(str(r.get("DB SKU") or "")) for r in db_only if r.get("DB SKU")}

    for db_sku_raw, to_sku in SKU_RENAME_TO_SHEET.items():
        db_sku = norm_sku(db_sku_raw)
        to = norm_sku(to_sku)
        if to in sheet_by_sku or db_sku in db_only_skus:
            rename_skus.append({"fromSku": db_sku_raw, "toSku": to_sku, "reason": "Sheet SKU is source of truth"})
            resolved_sheet.append(to_sku)

    for sku, slug in CREATE_VARIANT_ON.items():
        ns = norm_sku(sku)
        if ns in sheet_by_sku:
            m = master.get(ns, {})
            create_variants.append(
                {
                    "productSlug": slug,
                    "sku": m.get("sku") or sku,
                    "variantName": m.get("variantName") or sheet_by_sku[ns].get("Sheet Variant name") or "",
                    "productName": m.get("productName") or sheet_by_sku[ns].get("Sheet Product name") or "",
                    "reason": "Missing variant on existing product",
                }
            )
            resolved_sheet.append(m.get("sku") or sku)

    for sku in CREATE_PRODUCT_SKUS:
        ns = norm_sku(sku)
        if ns in sheet_by_sku:
            m = master.get(ns, {})
            create_products.append(
                {
                    "sku": m.get("sku") or sku,
                    "productName": m.get("productName") or sheet_by_sku[ns].get("Sheet Product name") or "",
                    "variantName": m.get("variantName") or sheet_by_sku[ns].get("Sheet Variant name") or "",
                    "reason": "Not on Lightsail — create simple product",
                }
            )
            resolved_sheet.append(m.get("sku") or sku)

    rename_from = {norm_sku(x["fromSku"]) for x in rename_skus}
    for row in db_only:
        db_sku = str(row.get("DB SKU") or "").strip()
        if not db_sku:
            continue
        ns = norm_sku(db_sku)
        if ns in rename_from:
            continue
        if ns in master_skus:
            continue
        draft_skus.append(
            {
                "sku": db_sku,
                "productName": row.get("DB Product name") or "",
                "variantName": row.get("DB Variant name") or "",
                "reason": "On Lightsail but not on team sheet — draft variant",
            }
        )

    remaining_sheet = [
        r
        for r in sheet_only
        if norm_sku(str(r.get("Sheet SKU") or "")) not in {norm_sku(s) for s in resolved_sheet}
        and norm_sku(str(r.get("Sheet SKU") or "")) not in {norm_sku(x["toSku"]) for x in rename_skus}
    ]

    plan = {
        "source": str(FUZZY),
        "summary": {
            "sheet_only_input": len(sheet_only),
            "db_only_input": len(db_only),
            "rename_skus": len(rename_skus),
            "create_variants": len(create_variants),
            "create_products": len(create_products),
            "draft_variants": len(draft_skus),
            "remaining_sheet_only": len(remaining_sheet),
        },
        "rename_skus": rename_skus,
        "create_variants": create_variants,
        "create_products": create_products,
        "draft_variants": draft_skus,
        "remaining_sheet_only": [
            {
                "sku": r.get("Sheet SKU"),
                "productName": r.get("Sheet Product name"),
                "variantName": r.get("Sheet Variant name"),
            }
            for r in remaining_sheet
        ],
    }

    OUT.write_text(json.dumps(plan, indent=2) + "\n")
    print(json.dumps(plan["summary"], indent=2))
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
