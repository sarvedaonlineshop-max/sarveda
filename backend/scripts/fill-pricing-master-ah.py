#!/usr/bin/env python3
"""Fill column AH (Existing Sale Price) from staging DB via public API."""

from __future__ import annotations

import json
import sys
import time
import urllib.error
import urllib.request
from collections import defaultdict
from pathlib import Path

import openpyxl

API_BASE = "https://sarveda-demo.xyz/api/products"
XLSX_PATH = Path(__file__).resolve().parents[2] / "data" / "Sarveda Product Pricing MASTER July 26.xlsx"
SHEET = "Website Catalog"
SKU_COL = 3
AH_COL = 34
DATA_START_ROW = 4
TOLERANCE = 0.01


def api_get(url: str, retries: int = 3) -> dict:
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(url, timeout=60) as resp:
                return json.loads(resp.read().decode())
        except (urllib.error.URLError, TimeoutError) as err:
            if attempt == retries - 1:
                raise
            time.sleep(1.5 * (attempt + 1))
    raise RuntimeError("unreachable")


def fetch_all_variant_prices() -> dict[str, float]:
    """SKU -> sale price in INR (rupees, from saleInPaise)."""
    page = 1
    slugs: list[str] = []
    while True:
        payload = api_get(f"{API_BASE}?page={page}&limit=100")
        items = payload["data"]["items"]
        slugs.extend(i["slug"] for i in items)
        pag = payload["data"]["pagination"]
        if page >= pag["totalPages"]:
            break
        page += 1

    print(f"Fetched {len(slugs)} product slugs from API")

    sku_prices: dict[str, float] = {}
    dup_db: dict[str, list[float]] = defaultdict(list)

    for idx, slug in enumerate(slugs, 1):
        if idx % 20 == 0:
            print(f"  product {idx}/{len(slugs)} …")
        detail = api_get(f"{API_BASE}/{slug}")
        product = detail["data"]["product"]
        for v in product.get("variants", []):
            sku = str(v["sku"]).strip()
            if not sku:
                continue
            price = round(v["saleInPaise"] / 100, 2)
            dup_db[sku].append(price)
            sku_prices[sku] = price

    db_duplicates = {sku: prices for sku, prices in dup_db.items() if len(prices) > 1}
    unique_conflicts = {
        sku: prices for sku, prices in db_duplicates.items() if len(set(prices)) > 1
    }
    if unique_conflicts:
        print("WARNING: DB SKU price conflicts:", len(unique_conflicts))

    return sku_prices, db_duplicates, unique_conflicts


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"Missing file: {XLSX_PATH}", file=sys.stderr)
        return 1

    print("Loading variant prices from staging API …")
    sku_prices, db_duplicates, db_price_conflicts = fetch_all_variant_prices()
    print(f"DB variants with SKU: {len(sku_prices)}")

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET]

    filled = 0
    already_correct = 0
    no_sku = 0
    sku_not_in_db = []
    conflicts = []
    excel_dup_skus: dict[str, list[int]] = defaultdict(list)

    max_row = ws.max_row
    for row in range(DATA_START_ROW, max_row + 1):
        raw_sku = ws.cell(row, SKU_COL).value
        if raw_sku is None or str(raw_sku).strip() == "":
            no_sku += 1
            continue
        sku = str(raw_sku).strip()
        excel_dup_skus[sku].append(row)

        if sku not in sku_prices:
            sku_not_in_db.append({"row": row, "sku": sku})
            continue

        if sku in db_price_conflicts:
            conflicts.append({
                "row": row,
                "sku": sku,
                "reason": "duplicate_sku_in_db_with_different_prices",
                "db_prices": db_price_conflicts[sku],
            })
            continue

        db_price = sku_prices[sku]
        existing = ws.cell(row, AH_COL).value

        if existing is not None and existing != "":
            try:
                existing_num = float(existing)
            except (TypeError, ValueError):
                conflicts.append({
                    "row": row,
                    "sku": sku,
                    "reason": "existing_ah_not_numeric",
                    "existing": existing,
                    "db_price": db_price,
                })
                continue
            if abs(existing_num - db_price) > TOLERANCE:
                conflicts.append({
                    "row": row,
                    "sku": sku,
                    "reason": "existing_ah_differs_from_db",
                    "existing": existing_num,
                    "db_price": db_price,
                })
                continue
            already_correct += 1
            continue

        ws.cell(row, AH_COL).value = db_price
        filled += 1

    excel_duplicate_rows = {
        sku: rows for sku, rows in excel_dup_skus.items() if len(rows) > 1
    }

    report = {
        "db_variant_count": len(sku_prices),
        "excel_rows_with_sku": sum(len(v) for v in excel_dup_skus.values()),
        "filled_ah": filled,
        "already_had_matching_price": already_correct,
        "rows_without_sku": no_sku,
        "sku_not_in_db_count": len(sku_not_in_db),
        "sku_not_in_db": sku_not_in_db,
        "conflicts_count": len(conflicts),
        "conflicts": conflicts,
        "excel_duplicate_sku_count": len(excel_duplicate_rows),
        "excel_duplicate_skus": excel_duplicate_rows,
        "db_duplicate_sku_with_different_prices": db_price_conflicts,
    }

    report_path = XLSX_PATH.parent / "pricing-master-ah-fill-report.json"
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    wb.save(XLSX_PATH)
    wb.close()

    print("\n=== SUMMARY ===")
    print(f"DB variants (by SKU):     {report['db_variant_count']}")
    print(f"Excel rows with SKU:      {report['excel_rows_with_sku']}")
    print(f"AH filled from DB:        {report['filled_ah']}")
    print(f"Already matched (skipped): {report['already_had_matching_price']}")
    print(f"No SKU in row:            {report['rows_without_sku']}")
    print(f"SKU not in DB:            {report['sku_not_in_db_count']}")
    print(f"Conflicts (left blank):   {report['conflicts_count']}")
    print(f"Duplicate SKUs in Excel:  {report['excel_duplicate_sku_count']}")
    print(f"Report: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
