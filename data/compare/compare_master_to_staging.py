#!/usr/bin/env python3
"""Compare data/Sarveda MASTER.xlsx to a staging variants CSV (SKU first, then name)."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook


def norm_sku(s: str) -> str:
    return " ".join(str(s).strip().split()).upper()


def norm_name(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def load_master(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_i = next(
        i for i, r in enumerate(rows[:20]) if r and str(r[0] or "").strip() == "Name"
    )
    by_name: dict[str, list] = defaultdict(list)
    sku_to_name: dict[str, str] = {}
    cur = None
    for r in rows[header_i + 1 :]:
        if not r:
            continue
        if r[0] is not None and str(r[0]).strip():
            cur = str(r[0]).strip()
        sku = r[2] if len(r) > 2 else None
        if not cur or sku is None or not str(sku).strip():
            continue
        sn = norm_sku(sku)
        by_name[cur].append(sn)
        sku_to_name[sn] = cur
    return by_name, sku_to_name


def load_staging_csv(path: Path, *, include_drafts: bool = False):
    by_name: dict[str, list] = defaultdict(list)
    sku_to_name: dict[str, str] = {}
    skipped_draft = 0
    with path.open(newline="") as f:
        for row in csv.DictReader(f):
            status = (row.get("product_status") or "").strip().upper()
            if not include_drafts and status == "DRAFT":
                skipped_draft += 1
                continue
            name = row["product_name"].strip()
            sn = norm_sku(row["sku"])
            by_name[name].append(sn)
            sku_to_name[sn] = name
    return by_name, sku_to_name, skipped_draft


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--master", type=Path, default=Path("data/Sarveda MASTER.xlsx"))
    ap.add_argument("--staging", type=Path, required=True)
    ap.add_argument("--out-dir", type=Path, default=Path("data/compare"))
    ap.add_argument(
        "--include-drafts",
        action="store_true",
        help="Include DRAFT products from staging (default: exclude)",
    )
    args = ap.parse_args()
    args.out_dir.mkdir(parents=True, exist_ok=True)

    sheet_by_name, sheet_sku = load_master(args.master)
    db_by_name, db_sku, skipped_draft = load_staging_csv(
        args.staging, include_drafts=args.include_drafts
    )

    sheet_to_db: dict[str, str] = {}
    db_to_sheet: dict[str, str] = {}
    for sku in set(sheet_sku) & set(db_sku):
        sn, dn = sheet_sku[sku], db_sku[sku]
        if sn not in sheet_to_db and dn not in db_to_sheet:
            sheet_to_db[sn] = dn
            db_to_sheet[dn] = sn

    sheet_norm = {norm_name(n): n for n in sheet_by_name}
    db_norm = {norm_name(n): n for n in db_by_name}
    for snorm, sn in sheet_norm.items():
        if sn in sheet_to_db:
            continue
        if snorm in db_norm and db_norm[snorm] not in db_to_sheet:
            sheet_to_db[sn] = db_norm[snorm]
            db_to_sheet[db_norm[snorm]] = sn

    unmatched_sheet = [n for n in sheet_by_name if n not in sheet_to_db]
    unmatched_db = [n for n in db_by_name if n not in db_to_sheet]
    for sn in list(unmatched_sheet):
        best, best_sc = None, 0.0
        for dn in unmatched_db:
            sc = SequenceMatcher(None, norm_name(sn), norm_name(dn)).ratio()
            if sc > best_sc:
                best, best_sc = dn, sc
        if best and best_sc >= 0.92:
            sheet_to_db[sn] = best
            db_to_sheet[best] = sn
            unmatched_sheet.remove(sn)
            unmatched_db.remove(best)

    exact, partial = [], []
    for sn, dn in sheet_to_db.items():
        s_set, d_set = set(sheet_by_name[sn]), set(db_by_name[dn])
        rec = {
            "sheet_name": sn,
            "db_name": dn,
            "sheet_n": len(s_set),
            "db_n": len(d_set),
            "shared": len(s_set & d_set),
            "only_sheet": len(s_set - d_set),
            "only_db": len(d_set - s_set),
        }
        (exact if s_set == d_set else partial).append(rec)

    only_sheet = sorted(n for n in sheet_by_name if n not in sheet_to_db)
    only_db = sorted(n for n in db_by_name if n not in db_to_sheet)
    sku_exact = set(sheet_sku) & set(db_sku)

    summary = {
        "master_file": str(args.master),
        "include_drafts": args.include_drafts,
        "skipped_draft_variant_rows": skipped_draft,
        "sheet_products": len(sheet_by_name),
        "sheet_skus": len(sheet_sku),
        "db_products": len(db_by_name),
        "db_skus": len(db_sku),
        "sku_exact": len(sku_exact),
        "sku_only_sheet": len(set(sheet_sku) - set(db_sku)),
        "sku_only_db": len(set(db_sku) - set(sheet_sku)),
        "products_exact": len(exact),
        "products_partial": len(partial),
        "products_only_sheet": len(only_sheet),
        "products_only_db": len(only_db),
    }
    (args.out_dir / "master_vs_staging_summary.json").write_text(
        json.dumps(summary, indent=2) + "\n"
    )
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
