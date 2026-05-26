#!/usr/bin/env python3
"""
Update Sarveda RTM xlsx — cell values only (preserves formatting).
Usage: PYTHONPATH=backend/.rtm-pip python3 scripts/update-rtm-status.py
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
RTM_PATH = ROOT / "Sarveda-RTM-v1.1.xlsx"
FALLBACK = ROOT / "Sarveda-RTM.xlsx"

# REQ ID -> (Dev Status, Test Status, note append)
UPDATES: dict[str, tuple[str, str, str]] = {
    "REQ-CART-006": (
        "Complete",
        "Not Tested",
        "May 2026: POST/DELETE /api/cart/coupon + checkout applies discount.",
    ),
    "REQ-MKT-004": (
        "In Progress",
        "Not Tested",
        "May 2026: Coupon engine at checkout; admin coupon UI still pending.",
    ),
    "REQ-MKT-007": (
        "In Progress",
        "Not Tested",
        "May 2026: WELCOME10 works if imported via import:coupons.",
    ),
}

HEADER_ROW = 4
COL_REQ = 1
COL_DEV = 7
COL_TEST = 8
COL_NOTES = 11
COL_REVISED = 13


def find_full_rtm_sheet(wb: openpyxl.Workbook):
    for name in wb.sheetnames:
        if "Full RTM" in name or name.strip().endswith("Full RTM"):
            return wb[name]
    return wb[wb.sheetnames[1]]


def update_sheet(ws, updates: dict[str, tuple[str, str, str]]) -> int:
    changed = 0
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        req = ws.cell(row, COL_REQ).value
        if not req or not str(req).startswith("REQ-"):
            continue
        req_id = str(req).strip()
        if req_id not in updates:
            continue
        dev, test, note = updates[req_id]
        ws.cell(row, COL_DEV).value = dev
        ws.cell(row, COL_TEST).value = test
        existing = ws.cell(row, COL_NOTES).value
        prefix = f"{note} "
        merged = prefix + (str(existing) if existing else "")
        ws.cell(row, COL_NOTES).value = merged.strip()
        ws.cell(row, COL_REVISED).value = date.today().strftime("%d-%b-%Y")
        changed += 1
    return changed


def main() -> None:
    path = RTM_PATH if RTM_PATH.exists() else FALLBACK
    if not path.exists():
        raise SystemExit(f"RTM not found: {RTM_PATH}")

    wb = openpyxl.load_workbook(path)
    ws = find_full_rtm_sheet(wb)
    n = update_sheet(ws, UPDATES)
    wb.save(path)
    print(f"Updated {n} rows in {path.name} (formatting preserved)")


if __name__ == "__main__":
    main()
